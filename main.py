import asyncio
import os
import logging
import pytz 
import io
import aiohttp
import json
import csv
import calendar
import re
from datetime import datetime, timedelta, date as d_date, time as d_time
from collections import defaultdict
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart, Command
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from geopy.distance import geodesic
from aiohttp import web
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import asyncpg
import pickle
import traceback

# --- LOGGING ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

def escape_md(text: str) -> str:
    """Markdown uchun xavfli belgilarni escape qilish"""
    if not text:
        return ''
    for ch in ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']:
        text = text.replace(ch, f'\\{ch}')
    return text

# --- SOZLAMALAR ---
TOKEN = os.environ.get("BOT_TOKEN")
if not TOKEN:
    raise ValueError("BOT_TOKEN topilmadi! Render.com da environment variable qo'shing")
DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise ValueError("DATABASE_URL topilmadi! Render.com da environment variable qo'shing")
BASE_URL = os.environ.get("WEBHOOK_URL", "https://www.kita.uz")
ADMIN_GROUP_ID = -1003885800610 
UZB_TZ = pytz.timezone('Asia/Tashkent') 

# Bot obyektini yaratish
bot = Bot(token=TOKEN)

# Dispatcher obyektini yaratish
dp = Dispatcher()

# Foydalanuvchi ma'lumotlari (RAM da vaqtinchalik)
user_names = {}
user_specialty = {}
user_status = {}
user_languages = {}
user_ids = set()
daily_attendance_log = set()
attendance_counter = {}
schedules = {}
user_schedules = defaultdict(list)
broadcast_history = []

# RAM da guruhlar va o'quvchilar (tezroq ishlash uchun, lekin asosiy PostgreSQL'da)
groups = {}  # group_id -> group ma'lumotlari
group_students = defaultdict(list)  # group_id -> o'quvchilar ro'yxati
group_attendance_files = {}  # group_id -> kumulativ Excel (bytes) — har dars uchun yangi ustun

# BARCHA LOKATSIYALAR - DATABASE DAN YUKLANADI
LOCATIONS = []
SALARY_STRUCTURES = {}
BUILDINGS = {}
PENALTY_TYPES = {}
LESSON_TYPES = {}

ALLOWED_DISTANCE = 500
TAX_RATE = 7.5  # Soliq stavkasi (%)

# --- OYLIK HISOBOT UCHUN STATE ---
class MonthlyReport(StatesGroup):
    waiting_for_date_range = State()

# --- OYLIK KALKULYATOR UCHUN STATE ---
class SalaryCalc(StatesGroup):
    selecting_specialty = State()
    selecting_teacher = State()
    # Ko'p filialli hisob uchun
    entering_branch_data = State() 
    entering_students = State()
    entering_lessons = State()
    selecting_percentage = State()
    entering_penalty_it_percent = State()  # IT uchun % jarima
    entering_penalty_kr_sum = State()      # Koreys tili uchun so'mda jarima
    entering_payment = State()              # IT uchun jami tushum

# --- VIZUAL JADVAL UCHUN STATE ---
class VisualSchedule(StatesGroup):
    selecting_branch = State()

# --- GURUH YARATISH UCHUN STATE (YANGI) ---
class CreateGroup(StatesGroup):
    selecting_branch = State()
    selecting_type = State()
    selecting_teacher = State()
    selecting_days = State()
    entering_day_times = State()   # har kun uchun alohida vaqt
    entering_group_name = State()
    waiting_excel = State()

# --- O'QUVCHILAR DAVOMATI UCHUN STATE (YANGI) ---
class StudentAttendance(StatesGroup):
    selecting_students = State()
    late_students = State()

# --- POSTGRESQL DATABASE CLASS ---
class Database:
    def __init__(self, url):
        self.url = url
        self.pool = None
    
    async def create_pool(self):
        try:
            self.pool = await asyncpg.create_pool(self.url, command_timeout=60, statement_cache_size=0)
            logging.info("✅ PostgreSQL ga ulandik!")
            return True
        except Exception as e:
            logging.error(f"❌ PostgreSQL ga ulanishda xato: {e}")
            return False
    
    async def init_tables(self):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    user_id BIGINT PRIMARY KEY,
                    full_name TEXT,
                    specialty TEXT,
                    status TEXT DEFAULT 'active',
                    language TEXT DEFAULT 'uz',
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            
            # Business expenses table (Korean style)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS business_expenses (
                    id SERIAL PRIMARY KEY,
                    month TEXT NOT NULL,
                    expense_type TEXT NOT NULL,
                    amount INT DEFAULT 0,
                    note TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT NOW(),
                    UNIQUE(month, expense_type)
                )
            """)
            
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS applications (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    phone TEXT NOT NULL,
                    course TEXT DEFAULT '',
                    message TEXT DEFAULT '',
                    status TEXT DEFAULT 'new',
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS bootcamp_applications (
                    id SERIAL PRIMARY KEY,
                    fname TEXT NOT NULL,
                    lname TEXT NOT NULL,
                    phone TEXT NOT NULL,
                    dob TEXT DEFAULT '',
                    email TEXT DEFAULT '',
                    about TEXT DEFAULT '',
                    skills TEXT DEFAULT '',
                    track TEXT DEFAULT '',
                    resume_url TEXT DEFAULT '',
                    resume_name TEXT DEFAULT '',
                    status TEXT DEFAULT 'new',
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            try:
                await conn.execute("ALTER TABLE bootcamp_applications ADD COLUMN IF NOT EXISTS resume_url TEXT DEFAULT ''")
                await conn.execute("ALTER TABLE bootcamp_applications ADD COLUMN IF NOT EXISTS resume_name TEXT DEFAULT ''")
            except: pass
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS news (
                    id SERIAL PRIMARY KEY,
                    title TEXT NOT NULL,
                    body TEXT NOT NULL,
                    title_ru TEXT DEFAULT '',
                    body_ru TEXT DEFAULT '',
                    title_kr TEXT DEFAULT '',
                    body_kr TEXT DEFAULT '',
                    image_url TEXT DEFAULT '',
                    is_published BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            # Existing news table migration - add columns if not exist
            for col in [('title_ru','TEXT DEFAULT \'\''),('body_ru','TEXT DEFAULT \'\''),('title_kr','TEXT DEFAULT \'\''),('body_kr','TEXT DEFAULT \'\'')]:
                try:
                    await conn.execute(f"ALTER TABLE news ADD COLUMN IF NOT EXISTS {col[0]} {col[1]}")
                except:
                    pass
            # Groups table migration - add student_count column if not exist
            try:
                await conn.execute("ALTER TABLE groups ADD COLUMN IF NOT EXISTS student_count INTEGER DEFAULT 0")
            except:
                pass
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS site_config (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TIMESTAMP DEFAULT NOW()
                )
            """)
            
            # Salary configurations table (Korean style)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS salary_configs (
                    id SERIAL PRIMARY KEY,
                    category TEXT NOT NULL,
                    category_kr TEXT DEFAULT '',
                    bin_key TEXT NOT NULL,
                    bin_name TEXT DEFAULT '',
                    amount INT DEFAULT 0,
                    updated_at TIMESTAMP DEFAULT NOW(),
                    UNIQUE(category, bin_key)
                )
            """)
            
            # Branches table
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS branches (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    lat DOUBLE PRECISION DEFAULT 0,
                    lon DOUBLE PRECISION DEFAULT 0,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            
            # Configurations table (JSON data storage)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS configurations (
                    key TEXT PRIMARY KEY,
                    value JSONB NOT NULL,
                    updated_at TIMESTAMP DEFAULT NOW()
                )
            """)
            defaults = [
                ('hero_title', 'KITA구 HANCOM Academy'),
                ('hero_subtitle', 'IT texnologiyalari, Koreys tili va ofis koʻnikmalarini professional darajada oʻrganing.'),
                ('stat_graduates', '500+'),
                ('stat_directions', '3'),
                ('stat_years', '5+'),
                ('stat_groups', '22'),
                ('contact_phone', '+998508182280'),
                ('contact_address', "Shota Rustaveli ko'chasi 156"),
                ('contact_telegram', '@Hancom_A'),
                ('contact_instagram', 'hancomacademy.uz'),
                ('course1_title', 'IT Texnologiyalari'),
                ('course1_desc', 'Dasturlash, kompyuter savodxonligi, ofis dasturlari.'),
                ('course1_img', 'https://i.ibb.co/FkykT2bN/1.jpg'),
                ('course2_title', 'Koreys Tili'),
                ('course2_desc', 'Hangeul dan boshlab TOPIK imtihonigacha.'),
                ('course2_img', 'https://i.ibb.co/PzmMWXF1/2.jpg'),
                ('course3_title', 'Ofis Xizmati'),
                ('course3_desc', 'Ish yuritish, hujjatlashtirish, ofis dasturlari.'),
                ('course3_img', 'https://i.ibb.co/1f01hPHS/3.jpg'),
                ('gallery_img1', 'https://i.ibb.co/67gxBCYS/4.jpg'),
                ('gallery_img2', 'https://i.ibb.co/hJfd9kQt/5.jpg'),
                ('gallery_img3', 'https://i.ibb.co/hF4R5r9F/6.jpg'),
                ('gallery_img4', 'https://i.ibb.co/5XJ4nPKP/7.jpg'),
                ('gallery_img5', 'https://i.ibb.co/sdQHfZs5/8.jpg'),
                ('about_text', "KITA (Korea IT Academy) — Toshkentda joylashgan zamonaviy o'quv markazi."),
                # Russian defaults
                ('hero_title_ru', 'KITA구 HANCOM Academy'),
                ('hero_subtitle_ru', 'Изучайте IT, корейский язык и офисные навыки на профессиональном уровне.'),
                ('about_text_ru', 'KITA (Korea IT Academy) — современный учебный центр в Ташкенте.'),
                ('course1_title_ru', 'IT Технологии'),
                ('course1_desc_ru', 'Программирование, компьютерная грамотность, офисные программы.'),
                ('course2_title_ru', 'Корейский язык'),
                ('course2_desc_ru', 'От хангыля до экзамена TOPIK.'),
                ('course3_title_ru', 'Офисная служба'),
                ('course3_desc_ru', 'Делопроизводство, документация, офисные программы.'),
                # Korean defaults
                ('hero_title_kr', 'KITA구 HANCOM Academy'),
                ('hero_subtitle_kr', 'IT 기술, 한국어, 사무 능력을 전문적인 수준으로 배우세요.'),
                ('about_text_kr', 'KITA (Korea IT Academy) — 타슈켄트에 위치한 현대적인 교육 센터입니다.'),
                ('course1_title_kr', 'IT 기술'),
                ('course1_desc_kr', '프로그래밍, 컴퓨터 활용, 오피스 프로그램.'),
                ('course2_title_kr', '한국어'),
                ('course2_desc_kr', '한글부터 TOPIK 시험까지.'),
                ('course3_title_kr', '사무 서비스'),
                ('course3_desc_kr', '사무 관리, 문서화, 오피스 프로그램.'),
            ]
            for key, value in defaults:
                await conn.execute("""
                    INSERT INTO site_config (key, value)
                    VALUES ($1, $2)
                    ON CONFLICT (key) DO NOTHING
                """, key, value)
            
            # Partners table
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS partners (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    logo_url TEXT DEFAULT '',
                    website_url TEXT DEFAULT '',
                    sort_order INT DEFAULT 0,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            logging.info("✅ Jadvallar yaratildi!")
    
    async def save_user(self, user_id, full_name, specialty=None, language='uz'):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO users (user_id, full_name, specialty, language, status)
                VALUES ($1, $2, $3, $4, 'active')
                ON CONFLICT (user_id) 
                DO UPDATE SET 
                    full_name = EXCLUDED.full_name,
                    specialty = COALESCE(EXCLUDED.specialty, users.specialty),
                    language = EXCLUDED.language
            """, user_id, full_name, specialty, language)
    
    async def update_user_status(self, user_id, status):
        async with self.pool.acquire() as conn:
            await conn.execute("UPDATE users SET status = $1 WHERE user_id = $2", status, user_id)
    
    async def get_all_users(self):
        async with self.pool.acquire() as conn:
            return await conn.fetch("SELECT * FROM users")
    
    async def get_user(self, user_id):
        async with self.pool.acquire() as conn:
            return await conn.fetchrow("SELECT * FROM users WHERE user_id = $1", user_id)
    
    async def load_branches(self):
        """Load branches from database into LOCATIONS"""
        global LOCATIONS
        try:
            async with self.pool.acquire() as conn:
                rows = await conn.fetch("SELECT * FROM branches ORDER BY id")
            
            LOCATIONS = [{'name': r['name'], 'lat': r['lat'], 'lon': r['lon']} for r in rows]
            logging.info(f"✅ Filiallar yuklandi: {len(LOCATIONS)} ta")
        except Exception as e:
            logging.error(f"Branches load error: {e}")
            LOCATIONS = []
    
    async def load_configurations(self):
        """Load all configurations from database"""
        import json as _json
        global SALARY_STRUCTURES, BUILDINGS, PENALTY_TYPES, LESSON_TYPES
        try:
            async with self.pool.acquire() as conn:
                rows = await conn.fetch("SELECT key, value FROM configurations")
            
            configs = {}
            for row in rows:
                val = row['value']
                if isinstance(val, str):
                    val = _json.loads(val)
                configs[row['key']] = val
            
            if not configs:
                # Seed default configurations
                await self.seed_configurations()
                async with self.pool.acquire() as conn:
                    rows = await conn.fetch("SELECT key, value FROM configurations")
                for row in rows:
                    val = row['value']
                    if isinstance(val, str):
                        val = _json.loads(val)
                    configs[row['key']] = val
            
            # Load into globals
            if 'salary_structures' in configs:
                SALARY_STRUCTURES = configs['salary_structures']
            if 'buildings' in configs:
                BUILDINGS = configs['buildings']
            if 'penalty_types' in configs:
                PENALTY_TYPES = configs['penalty_types']
            if 'lesson_types' in configs:
                LESSON_TYPES = configs['lesson_types']
            
            logging.info(f"✅ Konfiguratsiyalar yuklandi: {len(configs)} ta")
        except Exception as e:
            logging.error(f"Configurations load error: {e}")
    
    async def seed_configurations(self):
        """Seed default configurations"""
        import json as _json
        
        default_configs = [
            ('salary_structures', {
                'soeup': {'name': '수습', 'salaries': {'bin_1': 7500000}},
                'sawon': {'name': '사원', 'salaries': {'bin_1': 8500000, 'bin_2': 9500000, 'bin_3': 10500000}},
                'daeri': {'name': '대리', 'salaries': {'bin_1': 11500000, 'bin_2': 13000000, 'bin_3': 14500000}},
                'gwallija': {'name': '관리자', 'salaries': {'bin_1': 16000000, 'bin_2': 17500000, 'bin_3': 19000000}}
            }),
            ('buildings', {'bin_1': '1호봉', 'bin_2': '2호봉', 'bin_3': '3호봉'}),
            ('penalty_types', {
                'jigak': {'name': '지각', 'percent': 1.0, 'amounts': {'soeup': 75000, 'sawon': 95000, 'daeri': 115000, 'gwallija': 160000}},
                'mudan_jigak': {'name': '무단지각', 'percent': 3.0, 'amounts': {'soeup': 225000, 'sawon': 285000, 'daeri': 345000, 'gwallija': 480000}},
                'jote': {'name': '조퇴', 'percent': 2.5, 'amounts': {'soeup': 187500, 'sawon': 237500, 'daeri': 287500, 'gwallija': 400000}},
                'gyeolgun': {'name': '결근', 'percent': 5.5, 'amounts': {'soeup': 412500, 'sawon': 522500, 'daeri': 632500, 'gwallija': 880000}},
                'mudan_gyeolgun': {'name': '무단결근', 'percent': 12.0, 'amounts': {'soeup': 900000, 'sawon': 1140000, 'daeri': 1380000, 'gwallija': 1920000}},
                'ilil_eopmu_bogeo': {'name': '일일업무보고', 'percent': 1.0, 'amounts': {'soeup': 75000, 'sawon': 95000, 'daeri': 115000, 'gwallija': 160000}},
                'geojit_bogeo': {'name': '거짓보고', 'percent': 2.0, 'amounts': {'soeup': 150000, 'sawon': 190000, 'daeri': 230000, 'gwallija': 320000}},
                'mudan_ital': {'name': '무단이탈', 'percent': 5.0, 'amounts': {'soeup': 375000, 'sawon': 475000, 'daeri': 575000, 'gwallija': 800000}},
                'jisi_bulihhaeng': {'name': '지시불이행', 'percent': 8.0, 'amounts': {'soeup': 600000, 'sawon': 760000, 'daeri': 920000, 'gwallija': 1280000}},
                'mi_bogeo': {'name': '미보고', 'percent': 2.0, 'amounts': {'soeup': 150000, 'sawon': 190000, 'daeri': 230000, 'gwallija': 320000}}
            }),
            ('lesson_types', {'Koreys tili': 'korean', 'IT': 'it', 'Ofis xodimi': 'office'})
        ]
        
        async with self.pool.acquire() as conn:
            for key, value in default_configs:
                await conn.execute("""
                    INSERT INTO configurations (key, value) VALUES ($1, $2)
                    ON CONFLICT (key) DO NOTHING
                """, key, _json.dumps(value))
    
    async def save_attendance(self, user_id, branch, att_date, att_time):
        try:
            async with self.pool.acquire() as conn:
                from datetime import datetime, time
                date_obj = datetime.strptime(att_date, "%Y-%m-%d").date()
                time_parts = att_time.split(':')
                time_obj = time(int(time_parts[0]), int(time_parts[1]), int(time_parts[2]))
                
                await conn.execute("""
                    INSERT INTO attendance (user_id, branch, date, time)
                    VALUES ($1, $2, $3, $4)
                    ON CONFLICT (user_id, branch, date) DO NOTHING
                """, user_id, branch, date_obj, time_obj)
                logging.info(f"✅ Davomat saqlandi: user={user_id}, branch={branch}")
        except Exception as e:
            logging.error(f"❌ Davomat saqlashda xato: {e}")
    
    async def get_user_attendance(self, user_id):
        async with self.pool.acquire() as conn:
            return await conn.fetch("""
                SELECT * FROM attendance 
                WHERE user_id = $1 
                ORDER BY date DESC, time DESC
            """, user_id)
    
    async def get_attendance_by_date(self, att_date):
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT a.*, u.full_name, u.specialty 
                FROM attendance a
                JOIN users u ON a.user_id = u.user_id
                WHERE a.date = $1
                ORDER BY a.time
            """, att_date)
            return rows
    
    async def get_all_attendance(self):
        async with self.pool.acquire() as conn:
            return await conn.fetch("SELECT * FROM attendance")
    
    async def save_schedule(self, schedule_id, user_id, branch, lesson_type, days_dict):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO schedules (schedule_id, user_id, branch, lesson_type, days_data)
                VALUES ($1, $2, $3, $4, $5::jsonb)
            """, schedule_id, user_id, branch, lesson_type, json.dumps(days_dict))
    
    async def get_user_schedules(self, user_id):
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM schedules WHERE user_id = $1", user_id)
            result = []
            for row in rows:
                data = dict(row)
                data['days'] = json.loads(data['days_data'])
                result.append(data)
            return result
    
    async def get_all_schedules(self):
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM schedules")
            result = []
            for row in rows:
                data = dict(row)
                data['days'] = json.loads(data['days_data'])
                result.append(data)
            return result
    
    async def delete_schedule(self, schedule_id):
        async with self.pool.acquire() as conn:
            await conn.execute("DELETE FROM schedules WHERE schedule_id = $1", schedule_id)
    
    async def update_schedule(self, schedule_id, branch, lesson_type, days_dict):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                UPDATE schedules 
                SET branch = $1, lesson_type = $2, days_data = $3::jsonb
                WHERE schedule_id = $4
            """, branch, lesson_type, json.dumps(days_dict), schedule_id)
    
    async def save_broadcast(self, message_text, sent_count, failed_count, specialty):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO broadcast_history (message_text, sent_count, failed_count, specialty)
                VALUES ($1, $2, $3, $4)
            """, message_text, sent_count, failed_count, specialty)
    
    async def load_to_ram(self):
        """Ma'lumotlarni bazadan yangilash"""
        global user_names, user_specialty, user_status, user_languages, user_ids
        global daily_attendance_log, attendance_counter, schedules, user_schedules
        global groups, group_students
        
        # Global o'zgaruvchilarni tozalash
        user_names.clear(); user_specialty.clear(); user_status.clear()
        user_languages.clear(); user_ids.clear()
        daily_attendance_log.clear(); attendance_counter.clear()
        schedules.clear(); user_schedules.clear()
        groups.clear(); group_students.clear()
        
        users = await self.get_all_users()
        for u in users:
            user_ids.add(u['user_id'])
            user_names[u['user_id']] = u['full_name']
            user_specialty[u['user_id']] = u['specialty']
            user_status[u['user_id']] = u['status']
            user_languages[u['user_id']] = u['language']
        
        attendances = await self.get_all_attendance()
        for r in attendances:
            daily_attendance_log.add((
                r['user_id'],
                r['branch'],
                r['date'].isoformat(),
                r['time'].strftime("%H:%M:%S")
            ))
            month = r['date'].strftime("%Y-%m")
            key = (r['user_id'], r['branch'], month)
            attendance_counter[key] = attendance_counter.get(key, 0) + 1
        
        all_schedules = await self.get_all_schedules()
        for r in all_schedules:
            schedules[r['schedule_id']] = {
                'user_id': r['user_id'],
                'branch': r['branch'],
                'lesson_type': r['lesson_type'],
                'days': r['days']
            }
            user_schedules[r['user_id']].append(r['schedule_id'])
        
        # Guruh va o'quvchilarni yuklash
        async with self.pool.acquire() as conn:
            all_groups = await conn.fetch("SELECT * FROM groups")
            for g in all_groups:
                raw = json.loads(g['days_data'])
                # Yangi format: {kun: vaqt} dict
                if isinstance(raw, dict):
                    day_times = raw
                    days_list = list(raw.keys())
                    first_time = list(raw.values())[0] if raw else ''
                else:
                    # Eski list formatni day_times ga aylantiramiz
                    t = g['time_text'] or ''
                    day_times = {d: t for d in raw}
                    days_list = raw
                    first_time = t
                groups[g['id']] = {
                    'group_name': g['group_name'],
                    'branch': g['branch'],
                    'lesson_type': g['lesson_type'],
                    'teacher_id': g['teacher_id'],
                    'days': days_list,
                    'day_times': day_times,
                    'time': first_time,
                    'time_text': first_time,
                    'created_at': g['created_at'],
                    'student_count': g.get('student_count') or len(group_students.get(g['id'], []))
                }
                
                students = await conn.fetch("SELECT * FROM group_students WHERE group_id = $1", g['id'])
                group_students[g['id']] = [{'name': s['student_name'], 'phone': s['student_phone']} for s in students]

                # Excel faylni DB dan yuklash
                excel_row = await conn.fetchrow(
                    "SELECT file_data FROM group_excel_files WHERE group_id=$1", g['id']
                )
                if excel_row and excel_row['file_data']:
                    group_attendance_files[g['id']] = bytes(excel_row['file_data'])
        
        logging.info(f"✅ RAM ga yangilandi: {len(user_ids)} foydalanuvchi, {len(daily_attendance_log)} davomat, {len(groups)} guruh")


db = Database(DATABASE_URL)

class Registration(StatesGroup):
    waiting_for_name = State()
    waiting_for_specialty = State()

class Broadcast(StatesGroup):
    selecting_specialty = State()
    waiting_for_message = State()
    waiting_for_confirm = State()

class AddLocation(StatesGroup):
    waiting_for_name = State()
    waiting_for_coords = State()

class PDFReport(StatesGroup):
    waiting_for_date = State()

class TeacherAddStudent(StatesGroup):
    entering_name  = State()
    entering_phone = State()

class ProfileEdit(StatesGroup):
    waiting_for_new_name = State()

class AdminPDFReport(StatesGroup):
    waiting_for_report_type = State()

class EditGroupStudents(StatesGroup):
    entering_new_name = State()
    entering_new_phone = State()

class ExcelUploadGroup(StatesGroup):
    waiting_file = State()

class ExcelCreateGroup(StatesGroup):
    waiting_file = State()

class EditGroupSchedule(StatesGroup):
    selecting_days = State()
    entering_day_times = State()
    entering_time = State()

class EditGroupTeacher(StatesGroup):
    selecting_teacher = State()

WEEKDAYS = {
    'uz':['Dushanba', 'Seshanba', 'Chorshanba', 'Payshanba', 'Juma', 'Shanba', 'Yakshanba'],
}

WEEKDAY_ORDER = {
    'Dushanba': 0, 'Seshanba': 1, 'Chorshanba': 2, 'Payshanba': 3, 'Juma': 4, 'Shanba': 5, 'Yakshanba': 6
}

LESSON_TYPES = {
    'uz':['IT', 'Koreys tili'],
}

user_photo_cache = {}  # uid -> Telegram photo URL
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "7117")

def _check_admin_request(request) -> bool:
    # Cookie ni tekshirish (asosiy usul)
    password = request.cookies.get('admin_token', '')
    if password == ADMIN_PASSWORD:
        return True
    
    # URL parametrini tekshirish (fallback)
    password_query = request.rel_url.query.get('p', '')
    return password_query == ADMIN_PASSWORD

WEEKDAYS_UZ =["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]

TRANSLATIONS = {
    'uz': {
        'welcome': "🌟 HANCOM ACADEMYning o'qituvchilar uchun davomat botiga hush kelibsiz, {name}!",
        'ask_name': "👤 Iltimos, ism va familiyangizni kiriting:\n\nMasalan: Ali Karimov",
        'ask_specialty': "📚 Qaysi fan o'qituvchisisiz?",
        'specialty_it': "💻 IT",
        'specialty_korean': "🇰🇷 Koreys tili",
        'specialty_office': "🏢 Ofis xodimi",
        'stats': "📊 Sizning statistikangiz:",
        'no_stats': "💭 Hali davomat qilmagansiz",
        'branches': "🏢 Mavjud filiallar (lokatsiya):",
        'help': "🤖 Botdan foydalanish qo'llanmasi:\n\n📍 Davomat qilish uchun:\n• Pastdagi \"📍 Kelganimni tasdiqlash\" tugmasini bosing\n• Joylashuvingizni yuboring\n\n📊 Statistika:\n• \"📊 Mening statistikam\" - shaxsiy davomat tarixingiz\n• \"🏢 Filiallar\" - barcha mavjud filiallar ro'yxati\n\n⚠️ Eslatmalar:\n• Har bir filialda kuniga faqat 1 marta davomat qilish mumkin\n• Davomat faqat Toshkent vaqti bilan hisoblanadi",
        'attendance_success': "✅ Davomat tasdiqlandi!\n\n🏫 Filial: {branch}\n📅 Sana: {date}\n⏰ Vaqt: {time}\n📊 Bu oydagi tashriflar: {count} marta\n📍 Masofa: {distance:.1f} metr",
        'already_attended': "⚠️ Siz bugun {branch} hududida allaqachon davomatdan o'tgansiz!",
        'not_in_area': "❌ Siz belgilangan ta'lim muassasalari hududida emassiz!",
        'daily_reminder': "⏰ Eslatma! Bugun hali davomat qilmagansiz. Ish kuningizni boshlash uchun davomatni tasdiqlang!",
        'weekly_top': "🏆 Haftaning eng faol o'qituvchilari:\n\n{top_list}",
        'monthly_report': "📊 {month} oyi uchun hisobot\n\n{report}",
        'language_changed': "✅ Til o'zgartirildi: O'zbek tili",
        'language_prompt': "Iltimos, tilni tanlang:",
        'view_schedules': "📋 Dars jadvalim (PDF)",
        'my_groups': "👥 Guruhlarim",
        'my_schedule': "📅 Sizning dars jadvalingiz PDF formatida tayyorlandi!",
        'no_schedules': "📭 Sizga hali dars jadvali biriktirilmagan.",
        'schedule_updated': "📢 Sizning dars jadvalingiz yangilandi!",
        'schedule_deleted_notify': "📢 Sizning dars jadvalingiz o'chirildi.",
        'reminder': "⏰ Eslatma!\n\nBugun soat {time} da {branch} filialida darsingiz bor.\nDavomat qilishni unutmang!",
        'lesson_started_attended': "✅ Dars boshlandi va siz muvaffaqiyatli davomatni amalga oshirdingiz!\n\nE'tiboringizni darsga qaratishingiz mumkin.\nDarsga kelgan o'quvchilarni davomat qilishni yodingizdan chiqarmang.\n\nHayrli kun!",
        'lesson_started_not_attended': "⚠️ Sizning darsingiz boshlandi, lekin hali davomat qilmadingiz!\n\n📌 {branch} filialida soat {time} da darsingiz boshlangan.\n📍 Iltimos, darhol davomat qiling yoki sababini admin xabardor qiling.\n\nDavomat qilish uchun 📍 Kelganimni tasdiqlash tugmasini bosing.",
        'select_teacher': "👤 O'qituvchini tanlang:",
        'select_lesson_type': "📚 Dars turini tanlang:",
        'active_schedules': "📋 Faol dars jadvallari",
        'no_active_schedules': "📭 Hali dars jadvallari mavjud emas.",
        'schedule_info': "{teacher} [{specialty}]\n🏢 {branch}\n📚 {lesson_type}\n{days_times}",
        'enter_date': "📅 Hisobot olish uchun sanani kiriting (format: YYYY-MM-DD)\nMasalan: 2026-03-01",
        'invalid_date': "❌ Noto'g'ri sana formati. Qaytadan urinib ko'ring:",
        'select_broadcast_specialty': "📢 Qaysi fan o'qituvchilariga xabar yubormoqchisiz?",
        'all_teachers': "👥 Hammasi",
        'edit_schedule': "✏️ Dars jadvalini tahrirlash",
        'select_new_branch': "🏢 Yangi filialni tanlang:",
        'select_new_lesson_type': "📚 Yangi dars turini tanlang:",
        'select_new_weekdays': "📅 Yangi kunlarni tanlang:",
        'enter_new_time': "⏰ {weekday} kuni uchun yangi vaqtni kiriting:\n\nFormat: HH:MM (masalan: 09:00)",
        'ontime': "Vaqtida",
        'late': "Kechikkan",
        'my_profile': "👤 Mening profilim",
        'profile_info': "👤 Sizning profilingiz:\n\nIsm: {name}\nMutaxassislik: {specialty}\nTil: {lang}",
        'edit_name': "✏️ Ismni o'zgartirish",
        'edit_my_specialty': "📚 Faoliyat turini o'zgartirish",
        'enter_new_name': "Yangi ism va familiyangizni kiriting:",
        'name_updated': "✅ Ismingiz muvaffaqiyatli yangilandi!",
        'back_to_menu': "🔙 Menyuga qaytish",
        'select_new_spec': "Yangi faoliyat turini tanlang:",
        'spec_updated': "✅ Mutaxassislik yangilandi!",
        'back_btn': "🔙 Ortga",
        'pdf_title': "Dars Jadvali",
        'pdf_headers': ['Kun', 'Vaqt'],
        'pdf_created': "Yaratilgan sana",
        'group_students_title': "O'quvchilar davomati",
        'group_students_submit': "📤 Davomatni yuborish",
        'group_students_sent': "✅ O'quvchilar davomati yuborildi!",
        'blocked_user': "❌ Siz bloklangansiz. Admin bilan bog'laning.",
        'ask_language': "🌐 Tilni tanlang / Выберите язык:",
        'buttons': {
            'attendance': "📍 Kelganimni tasdiqlash",
            'my_stats': "📊 Mening statistikam",
            'branches': "🏢 Filiallar",
            'top_week': "🏆 Hafta topi",
            'view_schedules': "📋 Dars jadvalim (PDF)",
            'my_groups': "👥 Guruhlarim",
            'help': "❓ Yordam",
            'language': "🌐 Til",
            'profile': "👤 Mening profilim",
            'schedule': "📅 Dars jadvalim",
            'stats': "📊 Mening statistikam"
        }
    },
    'ru': {
        'welcome': "🌟 Добро пожаловать в бот HANCOM ACADEMY для регистрации преподавателей, {name}!",
        'ask_name': "👤 Пожалуйста, введите ваше имя и фамилию:\n\nНапример: Али Каримов",
        'ask_specialty': "📚 Какой предмет вы преподаете?",
        'specialty_it': "💻 IT",
        'specialty_korean': "🇰🇷 Корейский язык",
        'specialty_office': "🏢 Офисный сотрудник",
        'stats': "📊 Ваша статистика:",
        'no_stats': "💭 Вы еще не отмечали присутствие",
        'branches': "🏢 Доступные филиалы (локации):",
        'help': "🤖 Инструкция по использованию бота:\n\n📍 Для отметки присутствия:\n• Нажмите кнопку \"📍 Подтвердить приход\"\n• Отправьте ваше местоположение\n\n📊 Статистика:\n• \"📊 Моя статистика\" - ваша личная история посещений\n• \"🏢 Филиалы\" - список всех филиалов\n\n⚠️ Примечания:\n• В каждом филиале можно отмечаться только 1 раз в день\n• Отметка рассчитывается по времени Ташкента",
        'attendance_success': "✅ Присутствие подтверждено!\n\n🏫 Филиал: {branch}\n📅 Дата: {date}\n⏰ Время: {time}\n📊 Посещений в этом месяце: {count} раз\n📍 Расстояние: {distance:.1f} метров",
        'already_attended': "⚠️ Вы уже отметились в {branch} сегодня!",
        'not_in_area': "❌ Вы находитесь вне территории образовательных учреждений!",
        'daily_reminder': "⏰ Напоминание! Вы еще не отметились сегодня. Подтвердите присутствие для начала рабочего дня!",
        'weekly_top': "🏆 Самые активные преподаватели недели:\n\n{top_list}",
        'monthly_report': "📊 Отчет за {month}\n\n{report}",
        'language_changed': "✅ Язык изменен: Русский",
        'language_prompt': "Пожалуйста, выберите язык:",
        'view_schedules': "📋 Мое расписание (PDF)",
        'my_groups': "👥 Мои группы",
        'my_schedule': "📅 Ваше расписание в формате PDF готово!",
        'no_schedules': "📭 Вам еще не назначено расписание.",
        'schedule_updated': "📢 Ваше расписание обновлено!",
        'schedule_deleted_notify': "📢 Ваше расписание удалено.",
        'reminder': "⏰ Напоминание!\n\nСегодня в {time} у вас занятие в филиале {branch}.\nНе забудьте отметиться!",
        'lesson_started_attended': "✅ Занятие началось и вы успешно отметились!\n\nМожете сосредоточиться на уроке.\nНе забудьте отметить присутствие пришедших учеников.\n\nХорошего дня!",
        'lesson_started_not_attended': "⚠️ Ваше занятие началось, но вы еще не отметились!\n\n📌 Занятие в филиале {branch} началось в {time}.\n📍 Пожалуйста, немедленно отметьтесь или сообщите причину администратору.\n\nДля отметки нажмите кнопку 📍 Подтвердить приход.",
        'select_teacher': "👤 Выберите преподавателя:",
        'select_lesson_type': "📚 Выберите тип занятия:",
        'active_schedules': "📋 Активное расписание",
        'no_active_schedules': "📭 Активное расписание отсутствует.",
        'schedule_info': "{teacher} [{specialty}]\n🏢 {branch}\n📚 {lesson_type}\n{days_times}",
        'enter_date': "📅 Введите дату для отчета (формат: YYYY-MM-DD)\nНапример: 2026-03-01",
        'invalid_date': "❌ Неверный формат даты. Попробуйте снова:",
        'select_broadcast_specialty': "📢 Какой группе преподавателей отправить сообщение?",
        'all_teachers': "👥 Все",
        'edit_schedule': "✏️ Изменить расписание",
        'select_new_branch': "🏢 Выберите новый филиал:",
        'select_new_lesson_type': "📚 Выберите новый тип занятия:",
        'select_new_weekdays': "📅 Выберите новые дни:",
        'enter_new_time': "⏰ Введите новое время для {weekday}:\n\nФормат: HH:MM (например: 09:00)",
        'ontime': "Вовремя",
        'late': "Опоздал",
        'my_profile': "👤 Мой профиль",
        'profile_info': "👤 Ваш профиль:\n\nИмя: {name}\nСпециальность: {specialty}\nЯзык: {lang}",
        'edit_name': "✏️ Изменить имя",
        'edit_my_specialty': "📚 Изменить специальность",
        'enter_new_name': "Введите новое имя и фамилию:",
        'name_updated': "✅ Ваше имя успешно обновлено!",
        'back_to_menu': "🔙 Вернуться в меню",
        'select_new_spec': "Выберите новую специальность:",
        'spec_updated': "✅ Специальность обновлена!",
        'back_btn': "🔙 Назад",
        'pdf_title': "Расписание занятий",
        'pdf_headers': ['День', 'Время'],
        'pdf_created': "Дата создания",
        'group_students_title': "Посещаемость учеников",
        'group_students_submit': "📤 Отправить данные",
        'group_students_sent': "✅ Данные о посещаемости отправлены!",
        'blocked_user': "❌ Вы заблокированы. Свяжитесь с администратором.",
        'ask_language': "🌐 Выберите язык:",
        'buttons': {
            'attendance': "📍 Подтвердить приход",
            'my_stats': "📊 Моя статистика",
            'branches': "🏢 Филиалы",
            'top_week': "🏆 Топ недели",
            'view_schedules': "📋 Мое расписание (PDF)",
            'my_groups': "👥 Мои группы",
            'help': "❓ Помощь",
            'language': "🌐 Язык",
            'profile': "👤 Мой профиль",
            'schedule': "📅 Мое расписание",
            'stats': "📊 Моя статистика"
        }
    }
}

def get_text(user_id: int, key: str, **kwargs):
    lang = user_languages.get(user_id, 'uz')
    if lang not in TRANSLATIONS:
        lang = 'uz'
        user_languages[user_id] = 'uz'
    text = TRANSLATIONS[lang].get(key, '')
    if kwargs:
        try:
            text = text.format(**kwargs)
        except:
            pass
    return text

def get_button_text(user_id: int, button_key: str):
    lang = user_languages.get(user_id, 'uz')
    if lang not in TRANSLATIONS:
        lang = 'uz'
    return TRANSLATIONS[lang].get('buttons', {}).get(button_key, TRANSLATIONS['uz'].get('buttons', {}).get(button_key, button_key))

# Sizning shaxsiy Telegram ID ingiz (botga /start bosib, /myid yoki @userinfobot dan bilib oling)
ADMIN_USER_IDS = set()  # Keyin qo'shiladi

def check_admin(chat_id):
    return chat_id == ADMIN_GROUP_ID or chat_id in ADMIN_USER_IDS

def get_specialty_display(specialty: str, lang: str = 'uz') -> str:
    if specialty == 'IT':
        return "💻 IT"
    elif specialty == 'Koreys tili':
        return "🇰🇷 Koreys tili"
    else:
        return "🏢 Ofis xodimi"

def sort_weekdays(days_dict):
    order = {'Dushanba': 0, 'Seshanba': 1, 'Chorshanba': 2, 'Payshanba': 3, 'Juma': 4, 'Shanba': 5, 'Yakshanba': 6}
    return dict(sorted(days_dict.items(), key=lambda x: order.get(x[0], 7)))

def calculate_lateness(attendance_time: str, lesson_time: str) -> tuple:
    try:
        if not attendance_time or not lesson_time or '—' in str(lesson_time) or '—' in str(attendance_time):
            return True, 0
        att_parts = list(map(int, str(attendance_time).split(':')))
        les_parts = list(map(int, lesson_time.split(':')))
        
        att_seconds = att_parts[0] * 3600 + att_parts[1] * 60 + att_parts[2]
        les_seconds = les_parts[0] * 3600 + les_parts[1] * 60
        
        diff = att_seconds - les_seconds
        
        if diff <= 60:
            return True, 0
        else:
            return False, int(diff / 60)
    except Exception as e:
        logging.error(f"calculate_lateness error: {e}")
        return True, 0

def get_kr_exam_penalty(perc: int) -> int:
    """Koreys tili imtixon foiziga qarab jarima qaytaradi"""
    if perc < 10: return 900000
    elif perc < 20: return 800000
    elif perc < 30: return 700000
    elif perc < 40: return 600000
    elif perc < 50: return 500000
    elif perc < 60: return 400000
    elif perc < 70: return 300000
    elif perc < 80: return 200000
    elif perc < 90: return 100000
    else: return 0

async def get_combined_report_pdf(report_date: d_date) -> io.BytesIO:
    """Kunlik davomat hisoboti — Excel format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    report_date_str = report_date.strftime("%Y-%m-%d")
    report_weekday = WEEKDAYS_UZ[report_date.weekday()]

    check_ins = [list(att) for att in daily_attendance_log if att[2] == report_date_str]

    for s_id, s_data in schedules.items():
        uid = s_data['user_id']
        branch = s_data['branch']
        if report_weekday in s_data['days']:
            already_noted = any(c[0] == uid and c[1] == branch for c in check_ins)
            if not already_noted:
                check_ins.append([uid, branch, report_date_str, "00:00:00", "ABSENT"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Davomat"

    # Sarlavha
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Davomat Hisoboti: {report_date.strftime('%d.%m.%Y')} ({report_weekday})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['№', 'Davomat vaqti', 'Dars vaqti', "O'qituvchi", 'Mutaxassislik', 'Filial', 'Holat', 'Kechikish']
    header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Ma'lumotlar
    for i, att in enumerate(sorted(check_ins, key=lambda x: x[3] if x[3] != "00:00:00" else "23:59:59"), 1):
        uid = att[0]
        branch = att[1]
        att_time = att[3]
        is_absent = len(att) > 4 and att[4] == "ABSENT"
        teacher_name = user_names.get(uid, "Noma'lum")
        specialty = user_specialty.get(uid, "")
        lesson_time = "—"
        for gdata in groups.values():
            if gdata.get('teacher_id') == uid and gdata.get('branch') == branch:
                if report_weekday in gdata.get('days', []):
                    lesson_time = gdata.get('time_text') or gdata.get('time', '—')
                    break
        if is_absent or att_time == "00:00:00":
            status = "KELMAGAN"
            late_text = "—"
            att_time_disp = "—"
        else:
            ontime, mins = calculate_lateness(att_time, lesson_time)
            status = "Vaqtida" if ontime else "Kechikkan"
            late_text = "0" if ontime else f"{mins} min"
            att_time_disp = att_time

        row = [str(i), att_time_disp, lesson_time, teacher_name, specialty, branch, status, late_text]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+2, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            if status == "KELMAGAN":
                cell.fill = PatternFill(start_color='EF5350', end_color='EF5350', fill_type='solid')
                cell.font = Font(color='FFFFFF')
            elif status == "Kechikkan" and col == 7:
                cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                cell.font = Font(color='CC0000')
            elif status == "Vaqtida" and col == 7:
                cell.font = Font(color='007700')

    # Ustun kengliklari
    col_widths = [5, 14, 12, 22, 15, 18, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def main_keyboard(user_id: int):
    builder = ReplyKeyboardBuilder()
    lang = user_languages.get(user_id, 'uz')
    if lang == 'ru':
        app_btn_text = '📱 HANCOM Teacher'
        lang_btn_text = '\U0001F310 Язык'
    else:
        app_btn_text = '📱 HANCOM Teacher'
        lang_btn_text = '\U0001F310 Til'
    webapp_url = f"{BASE_URL}/teacher?user_id={user_id}&section=stats"
    buttons = [
        KeyboardButton(text=get_button_text(user_id, 'attendance'), request_location=True),
        KeyboardButton(
            text=app_btn_text,
            web_app=types.WebAppInfo(url=webapp_url)
        ),
        KeyboardButton(text=lang_btn_text),
    ]
    builder.add(*buttons)
    builder.adjust(2, 1)
    return builder.as_markup(resize_keyboard=True)


async def language_selection_keyboard():
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="🇺🇿 O'zbekcha", callback_data="lang_uz"),
        InlineKeyboardButton(text="🇷🇺 Русский", callback_data="lang_ru"),
    )
    return builder.as_markup()

async def specialty_keyboard(user_id: int):
    lang = user_languages.get(user_id, 'uz')
    builder = ReplyKeyboardBuilder()
    builder.add(
        KeyboardButton(text=TRANSLATIONS[lang]['specialty_it']),
        KeyboardButton(text=TRANSLATIONS[lang]['specialty_korean']),
        KeyboardButton(text=TRANSLATIONS[lang]['specialty_office'])
    )
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True)

def get_yandex_maps_link(lat: float, lon: float) -> str:
    return f"https://yandex.com/maps/?pt={lon},{lat}&z=17&l=map"

# --- STANDARTLASHTIRILGAN PDF FUNKSIYASI (XUDDI ADMIN PANELIDAGIDEK) ---
async def create_schedule_pdf(user_id: int) -> io.BytesIO:
    """O'qituvchining shaxsiy dars jadvali — Excel format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Dars Jadvali"

    name = user_names.get(user_id, "O'qituvchi")
    specialty_raw = user_specialty.get(user_id, '')
    spec_display = get_specialty_display(specialty_raw)

    # Sarlavha
    ws.merge_cells('A1:C1')
    ws['A1'] = "HANCOM ACADEMY — O'QITUVCHI DARS JADVALI"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = f"{name} | {spec_display}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    teacher_groups = [(gid, gd) for gid, gd in groups.items() if gd['teacher_id'] == user_id]

    if not teacher_groups:
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "Sizga hali guruh biriktirilmagan."
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
    else:
        for gid, gdata in teacher_groups:
            branch = gdata.get('branch', '')
            group_name = gdata.get('group_name', '')
            lesson_type = gdata.get('lesson_type', '')
            days_list = gdata.get('days', [])
            day_times_g = gdata.get('day_times', {})
            time_text = gdata.get('time_text') or gdata.get('time', '')

            # Guruh sarlavhasi
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = f"{group_name} | {branch} | {lesson_type}"
            ws[f'A{row}'].font = Font(bold=True, size=12, color='C62828')
            ws[f'A{row}'].alignment = Alignment(horizontal='left')
            row += 1

            # Header
            for col, h in enumerate(['Hafta kuni', 'Dars vaqti', 'Filial'], 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            row += 1

            for day in days_list:
                vaqt = day_times_g.get(day, time_text)
                for col, val in enumerate([day, vaqt, branch], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                row += 1
            row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

async def get_student_attendance_kb(group_id, selected_indices):
    """Guruh o'quvchilarini tanlash uchun klaviatura yaratish"""
    builder = InlineKeyboardBuilder()
    students = group_students.get(group_id, [])
    lang = 'uz'  # Default til
    
    for i, std in enumerate(students):
        status = "✅ " if i in selected_indices else "⬜ "
        builder.row(InlineKeyboardButton(
            text=f"{status}{std['name']}", 
            callback_data=f"std_check_{i}"
        ))
    
    builder.row(InlineKeyboardButton(
        text=TRANSLATIONS[lang]['group_students_submit'], 
        callback_data="std_submit"
    ))
    return builder.as_markup()

async def handle(request):
    """Asosiy sahifa — KITA landing page"""
    import os
    html_path = os.path.join(os.path.dirname(__file__), 'index.html')
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            html = f.read()
        return web.Response(text=html, content_type='text/html', charset='utf-8')
    except FileNotFoundError:
        now_uzb = datetime.now(UZB_TZ)
        return web.Response(
            text=f"Bot is running! ✅\n\n"
                 f"📅 Sana: {now_uzb.strftime('%Y-%m-%d')}\n"
                 f"⏰ Vaqt: {now_uzb.strftime('%H:%M:%S')}\n"
                 f"👥 Foydalanuvchilar: {len(user_ids)} ta\n"
                 f"📊 Bugungi davomatlar: {len([k for k in daily_attendance_log if k[2] == now_uzb.strftime('%Y-%m-%d')])} ta\n"
                 f"👥 Guruhlar: {len(groups)} ta"
        )

async def admin_panel_page(request):
    """Admin panel"""
    import os
    # Parol to'g'ri bo'lsa panelni ko'rsat
    if _check_admin_request(request):
        html_path = os.path.join(os.path.dirname(__file__), 'admin_panel.html')
        try:
            with open(html_path, 'r', encoding='utf-8') as f:
                html = f.read()
            return web.Response(text=html, content_type='text/html', charset='utf-8')
        except FileNotFoundError:
            return web.Response(text='Admin panel topilmadi', status=404)
    # Login sahifasi - alohida string qurish
    parts = [
        '<!DOCTYPE html><html><head>',
        '<meta charset="UTF-8">',
        '<meta name="viewport" content="width=device-width,initial-scale=1">',
        '<title>KITA Admin</title><link rel="icon" type="image/svg+xml" href="/static/white_logo.svg">',
        '<style>',
        '*{margin:0;padding:0;box-sizing:border-box}',
        'body{background:#0d1117;color:#e6edf3;font-family:Inter,sans-serif;',
        'min-height:100vh;display:flex;align-items:center;justify-content:center}',
        '.box{background:#161b22;border:1px solid #30363d;border-radius:16px;',
        'padding:40px 32px;width:100%;max-width:340px;text-align:center}',
        '.logo{margin-bottom:16px;display:flex;justify-content:center}',
        'input{width:100%;background:#0d1117;border:1px solid #30363d;border-radius:10px;',
        'padding:14px;color:#e6edf3;font-size:22px;letter-spacing:8px;text-align:center;outline:none}',
        'input:focus{border-color:#58a6ff}',
        'button{width:100%;background:#238636;border:none;border-radius:10px;',
        'padding:14px;color:#fff;font-size:15px;font-weight:700;cursor:pointer;margin-top:12px}',
        '.err{color:#f87171;font-size:12px;margin-top:10px;min-height:18px}',
        '</style></head><body>',
        '<div class="box">',
        '<div class="logo"><img src="/static/white_logo.svg" style="width:180px;height:auto" alt="KITA"></div>',
        '<input type="password" id="pw" placeholder="****" maxlength="20"',
        ' onkeydown="if(event.key===\'Enter\')login()">',
        '<button onclick="login()">Kirish &#8594;</button>',
        '<div class="err" id="err"></div>',
        '</div>',
        '<script>',
        'async function login(){',
        '  var pw=document.getElementById("pw").value;',
        '  if(!pw){document.getElementById("err").textContent="Parol kiriting";return;}',
        '  try{',
        '    var r=await fetch("/admin/login",{method:"POST",',
        '      headers:{"Content-Type":"application/json"},',
        '      body:JSON.stringify({password:pw})});',
        '    var d=await r.json();',
        '    if(d.ok){',
        '      location.href="/admin";',
        '    } else {',
        '      document.getElementById("err").textContent="Noto\'g\'ri parol";',
        '      document.getElementById("pw").value="";',
        '    }',
        '  }catch(e){document.getElementById("err").textContent="Xatolik: "+e.message;}',
        '}',
        '</script></body></html>',
    ]
    login_html = ''.join(parts)
    return web.Response(text=login_html, content_type='text/html', charset='utf-8')

async def admin_login(request):
    import json as _json
    try:
        data = await request.json()
        pw = data.get('password', '')
        if pw == ADMIN_PASSWORD:
            resp = web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
            resp.set_cookie('admin_token', pw, path='/', httponly=True)
            return resp
        return web.Response(text=_json.dumps({'ok': False}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_logout(request):
    """Admin logout"""
    import json as _json
    resp = web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    resp.del_cookie('admin_token', path='/')
    return resp

async def admin_api_attendance(request):
    """Tanlangan kun bo'yicha davomat"""
    import json as _json
    date_str = request.rel_url.query.get('date', datetime.now(UZB_TZ).strftime('%Y-%m-%d'))
    att = [
        {'user_id': a[0], 'branch': a[1], 'time': a[3] if len(a) > 3 else ''}
        for a in daily_attendance_log if a[2] == date_str
    ]
    return web.Response(
        text=_json.dumps({'attendance': att}, ensure_ascii=False, default=str),
        content_type='application/json', charset='utf-8'
    )

async def admin_api_user_status(request):
    """Foydalanuvchi statusini o'zgartirish"""
    import json as _json
    try:
        data = await request.json()
        uid = int(data['user_id'])
        status = data['status']
        if status not in ('active', 'blocked'):
            return web.Response(text=_json.dumps({'ok': False}), content_type='application/json')
        await db.update_user_status(uid, status)
        user_status[uid] = status
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_broadcast(request):
    """Broadcast yuborish — matn va/yoki fayl"""
    import json as _json, asyncio as _asyncio
    try:
        content_type = request.content_type or ''
        if 'multipart' in content_type:
            data = await request.post()
            text = (data.get('text') or '').strip()
            specialty = data.get('specialty', '')
            file_field = data.get('file')
        else:
            data = await request.json()
            text = data.get('text', '').strip()
            specialty = data.get('specialty', '')
            file_field = None

        if not text and not file_field:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Matn yoki fayl kerak'}), content_type='application/json')

        targets = [uid for uid in user_ids
                   if user_status.get(uid) != 'blocked'
                   and (not specialty or user_specialty.get(uid) == specialty)]
        sent, failed = 0, 0

        file_bytes = None
        file_name = None
        file_mime = None
        if file_field and hasattr(file_field, 'file'):
            file_bytes = file_field.file.read()
            file_name = file_field.filename or 'fayl'
            file_mime = file_field.content_type or 'application/octet-stream'

        for uid in targets:
            try:
                if file_bytes:
                    if 'image' in (file_mime or ''):
                        await bot.send_photo(uid, types.BufferedInputFile(file_bytes, filename=file_name), caption=text or None)
                    else:
                        await bot.send_document(uid, types.BufferedInputFile(file_bytes, filename=file_name), caption=text or None)
                elif text:
                    await bot.send_message(uid, text)
                sent += 1
                await _asyncio.sleep(0.05)
            except Exception as e:
                failed += 1
                logging.error(f"Broadcast uid={uid}: {e}")

        await db.save_broadcast(text or f'[Fayl: {file_name}]', sent, failed, specialty)
        return web.Response(text=_json.dumps({'ok': True, 'sent': sent, 'failed': failed}), content_type='application/json')
    except Exception as e:
        logging.error(f"admin_api_broadcast: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_branch_add(request):
    """Yangi filial qo'shish"""
    import json as _json
    try:
        data = await request.json()
        name = data['name'].strip()
        lat = float(data['lat'])
        lon = float(data['lon'])
        if not name or not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Noto\'g\'ri ma\'lumot'}), content_type='application/json')
        # Takrorlanishni tekshirish
        if any(b['name'] == name for b in LOCATIONS):
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Bu nom allaqachon bor'}), content_type='application/json')
        
        # Save to database
        async with db.pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO branches (name, lat, lon) VALUES ($1, $2, $3)
            """, name, lat, lon)
        
        new_branch = {'name': name, 'lat': lat, 'lon': lon}
        LOCATIONS.append(new_branch)
        # Barcha keyboard yangilash
        await update_all_keyboards()
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except ValueError as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': 'Koordinatalar raqam bo\'lishi kerak'}), content_type='application/json')
    except Exception as e:
        logging.error(f"admin_api_branch_add: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_report(request):
    """Hisobot yuklash"""
    report_type = request.match_info.get('type', 'excel')
    try:
        now_uzb = datetime.now(UZB_TZ)
        if report_type == 'excel':
            # create_monthly_excel requires year and month
            buf = await create_monthly_excel(now_uzb.year, now_uzb.month)
            return web.Response(
                body=buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename=oylik_hisobot_{now_uzb.strftime("%Y_%m")}.xlsx'}
            )
        elif report_type == 'pdf':
            buf = await get_combined_report_pdf(now_uzb.date())
            return web.Response(
                body=buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename=davomat_{now_uzb.strftime("%Y_%m_%d")}.xlsx'}
            )
        elif report_type == 'schedule_pdf':
            buf = await create_all_schedules_pdf()
            return web.Response(
                body=buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': 'attachment; filename=jadval.xlsx'}
            )
        else:
            return web.Response(status=404)
    except Exception as e:
        logging.error(f"Report xato: {e}")
        return web.Response(status=500, text=str(e))

async def admin_api_stats(request):
    """Umumiy statistika"""
    import json as _json
    from collections import defaultdict
    now_uzb = datetime.now(UZB_TZ)
    today = now_uzb.strftime("%Y-%m-%d")
    current_month = now_uzb.strftime("%Y-%m")

    total_users = len(user_ids)
    active_users = len([uid for uid in user_ids if user_status.get(uid) != 'blocked'])
    blocked_users = len([uid for uid in user_ids if user_status.get(uid) == 'blocked'])
    today_att = len([k for k in daily_attendance_log if k[2] == today])
    month_att = len([k for k in daily_attendance_log if k[2].startswith(current_month)])
    total_att = len(daily_attendance_log)

    it_c = len([uid for uid in user_ids if user_specialty.get(uid) == 'IT'])
    kr_c = len([uid for uid in user_ids if user_specialty.get(uid) == 'Koreys tili'])
    of_c = len([uid for uid in user_ids if user_specialty.get(uid) == 'Ofis xodimi'])

    branch_stats = defaultdict(int)
    teacher_stats = defaultdict(int)
    for a in daily_attendance_log:
        branch_stats[a[1]] += 1
        teacher_stats[a[0]] += 1

    top_branches = sorted(branch_stats.items(), key=lambda x: x[1], reverse=True)[:10]
    top_teachers = sorted(teacher_stats.items(), key=lambda x: x[1], reverse=True)[:10]
    top_teachers_data = [{'name': user_names.get(uid, str(uid)), 'specialty': user_specialty.get(uid, ''), 'count': cnt} for uid, cnt in top_teachers]

    return web.Response(
        text=_json.dumps({
            'total_users': total_users, 'active_users': active_users, 'blocked_users': blocked_users,
            'today_att': today_att, 'month_att': month_att, 'total_att': total_att,
            'it': it_c, 'kr': kr_c, 'office': of_c,
            'top_branches': [{'branch': b, 'count': c} for b, c in top_branches],
            'top_teachers': top_teachers_data,
            'total_groups': len(groups),
        }, ensure_ascii=False, default=str),
        content_type='application/json', charset='utf-8'
    )

async def admin_api_salary_calc(request):
    """Maosh hisoblash"""
    import json as _json
    try:
        data = await request.json()
        spec = data['specialty']
        teacher_id = int(data['teacher_id'])
        branch_results = data['branches']  # [{branch, students, lessons, perc, penalty, payment}]

        results = []
        for br in branch_results:
            students = int(br.get('students', 0))
            lessons = int(br.get('lessons', 12))
            perc = int(br.get('perc', 35))
            penalty = float(br.get('penalty', 0))
            payment = float(br.get('payment', 0))
            gross = 0
            exam_pen = 0

            if spec == 'IT':
                # 250k guruh: (250000 × s250 / 8) × l250 × perc%
                # 400k guruh: (400000 × s400 / 12) × l400 × perc%
                dtype = br.get('dtype', '')  # '250' yoki '400'
                s250 = int(br.get('s250', 0))
                l250 = int(br.get('l250', 8))
                s400 = int(br.get('s400', 0))
                l400 = int(br.get('l400', 12))
                penalty_pct = float(br.get('penalty', 0))

                gross = 0
                if dtype == '250':
                    rate250 = (250000 * s250 / 8) if s250 > 0 else 0
                    gross = rate250 * l250 * perc / 100
                elif dtype == '400':
                    rate400 = (400000 * s400 / 12) if s400 > 0 else 0
                    gross = rate400 * l400 * perc / 100
                else:
                    rate250 = (250000 * s250 / 8) if s250 > 0 else 0
                    gross250 = rate250 * l250 * perc / 100
                    rate400 = (400000 * s400 / 12) if s400 > 0 else 0
                    gross400 = rate400 * l400 * perc / 100
                    gross = gross250 + gross400
                
                gross_before_penalty = gross
                gross = gross_before_penalty * (1 - penalty_pct / 100)
                bonus = float(br.get('bonus', 0))
                gross += bonus
                pen_disp = f"{int(penalty_pct)}%"
            else:
                # KR formula - 1 ta dars narxini hisoblab, real darslar soniga ko'paytiramiz
                base = 1800000 + (students * 100000 if students > 10 else 0)
                ep_table = {(0,10):900000,(10,20):800000,(20,30):700000,(30,40):600000,(40,50):500000,(50,60):400000,(60,70):300000,(70,80):200000,(80,90):100000,(90,101):0}
                exam_pen = next((v for (lo,hi),v in ep_table.items() if lo<=perc<hi), 0)
                mid = base - exam_pen
                # 1 ta dars narxi = mid / 12 (standart 12 ta dars uchun)
                per_lesson_rate = mid / 12
                # Real darslar soniga ko'paytiramiz
                gross_before_penalty = per_lesson_rate * lessons
                penalty_pct = float(br.get('penalty', 0))
                bonus = float(br.get('bonus', 0))
                gross = gross_before_penalty * (1 - penalty_pct / 100) + bonus
                pen_disp = f"{int(penalty_pct)}%"

            results.append({
                'branch': br['branch'],
                'group_name': br.get('group_name', br['branch']),
                'students': students, 'lessons': lessons, 'perc': perc,
                'penalty': pen_disp, 'payment': payment, 'gross': gross,
                's250': int(br.get('s250', 0)), 'l250': int(br.get('l250', 8)),
                's400': int(br.get('s400', 0)), 'l400': int(br.get('l400', 12)),
                'bonus': float(br.get('bonus', 0)),
            })

        total_gross = sum(r['gross'] for r in results)
        tax = total_gross * 0.075
        net = total_gross - tax
        teacher_name = user_names.get(teacher_id, str(teacher_id))

        return web.Response(
            text=_json.dumps({
                'ok': True, 'teacher_name': teacher_name, 'specialty': spec,
                'results': results,
                'total_gross': total_gross, 'tax': tax, 'net': net
            }, ensure_ascii=False, default=str),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_office_salary_calc(request):
    """Office xodimi oylik hisoblash - Korean system"""
    import json as _json
    try:
        data = await request.json()
        employee_id = int(data['employee_id'])
        position = data['position']  # soeup, sawon, daeri, gwallija
        building = data['building']  # bin_1, bin_2, bin_3
        penalties = data.get('penalties', {})  # {penalty_type: count}
        expenses = data.get('expenses', 0)  # Xarajat cheklari summasi
        
        # Get base salary from database
        async with db.pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT amount, category_kr, bin_name FROM salary_configs WHERE category=$1 AND bin_key=$2",
                position, building
            )
        
        if not row:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Noto\'g\'ri lavozim yoki 호봉'}), content_type='application/json')
        
        base_salary = row['amount']
        pos_name_kr = row['category_kr']
        bin_name = row['bin_name']
        
        # Calculate penalties using fixed amounts from the image
        total_penalty = 0
        penalty_details = []
        
        for penalty_type, count in penalties.items():
            if penalty_type in PENALTY_TYPES and count > 0:
                penalty_info = PENALTY_TYPES[penalty_type]
                # Get fixed amount for this position
                fixed_amount = penalty_info['amounts'].get(position, 0)
                penalty_amount = fixed_amount * count
                total_penalty += penalty_amount
                penalty_details.append({
                    'type': penalty_type,
                    'name': penalty_info['name'],
                    'count': count,
                    'percent': penalty_info['percent'],
                    'base_amount': fixed_amount,
                    'amount': penalty_amount
                })
        
        # Calculate salary (expenses ADDED - reimbursement for employee)
        gross_salary = base_salary - total_penalty + expenses
        tax_amount = gross_salary * TAX_RATE / 100
        net_salary = gross_salary - tax_amount
        
        employee_name = user_names.get(employee_id, str(employee_id))
        
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'employee_name': employee_name,
                'position': pos_name_kr,
                'position_key': position,
                'building': bin_name,
                'base_salary': base_salary,
                'penalties': penalty_details,
                'total_penalty': total_penalty,
                'expenses': expenses,
                'gross_salary': gross_salary,
                'tax_rate': TAX_RATE,
                'tax_amount': tax_amount,
                'net_salary': net_salary
            }, ensure_ascii=False, default=str),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')


async def admin_api_office_salary_excel(request):
    """Office xodimi oylik Excel"""
    import json as _json
    try:
        data = await request.json()
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        employee_name = data.get('employee_name', 'Noma\'lum')
        position = data.get('position', '')
        building = data.get('building', '')
        base_salary = float(data.get('base_salary', 0))
        penalties = data.get('penalties', [])
        total_penalty = float(data.get('total_penalty', 0))
        expenses = float(data.get('expenses', 0))
        gross_salary = float(data.get('gross_salary', 0))
        tax_rate = float(data.get('tax_rate', 7.5))
        tax_amount = float(data.get('tax_amount', 0))
        net_salary = float(data.get('net_salary', 0))
        
        now_uzb = datetime.now(UZB_TZ)
        month_str = now_uzb.strftime('%Y-%m')
        
        COLOR_HDR = "1a3a5c"
        COLOR_COL = "2E86AB"
        COLOR_TOTAL = "4472C4"
        COLOR_TAX = "C00000"
        COLOR_NET = "006100"
        
        thin = Side(border_style="thin", color="AAAAAA")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Oylik"
        ws.row_dimensions[1].height = 28
        
        # Title
        ws.merge_cells('A1:B1')
        ws['A1'] = f"{employee_name} - Oylik hisobot"
        ws['A1'].font = Font(bold=True, size=16, color=COLOR_HDR)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Month
        ws['A2'] = "Oy"
        ws['B2'] = month_str
        ws['A2'].font = Font(bold=True)
        ws['B2'].font = Font(bold=True)
        
        # Headers
        ws['A4'] = "Ko'rsatkich"
        ws['B4'] = "Summa (so'm)"
        for col in ['A4', 'B4']:
            c = ws[col]
            c.fill = PatternFill(start_color=COLOR_COL, end_color=COLOR_COL, fill_type="solid")
            c.font = Font(bold=True, size=11, color="FFFFFF")
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 5
        # Base salary
        ws[f'A{row}'] = "Asosiy oylik"
        ws[f'B{row}'] = f"{int(base_salary):,}"
        ws[f'A{row}'].font = Font(size=11)
        ws[f'B{row}'].font = Font(size=11)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        
        # Penalties
        row += 1
        if penalties:
            for p in penalties:
                ws[f'A{row}'] = f"{p['name']} ({p['count']} × {p['percent']}%)"
                ws[f'B{row}'] = f"-{int(p['amount']):,}"
                ws[f'A{row}'].font = Font(size=10, color=COLOR_TAX)
                ws[f'B{row}'].font = Font(size=10, color=COLOR_TAX)
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
        
        # Total penalty
        if total_penalty > 0:
            ws[f'A{row}'] = "Jami jarima"
            ws[f'B{row}'] = f"-{int(total_penalty):,}"
            ws[f'A{row}'].font = Font(bold=True, size=10, color=COLOR_TAX)
            ws[f'B{row}'].font = Font(bold=True, size=10, color=COLOR_TAX)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Expenses
        if expenses > 0:
            ws[f'A{row}'] = "Xarajat cheklari"
            ws[f'B{row}'] = f"+{int(expenses):,}"
            ws[f'A{row}'].font = Font(size=10, color=COLOR_NET)
            ws[f'B{row}'].font = Font(size=10, color=COLOR_NET)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Gross
        row += 1
        ws[f'A{row}'] = "Jami (soliqsiz)"
        ws[f'B{row}'] = f"{int(gross_salary):,}"
        ws[f'A{row}'].font = Font(bold=True, size=11, color=COLOR_TOTAL)
        ws[f'B{row}'].font = Font(bold=True, size=11, color=COLOR_TOTAL)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        
        # Tax
        row += 1
        ws[f'A{row}'] = f"Soliq ({tax_rate}%)"
        ws[f'B{row}'] = f"-{int(tax_amount):,}"
        ws[f'A{row}'].font = Font(bold=True, size=11, color=COLOR_TAX)
        ws[f'B{row}'].font = Font(bold=True, size=11, color=COLOR_TAX)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        
        # Net
        row += 1
        ws[f'A{row}'] = "Qo'lga tegadi"
        ws[f'B{row}'] = f"{int(net_salary):,}"
        ws[f'A{row}'].font = Font(bold=True, size=12, color=COLOR_NET)
        ws[f'B{row}'].font = Font(bold=True, size=12, color=COLOR_NET)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"Oylik_{employee_name}_{month_str}.xlsx"
        return web.Response(
            body=buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'}
        )
    except Exception as e:
        logging.error(f"office_salary_excel error: {e}", exc_info=True)
        return web.Response(status=500, text=str(e))


async def admin_api_office_employees_list(request):
    """Office xodimlar ro'yxati"""
    import json as _json
    try:
        # Get office workers (Ofis xodimi specialty)
        result = []
        for uid in user_ids:
            spec = user_specialty.get(uid, '')
            if spec == 'Ofis xodimi' or spec == 'Office':
                result.append({
                    'user_id': uid,
                    'name': user_names.get(uid, 'Noma\'lum'),
                    'specialty': spec
                })
        
        return web.Response(
            text=_json.dumps({'ok': True, 'employees': result}, ensure_ascii=False),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')


async def admin_api_salary_structure(request):
    """Oylik tuzilmasi va jarima turlarini qaytarish - salary_configs dan"""
    import json as _json
    try:
        # Get salary structures from salary_configs table
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM salary_configs ORDER BY category, bin_key")
        
        salary_structures = {}
        for row in rows:
            cat = row['category']
            if cat not in salary_structures:
                salary_structures[cat] = {'name': row['category_kr'], 'salaries': {}}
            salary_structures[cat]['salaries'][row['bin_key']] = row['amount']
        
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'salary_structures': salary_structures,
                'buildings': BUILDINGS,
                'penalty_types': PENALTY_TYPES,
                'tax_rate': TAX_RATE
            }, ensure_ascii=False),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')


async def admin_api_teachers_list(request):
    """O'qituvchilar ro'yxati (maosh uchun) — guruhlar bilan"""
    import json as _json
    spec = request.rel_url.query.get('specialty', '')
    result = []
    for uid in user_ids:
        if user_status.get(uid) == 'blocked':
            continue
        if spec and user_specialty.get(uid) != spec:
            continue
        teacher_grp_list = []
        for gid, gdata in groups.items():
            if gdata.get('teacher_id') != uid:
                continue
            studs = group_students.get(gid, [])
            teacher_grp_list.append({
                'id': gid,
                'name': gdata.get('group_name', ''),
                'branch': gdata.get('branch', ''),
                'lesson_type': gdata.get('lesson_type', ''),
                'student_count': len(studs),
            })
        teacher_grp_list.sort(key=lambda g: (g['branch'], g['name']))
        result.append({
            'user_id': uid,
            'name': user_names.get(uid, ''),
            'specialty': user_specialty.get(uid, ''),
            'teacher_groups': teacher_grp_list,
            'branches': list({g['branch'] for g in teacher_grp_list})  # moslik uchun
        })
    return web.Response(text=_json.dumps(result, ensure_ascii=False), content_type='application/json', charset='utf-8')



async def admin_api_payments_summary(request):
    """Admin: barcha guruhlar to'lov summarysi — oy bo'yicha"""
    import json as _json
    try:
        month = request.rel_url.query.get('month', datetime.now(UZB_TZ).strftime('%Y-%m'))
        
        result = []
        async with db.pool.acquire() as conn:
            # Barcha to'lovlar oy bo'yicha
            pay_rows = await conn.fetch(
                "SELECT group_id, student_name, paid, amount, note FROM student_payments WHERE month=$1",
                month
            )
        
        pay_by_grp = {}
        for r in pay_rows:
            gid = r['group_id']
            if gid not in pay_by_grp:
                pay_by_grp[gid] = {'paid': 0, 'total_amt': 0, 'school_amt': 0, 'rows': []}
            if r['student_name'] == '__school__' and r['note'] == 'school':
                # Maktab to'lovi — alohida
                pay_by_grp[gid]['school_amt'] += r['amount']
                pay_by_grp[gid]['total_amt'] += r['amount']
            elif r['paid']:
                pay_by_grp[gid]['paid'] += 1
                pay_by_grp[gid]['total_amt'] += r['amount']
            pay_by_grp[gid]['rows'].append(r)
        
        total_students = 0
        total_paid = 0
        total_amount = 0
        
        for gid, gdata in groups.items():
            studs = group_students.get(gid, [])
            if not studs:
                continue
            grp_pay = pay_by_grp.get(gid, {'paid': 0, 'total_amt': 0, 'rows': []})
            paid_cnt = grp_pay['paid']
            total_cnt = len(studs)
            pct = round(paid_cnt / total_cnt * 100) if total_cnt else 0
            teacher_name = user_names.get(gdata.get('teacher_id'), '—')
            
            school_amt = grp_pay.get('school_amt', 0)
            result.append({
                'group_id': gid,
                'group_name': gdata.get('group_name', ''),
                'branch': gdata.get('branch', ''),
                'lesson_type': gdata.get('lesson_type', ''),
                'teacher_name': teacher_name,
                'total': total_cnt,
                'paid': paid_cnt,
                'unpaid': total_cnt - paid_cnt,
                'pct': pct,
                'amount': grp_pay['total_amt'],
                'school_amount': school_amt,
            })
            total_students += total_cnt
            total_paid += paid_cnt
            total_amount += grp_pay['total_amt']
        
        # Foiz bo'yicha sort
        result.sort(key=lambda x: -x['pct'])
        
        overall_pct = round(total_paid / total_students * 100) if total_students else 0
        
        total_school_amount = sum(g.get('school_amount', 0) for g in result)
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'month': month,
                'groups': result,
                'total_students': total_students,
                'total_paid': total_paid,
                'total_unpaid': total_students - total_paid,
                'total_amount': total_amount,
                'total_school_amount': total_school_amount,
                'overall_pct': overall_pct,
            }, ensure_ascii=False, default=str),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        logging.error(f"payments_summary error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_student_payments(request):
    """Admin: guruh o'quvchilar to'lovi — oy bo'yicha"""
    import json as _json
    try:
        gid = int(request.rel_url.query.get('group_id', 0))
        month = request.rel_url.query.get('month', datetime.now(UZB_TZ).strftime('%Y-%m'))
        if not gid:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'group_id kerak'}), content_type='application/json')
        
        studs = group_students.get(gid, [])
        
        # To'lovlar DB dan
        payments = {}
        try:
            async with db.pool.acquire() as conn:
                rows = await conn.fetch(
                    "SELECT student_name, paid, amount, note FROM student_payments WHERE group_id=$1 AND month=$2",
                    gid, month
                )
                for r in rows:
                    payments[r['student_name']] = {'paid': r['paid'], 'amount': r['amount'], 'note': r['note'] or ''}
        except Exception as e:
            logging.error(f"admin payments fetch: {e}")
        
        # Oylar ro'yxati (to'lov qilingan oylar)
        months_list = []
        try:
            async with db.pool.acquire() as conn:
                rows = await conn.fetch(
                    "SELECT DISTINCT month FROM student_payments WHERE group_id=$1 ORDER BY month DESC",
                    gid
                )
                months_list = [r['month'] for r in rows]
        except: pass
        if month not in months_list:
            months_list.insert(0, month)
        
        students_data = []
        total_paid = 0
        total_amount = 0
        for s in studs:
            pay = payments.get(s['name'], {'paid': False, 'amount': 0, 'note': ''})
            students_data.append({
                'name': s['name'], 'phone': s['phone'],
                'paid': pay['paid'], 'amount': pay['amount'], 'note': pay['note']
            })
            if pay['paid']:
                total_paid += 1
                total_amount += pay['amount']
        
        total = len(studs)
        pct = round(total_paid / total * 100) if total else 0
        
        # Maktab to'lovini olish
        school_payment = payments.get('__school__', {'paid': True, 'amount': 0, 'note': 'school'})
        school_amount = school_payment['amount'] if school_payment.get('note') == 'school' else 0

        return web.Response(
            text=_json.dumps({
                'ok': True,
                'students': students_data,
                'month': month,
                'months_list': months_list,
                'total': total,
                'paid_count': total_paid,
                'unpaid_count': total - total_paid,
                'paid_pct': pct,
                'total_amount': total_amount,
                'school_amount': school_amount,
                'group_name': groups.get(gid, {}).get('group_name', ''),
            }, ensure_ascii=False),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        logging.error(f"admin_api_student_payments error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_student_att(request):
    """O'quvchilar davomati"""
    import json as _json
    group_id = int(request.rel_url.query.get('group_id', 0))
    month_str = request.rel_url.query.get('month', datetime.now(UZB_TZ).strftime('%Y-%m'))
    date_str = request.rel_url.query.get('date', '')  # Kun filtri (YYYY-MM-DD)
    try:
        async with db.pool.acquire() as conn:
            if date_str:
                rows = await conn.fetch("""
                    SELECT student_name, student_phone, lesson_date, status
                    FROM student_attendance
                    WHERE group_id = $1 AND TO_CHAR(lesson_date, 'YYYY-MM-DD') = $2
                    ORDER BY student_name
                """, group_id, date_str)
            else:
                rows = await conn.fetch("""
                    SELECT student_name, student_phone, lesson_date, status
                    FROM student_attendance
                    WHERE group_id = $1 AND TO_CHAR(lesson_date, 'YYYY-MM') = $2
                    ORDER BY lesson_date, student_name
                """, group_id, month_str)
        data = [{'name': r['student_name'], 'phone': r['student_phone'],
                 'date': str(r['lesson_date']), 'status': r['status']} for r in rows]
        return web.Response(text=_json.dumps({'ok': True, 'data': data}, ensure_ascii=False), content_type='application/json', charset='utf-8')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_user_delete(request):
    """Foydalanuvchini arxivlash (o'chirish)"""
    import json as _json
    try:
        data = await request.json()
        uid = int(data['user_id'])
        archived_name = user_names.get(uid, f"ID:{uid}")
        archived_spec = user_specialty.get(uid, '')
        async with db.pool.acquire() as conn:
            await conn.execute(
                "UPDATE users SET status='deleted', full_name=$1 WHERE user_id=$2",
                f"[ARXIV] {archived_name}", uid
            )
            await conn.execute("DELETE FROM schedules WHERE user_id = $1", uid)
            teacher_group_ids = await conn.fetch("SELECT id FROM groups WHERE teacher_id = $1", uid)
            for grp_row in teacher_group_ids:
                await conn.execute("UPDATE groups SET teacher_id=NULL WHERE id=$1", grp_row['id'])
                if grp_row['id'] in groups:
                    groups[grp_row['id']]['teacher_id'] = None
        if uid in user_ids:
            user_ids.remove(uid)
        user_status[uid] = 'deleted'
        user_names[uid] = f"[ARXIV] {archived_name}"
        if uid in user_schedules:
            for schedule_id in user_schedules[uid]:
                schedules.pop(schedule_id, None)
            user_schedules.pop(uid, None)
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        logging.error(f"admin_api_user_delete error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_user_restore(request):
    """Foydalanuvchini tiklash (arxivdan chiqarish)"""
    import json as _json
    try:
        data = await request.json()
        uid = int(data['user_id'])
        async with db.pool.acquire() as conn:
            # Get archived name and restore
            row = await conn.fetchrow("SELECT full_name FROM users WHERE user_id=$1", uid)
            if row and row['full_name']:
                clean_name = row['full_name'].replace('[ARXIV]', '').strip()
                await conn.execute(
                    "UPDATE users SET status='active', full_name=$1 WHERE user_id=$2",
                    clean_name, uid
                )
            else:
                await conn.execute("UPDATE users SET status='active' WHERE user_id=$1", uid)
        user_status[uid] = 'active'
        if row and row['full_name']:
            user_names[uid] = row['full_name'].replace('[ARXIV]', '').strip()
        if uid not in user_ids:
            user_ids.append(uid)
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        logging.error(f"admin_api_user_restore error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_user_permanent_delete(request):
    """Foydalanuvchini UMUMAN o'chirish (bazadan)"""
    import json as _json
    try:
        data = await request.json()
        uid = int(data['user_id'])
        async with db.pool.acquire() as conn:
            # Delete all related data (faqat mavjud jadvallar)
            await conn.execute("DELETE FROM attendance WHERE user_id=$1", uid)
            await conn.execute("DELETE FROM schedules WHERE user_id=$1", uid)
            # Remove teacher from groups
            await conn.execute("UPDATE groups SET teacher_id=NULL WHERE teacher_id=$1", uid)
            # Delete user
            await conn.execute("DELETE FROM users WHERE user_id=$1", uid)
        
        # Clear from memory
        if uid in user_ids:
            user_ids.remove(uid)
        user_status.pop(uid, None)
        user_names.pop(uid, None)
        user_specialty.pop(uid, None)
        user_languages.pop(uid, None)
        user_photo_cache.pop(uid, None)
        if uid in user_schedules:
            for schedule_id in user_schedules[uid]:
                schedules.pop(schedule_id, None)
            user_schedules.pop(uid, None)
        
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        logging.error(f"admin_api_user_permanent_delete error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_user_stats(request):
    """Foydalanuvchi batafsil statistikasi"""
    import json as _json
    try:
        uid = int(request.rel_url.query.get('user_id', 0))
        branch_stats = {}
        month_stats = {}
        for (user_id, branch, date, time) in daily_attendance_log:
            if user_id == uid:
                branch_stats[branch] = branch_stats.get(branch, 0) + 1
                m = date[:7]
                month_stats[m] = month_stats.get(m, 0) + 1
        teacher_groups = [{'id': gid, 'name': gdata['group_name'], 'branch': gdata['branch']}
                          for gid, gdata in groups.items() if gdata.get('teacher_id') == uid]
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'name': user_names.get(uid, ''),
                'specialty': user_specialty.get(uid, ''),
                'status': user_status.get(uid, 'active'),
                'branch_stats': [{'branch': b, 'count': c} for b, c in sorted(branch_stats.items(), key=lambda x: -x[1])],
                'month_stats': [{'month': m, 'count': c} for m, c in sorted(month_stats.items(), reverse=True)],
                'groups': teacher_groups,
                'total': sum(branch_stats.values()),
            }, ensure_ascii=False),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_group_detail(request):
    """Guruh to'liq ma'lumoti (o'quvchilar, jadval)"""
    import json as _json
    try:
        gid = int(request.rel_url.query.get('id', 0))
        async with db.pool.acquire() as conn:
            grp = await conn.fetchrow("SELECT * FROM groups WHERE id=$1", gid)
            stds = await conn.fetch("SELECT * FROM group_students WHERE group_id=$1 ORDER BY id", gid)
        if not grp:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Topilmadi'}), content_type='application/json')
        raw = grp['days_data']
        if isinstance(raw, str):
            raw = _json.loads(raw)
        teacher_id = grp['teacher_id']
        return web.Response(text=_json.dumps({
            'ok': True,
            'id': grp['id'],
            'group_name': grp['group_name'],
            'branch': grp['branch'],
            'lesson_type': grp['lesson_type'],
            'teacher_id': teacher_id,
            'teacher_name': user_names.get(teacher_id, '—') if teacher_id else '—',
            'day_times': raw if isinstance(raw, dict) else {d: grp['time_text'] or '—' for d in (raw or [])},
            'days': list(raw.keys()) if isinstance(raw, dict) else (raw or []),
            'time_text': grp['time_text'],
            'students': [{'id': s['id'], 'name': s['student_name'], 'phone': s['student_phone']} for s in stds],
            'student_count': grp.get('student_count') or len(stds),
        }, ensure_ascii=False, default=str), content_type='application/json', charset='utf-8')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_group_edit_schedule(request):
    """Guruh jadvalini tahrirlash"""
    import json as _json
    try:
        data = await request.json()
        gid = int(data['group_id'])
        day_times = data['day_times']  # {day: time}
        days = list(day_times.keys())
        first_time = list(day_times.values())[0] if day_times else '—'
        async with db.pool.acquire() as conn:
            await conn.execute(
                "UPDATE groups SET days_data=$1, time_text=$2 WHERE id=$3",
                _json.dumps(day_times), first_time, gid
            )
        if gid in groups:
            groups[gid]['days'] = days
            groups[gid]['day_times'] = day_times
            groups[gid]['time'] = first_time
            groups[gid]['time_text'] = first_time
        # O'qituvchiga xabar
        grp = groups.get(gid, {})
        teacher_id = grp.get('teacher_id')
        if teacher_id:
            time_display = ", ".join([f"{d}: {t}" for d, t in day_times.items()])
            try:
                await bot.send_message(
                    teacher_id,
                    f"📅 *{grp.get('group_name','Guruh')}* guruhingizning dars jadvali yangilandi!\n\n"
                    f"📆 Yangi kunlar va vaqtlar:\n{time_display}",
                    parse_mode="Markdown"
                )
            except:
                pass
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_group_edit_teacher(request):
    """Guruh o'qituvchisini almashtirish"""
    import json as _json
    try:
        data = await request.json()
        gid = int(data['group_id'])
        new_tid = int(data['teacher_id'])
        grp = groups.get(gid, {})
        old_tid = grp.get('teacher_id')
        async with db.pool.acquire() as conn:
            await conn.execute("UPDATE groups SET teacher_id=$1 WHERE id=$2", new_tid, gid)
        if gid in groups:
            groups[gid]['teacher_id'] = new_tid
        new_name = user_names.get(new_tid, str(new_tid))
        # Eski o'qituvchiga xabar
        if old_tid and old_tid != new_tid:
            try:
                old_lang = user_languages.get(old_tid, 'uz')
                if old_lang == 'ru':
                    old_msg = f"ℹ️ Группа *{grp.get('group_name','Guruh')}* передана другому преподавателю."
                else:
                    old_msg = f"ℹ️ *{grp.get('group_name','Guruh')}* guruhi sizdan boshqa o'qituvchiga o'tkazildi."
                await bot.send_message(old_tid, old_msg, parse_mode="Markdown")
            except: pass
        # Yangi o'qituvchiga xabar
        try:
            new_lang = user_languages.get(new_tid, 'uz')
            if new_lang == 'ru':
                new_msg = (
                    f"🎉 Вы назначены новым преподавателем группы *{grp.get('group_name','Guruh')}*!\n\n"
                    f"🏢 Филиал: {grp.get('branch','—')}\n"
                    f"📆 Дни: {', '.join(grp.get('days',[]))}\n"
                    f"⏰ Время: {grp.get('time_text','—')}"
                )
            else:
                new_msg = (
                    f"🎉 Siz *{grp.get('group_name','Guruh')}* guruhining yangi o'qituvchisi sifatida tayinlandingiz!\n\n"
                    f"🏢 Filial: {grp.get('branch','—')}\n"
                    f"📆 Kunlar: {', '.join(grp.get('days',[]))}\n"
                    f"⏰ Vaqt: {grp.get('time_text','—')}"
                )
            await bot.send_message(new_tid, new_msg, parse_mode="Markdown")
        except: pass
        return web.Response(text=_json.dumps({'ok': True, 'teacher_name': new_name}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_group_edit_branch(request):
    """Guruh filialini almashtirish"""
    import json as _json
    try:
        data = await request.json()
        gid = int(data['group_id'])
        new_branch = data['branch'].strip()
        if not new_branch:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Filial nomi kerak'}), content_type='application/json')
        if new_branch not in [loc['name'] for loc in LOCATIONS]:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Noto\'g\'ri filial'}), content_type='application/json')
        async with db.pool.acquire() as conn:
            await conn.execute("UPDATE groups SET branch=$1 WHERE id=$2", new_branch, gid)
        if gid in groups:
            groups[gid]['branch'] = new_branch
        return web.Response(text=_json.dumps({'ok': True, 'branch': new_branch}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')



async def api_submit_application(request):
    """Ariza yuborish — saytdan"""
    import json as _json
    try:
        data = await request.json()
        name = data.get('name','').strip()
        phone = data.get('phone','').strip()
        course = data.get('course','').strip()
        message = data.get('message','').strip()
        if not name or not phone:
            return web.Response(text=_json.dumps({'ok':False,'error':'Ism va telefon kerak'}), content_type='application/json')
        async with db.pool.acquire() as conn:
            await conn.execute(
                "INSERT INTO applications (name, phone, course, message) VALUES ($1,$2,$3,$4)",
                name, phone, course, message
            )
        # Admin ga xabar
        try:
            msg = "\U0001f4e9 Yangi ariza!\n\U0001f464 " + str(name) + "\n\U0001f4de " + str(phone)
            if course: msg += "\n\U0001f4da Kurs: " + str(course)
            if message: msg += "\n\U0001f4ac " + str(message)
            await bot.send_message(ADMIN_GROUP_ID, msg)
        except: pass
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def api_bootcamp_apply(request):
    """Bootcampga ariza yuborish"""
    import json as _json
    try:
        data = await request.json()
        fname = data.get('fname','').strip()
        lname = data.get('lname','').strip()
        phone = data.get('phone','').strip()
        dob = data.get('dob','').strip()
        email = data.get('email','').strip()
        about = data.get('about','').strip()
        skills = data.get('skills','').strip()
        track = data.get('track','').strip()
        resume_url = data.get('resume_url','').strip()
        resume_name = data.get('resume_name','').strip()
        
        if not fname or not lname or not phone:
            return web.Response(text=_json.dumps({'ok':False,'error':'Ism va telefon kerak'}), content_type='application/json')
        async with db.pool.acquire() as conn:
            await conn.execute(
                "INSERT INTO bootcamp_applications (fname, lname, phone, dob, email, about, skills, track, resume_url, resume_name) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)",
                fname, lname, phone, dob, email, about, skills, track, resume_url, resume_name
            )
        try:
            track_text = "Rassom" if track == "artist" else "Dasturchi"
            msg = f"\U0001f393 Yangi Bootcamp ariza!\n\U0001f464 {fname} {lname}\n\U0001f4de {phone}\n\U0001f4c5 Tug'ilgan sana: {dob}\n\U0001f4e7 Email: {email}\n\U0001f4dd Yo'nalish: {track_text}"
            if resume_url:
                msg += f"\n\U0001f4c4 Resume: {resume_url}"
            await bot.send_message(ADMIN_GROUP_ID, msg)
        except: pass
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_bootcamp_applications_get(request):
    """Bootcamp arizalarini olish"""
    import json as _json
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM bootcamp_applications ORDER BY created_at DESC LIMIT 100")
            apps = [dict(r) for r in rows]
            for a in apps:
                if a.get('created_at'):
                    a['created_at'] = a['created_at'].strftime('%d.%m.%Y %H:%M')
        return web.Response(text=_json.dumps({'ok':True,'applications':apps}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_bootcamp_application_status(request):
    """Bootcamp ariza statusini yangilash"""
    import json as _json
    try:
        data = await request.json()
        app_id = int(data.get('id',0))
        status = data.get('status','new')
        async with db.pool.acquire() as conn:
            await conn.execute("UPDATE bootcamp_applications SET status=$1 WHERE id=$2", status, app_id)
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_bootcamp_application_delete(request):
    """Bootcamp arizani o'chirish"""
    import json as _json
    try:
        data = await request.json()
        app_id = int(data.get('id',0))
        async with db.pool.acquire() as conn:
            await conn.execute("DELETE FROM bootcamp_applications WHERE id=$1", app_id)
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_applications_get(request):
    """Arizalarni olish"""
    import json as _json
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM applications ORDER BY created_at DESC LIMIT 100")
            apps = [dict(r) for r in rows]
            for a in apps:
                if a.get('created_at'):
                    a['created_at'] = a['created_at'].strftime('%d.%m.%Y %H:%M')
        return web.Response(text=_json.dumps({'ok':True,'applications':apps}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_application_status(request):
    """Ariza statusini yangilash"""
    import json as _json
    try:
        data = await request.json()
        app_id = int(data['id'])
        status = data['status']
        async with db.pool.acquire() as conn:
            await conn.execute("UPDATE applications SET status=$1 WHERE id=$2", status, app_id)
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_application_delete(request):
    """Arizani o'chirish"""
    import json as _json
    try:
        data = await request.json()
        app_id = int(data['id'])
        async with db.pool.acquire() as conn:
            await conn.execute("DELETE FROM applications WHERE id=$1", app_id)
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def api_get_news(request):
    """Yangiliklar — saytdan"""
    import json as _json
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM news WHERE is_published=TRUE ORDER BY created_at DESC LIMIT 10")
            news = [dict(r) for r in rows]
            for n in news:
                if n.get('created_at'):
                    n['created_at'] = n['created_at'].strftime('%d.%m.%Y')
        return web.Response(text=_json.dumps({'ok':True,'news':news}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def api_branches_map(request):
    """Sayt uchun: filiallar + ularning faol kurslari (dars turi + vaqt)"""
    import json as _json
    try:
        result = []
        for loc in LOCATIONS:
            # Bu filialdagi faol guruhlarni topamiz
            branch_groups = [
                gdata for gdata in groups.values()
                if gdata.get('branch') == loc['name']
            ]
            if not branch_groups:
                continue
            # Dars turlarini yig'amiz (takrorlanmasdan)
            lessons = []
            seen = set()
            for g in branch_groups:
                lt = g.get('lesson_type', '')
                day_times = g.get('day_times', {})
                days = g.get('days', list(day_times.keys()))
                for day, time_val in day_times.items():
                    key = (lt, day, time_val)
                    if key not in seen:
                        seen.add(key)
                        lessons.append({'type': lt, 'day': day, 'time': time_val})
            if not lessons:
                continue
            result.append({
                'name': loc['name'],
                'lat': loc['lat'],
                'lon': loc['lon'],
                'lessons': lessons
            })
        return web.Response(
            text=_json.dumps({'ok': True, 'branches': result}, ensure_ascii=False),
            content_type='application/json',
            charset='utf-8'
        )
    except Exception as e:
        logging.error(f"api_branches_map error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_news_save(request):
    """Yangilik saqlash — 3 tilli"""
    import json as _json
    try:
        data = await request.json()
        nid = data.get('id')
        title = data.get('title','').strip()
        body = data.get('body','').strip()
        title_ru = data.get('title_ru','').strip()
        body_ru = data.get('body_ru','').strip()
        title_kr = data.get('title_kr','').strip()
        body_kr = data.get('body_kr','').strip()
        image_url = data.get('image_url','').strip()
        is_published = data.get('is_published', True)
        if not title or not body:
            return web.Response(text=_json.dumps({'ok':False,'error':'Sarlavha va matn kerak'}), content_type='application/json')
        async with db.pool.acquire() as conn:
            if nid:
                await conn.execute(
                    "UPDATE news SET title=$1,body=$2,title_ru=$3,body_ru=$4,title_kr=$5,body_kr=$6,image_url=$7,is_published=$8 WHERE id=$9",
                    title, body, title_ru, body_ru, title_kr, body_kr, image_url, is_published, int(nid)
                )
            else:
                await conn.execute(
                    "INSERT INTO news (title,body,title_ru,body_ru,title_kr,body_kr,image_url,is_published) VALUES ($1,$2,$3,$4,$5,$6,$7,$8)",
                    title, body, title_ru, body_ru, title_kr, body_kr, image_url, is_published
                )
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_news_delete(request):
    """Yangilik o'chirish"""
    import json as _json
    try:
        data = await request.json()
        nid = int(data['id'])
        async with db.pool.acquire() as conn:
            await conn.execute("DELETE FROM news WHERE id=$1", nid)
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_news_get(request):
    """Admin: barcha yangiliklar"""
    import json as _json
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM news ORDER BY created_at DESC")
            news = [dict(r) for r in rows]
            for n in news:
                if n.get('created_at'):
                    n['created_at'] = n['created_at'].strftime('%d.%m.%Y %H:%M')
        return web.Response(text=_json.dumps({'ok':True,'news':news}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def api_get_partners(request):
    """Public: sayt uchun partnerlar ro'yxati"""
    import json as _json
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM partners ORDER BY sort_order ASC, id ASC")
            partners = []
            for r in rows:
                p = dict(r)
                if p.get('created_at'):
                    p['created_at'] = p['created_at'].strftime('%d.%m.%Y %H:%M')
                partners.append(p)
        return web.Response(text=_json.dumps({'ok':True,'partners':partners}), content_type='application/json')
    except Exception as e:
        logging.error(f"api_get_partners error: {e}")
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_partners_get(request):
    """Admin: barcha partnerlar"""
    import json as _json
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM partners ORDER BY sort_order ASC, id DESC")
            partners = []
            for r in rows:
                p = dict(r)
                if p.get('created_at'):
                    p['created_at'] = p['created_at'].strftime('%d.%m.%Y %H:%M')
                partners.append(p)
        return web.Response(text=_json.dumps({'ok':True,'partners':partners}), content_type='application/json')
    except Exception as e:
        logging.error(f"admin_api_partners_get error: {e}")
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_partners_save(request):
    """Partner saqlash"""
    import json as _json
    try:
        data = await request.json()
        logging.info(f"Partner save data: {data}")
        pid = data.get('id')
        name = data.get('name','').strip()
        logo_url = data.get('logo_url','').strip()
        website_url = data.get('website_url','').strip()
        sort_order = int(data.get('sort_order') or 0)
        if not name:
            return web.Response(text=_json.dumps({'ok':False,'error':'Nomi kerak'}), content_type='application/json')
        async with db.pool.acquire() as conn:
            if pid:
                logging.info(f"Updating partner {pid}: {name}")
                await conn.execute(
                    "UPDATE partners SET name=$1, logo_url=$2, website_url=$3, sort_order=$4 WHERE id=$5",
                    name, logo_url, website_url, sort_order, int(pid)
                )
            else:
                logging.info(f"Inserting new partner: {name}")
                await conn.execute(
                    "INSERT INTO partners (name, logo_url, website_url, sort_order) VALUES ($1,$2,$3,$4)",
                    name, logo_url, website_url, sort_order
                )
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_partners_delete(request):
    """Partner o'chirish"""
    import json as _json
    try:
        data = await request.json()
        pid = int(data.get('id', 0))
        if not pid:
            return web.Response(text=_json.dumps({'ok':False,'error':'ID kerak'}), content_type='application/json')
        async with db.pool.acquire() as conn:
            await conn.execute("DELETE FROM partners WHERE id=$1", pid)
        return web.Response(text=_json.dumps({'ok':True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok':False,'error':str(e)}), content_type='application/json')

async def admin_api_site_config_get(request):
    """Sayt konfiguratsiyasini olish"""
    import json as _json
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT key, value FROM site_config ORDER BY key")
            config = {r['key']: r['value'] for r in rows}
        return web.Response(text=_json.dumps({'ok': True, 'config': config}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_site_config_save(request):
    """Sayt konfiguratsiyasini saqlash"""
    import json as _json
    try:
        data = await request.json()
        async with db.pool.acquire() as conn:
            for key, value in data.items():
                await conn.execute("""
                    INSERT INTO site_config (key, value, updated_at)
                    VALUES ($1, $2, NOW())
                    ON CONFLICT (key) DO UPDATE SET value=$2, updated_at=NOW()
                """, key, str(value))
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_upload_image(request):
    """Rasm yuklash"""
    import json as _json
    try:
        reader = await request.multipart()
        async for part in reader:
            if part.name == 'image':
                filename = part.filename
                if not filename:
                    return web.Response(text=_json.dumps({'ok': False, 'error': 'Fayl tanlanmadi'}), content_type='application/json')
                import uuid
                ext = os.path.splitext(filename)[1].lower()
                if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg'):
                    return web.Response(text=_json.dumps({'ok': False, 'error': 'Rasm formati noto\'g\'ri'}), content_type='application/json')
                new_filename = f"{uuid.uuid4().hex}{ext}"
                static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
                os.makedirs(static_dir, exist_ok=True)
                filepath = os.path.join(static_dir, new_filename)
                data = await part.read()
                with open(filepath, 'wb') as f:
                    f.write(data)
                return web.Response(text=_json.dumps({'ok': True, 'url': f'/static/{new_filename}'}), content_type='application/json')
        return web.Response(text=_json.dumps({'ok': False, 'error': 'Fayl topilmadi'}), content_type='application/json')
    except Exception as e:
        logging.error(f"upload_image error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def api_upload_resume(request):
    """Resume fayli yuklash"""
    import json as _json
    try:
        reader = await request.multipart()
        async for part in reader:
            if part.name == 'resume':
                filename = part.filename
                if not filename:
                    return web.Response(text=_json.dumps({'ok': False, 'error': 'Fayl tanlanmadi'}), content_type='application/json')
                import uuid
                ext = os.path.splitext(filename)[1].lower()
                if ext not in ('.pdf', '.doc', '.docx', '.jpg', '.jpeg', '.png'):
                    return web.Response(text=_json.dumps({'ok': False, 'error': 'Noto\'g\'ri format'}), content_type='application/json')
                new_filename = f"{uuid.uuid4().hex}{ext}"
                resumes_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resumes')
                os.makedirs(resumes_dir, exist_ok=True)
                filepath = os.path.join(resumes_dir, new_filename)
                data = await part.read()
                with open(filepath, 'wb') as f:
                    f.write(data)
                return web.Response(text=_json.dumps({'ok': True, 'url': f'/resumes/{new_filename}', 'name': filename}), content_type='application/json')
        return web.Response(text=_json.dumps({'ok': False, 'error': 'Fayl topilmadi'}), content_type='application/json')
    except Exception as e:
        logging.error(f"upload_resume error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_group_edit_name(request):
    """Guruh nomini o'zgartirish"""
    import json as _json
    try:
        data = await request.json()
        gid = int(data['group_id'])
        new_name = data['name'].strip()
        if not new_name:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Guruh nomi kerak'}), content_type='application/json')
        async with db.pool.acquire() as conn:
            await conn.execute("UPDATE groups SET group_name=$1 WHERE id=$2", new_name, gid)
        if gid in groups:
            groups[gid]['group_name'] = new_name
        return web.Response(text=_json.dumps({'ok': True, 'name': new_name}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_student_add(request):
    """O'quvchi qo'shish"""
    import json as _json
    try:
        data = await request.json()
        gid = int(data['group_id'])
        name = data['name'].strip()
        phone = data['phone'].strip()
        if not name or not phone:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Ism va telefon kerak'}), content_type='application/json')
        async with db.pool.acquire() as conn:
            sid = await conn.fetchval(
                "INSERT INTO group_students(group_id, student_name, student_phone) VALUES($1,$2,$3) RETURNING id",
                gid, name, phone
            )
        return web.Response(text=_json.dumps({'ok': True, 'id': sid, 'name': name, 'phone': phone}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_student_edit(request):
    """O'quvchi ma'lumotini tahrirlash"""
    import json as _json
    try:
        data = await request.json()
        sid = int(data['student_id'])
        name = data.get('name', '').strip()
        phone = data.get('phone', '').strip()
        if name:
            async with db.pool.acquire() as conn:
                await conn.execute("UPDATE group_students SET student_name=$1 WHERE id=$2", name, sid)
        if phone:
            async with db.pool.acquire() as conn:
                await conn.execute("UPDATE group_students SET student_phone=$1 WHERE id=$2", phone, sid)
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_student_delete(request):
    """O'quvchini o'chirish"""
    import json as _json
    try:
        data = await request.json()
        sid = int(data['student_id'])
        async with db.pool.acquire() as conn:
            std = await conn.fetchrow("SELECT * FROM group_students WHERE id=$1", sid)
            await conn.execute("DELETE FROM group_students WHERE id=$1", sid)
        return web.Response(text=_json.dumps({'ok': True, 'name': std['student_name'] if std else ''}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_group_create(request):
    """Yangi guruh yaratish"""
    import json as _json
    try:
        data = await request.json()
        branch = data['branch']
        lesson_type = data['lesson_type']
        teacher_id_raw = data.get('teacher_id')
        teacher_id = int(teacher_id_raw) if teacher_id_raw else None
        group_name = data['group_name'].strip()
        day_times = data.get('day_times', {})  # {day: time}
        students = data.get('students', [])  # [{name, phone}]
        student_count = data.get('student_count', 0)  # FREE TEACHER uchun o'quvchi soni
        if not group_name:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Guruh nomi kerak'}), content_type='application/json')
        days = list(day_times.keys()) if day_times else []
        first_time = list(day_times.values())[0] if day_times else ''
        async with db.pool.acquire() as conn:
            gid = await conn.fetchval(
                "INSERT INTO groups(group_name, branch, lesson_type, teacher_id, days_data, time_text, student_count) VALUES($1,$2,$3,$4,$5,$6,$7) RETURNING id",
                group_name, branch, lesson_type, teacher_id, _json.dumps(day_times), first_time, student_count
            )
            for std in students:
                if std.get('name') and std.get('phone'):
                    await conn.execute(
                        "INSERT INTO group_students(group_id, student_name, student_phone) VALUES($1,$2,$3)",
                        gid, std['name'], std['phone']
                    )
        # RAM ga qo'shish
        groups[gid] = {
            'group_name': group_name, 'branch': branch, 'lesson_type': lesson_type,
            'teacher_id': teacher_id, 'days': days, 'day_times': day_times,
            'time': first_time, 'time_text': first_time, 'created_at': datetime.now(UZB_TZ),
            'student_count': student_count if student_count else len(students)
        }
        # O'quvchilarni ham RAM ga saqlash
        group_students[gid] = [
            {'name': s['name'], 'phone': s['phone']}
            for s in students if s.get('name') and s.get('phone')
        ]
        # O'qituvchiga xabar
        try:
            time_disp = ", ".join([f"{d}: {t}" for d, t in day_times.items()])
            lang = user_languages.get(teacher_id, 'uz')
            if lang == 'ru':
                msg = (
                    f"🎉 Вам назначена новая группа!\n\n"
                    f"👥 Группа: *{group_name}*\n"
                    f"🏢 Филиал: {branch}\n"
                    f"📚 Предмет: {lesson_type}\n"
                    f"📆 Расписание: {time_disp}\n"
                    f"🧑\u200d🎓 Учеников: {len(students)}"
                )
            else:
                msg = (
                    f"🎉 Sizga yangi guruh biriktirildi!\n\n"
                    f"👥 Guruh: *{group_name}*\n"
                    f"🏢 Filial: {branch}\n"
                    f"📚 Fan: {lesson_type}\n"
                    f"📆 Jadval: {time_disp}\n"
                    f"🧑\u200d🎓 O'quvchilar: {len(students)} ta"
                )
            await bot.send_message(teacher_id, msg, parse_mode="Markdown")
        except: pass
        return web.Response(text=_json.dumps({'ok': True, 'group_id': gid}), content_type='application/json')
    except Exception as e:
        logging.error(f"admin_api_group_create error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_group_excel(request):
    """Guruh Excel faylini yuklash"""
    import json as _json
    try:
        gid = int(request.rel_url.query.get('id', 0))
        async with db.pool.acquire() as conn:
            grp = await conn.fetchrow("SELECT * FROM groups WHERE id=$1", gid)
            stds = await conn.fetch("SELECT * FROM group_students WHERE group_id=$1 ORDER BY id", gid)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "O'quvchilar"
        ws["A1"] = f"Guruh: {grp['group_name']} | {grp['branch']}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="4472C4")
        ws.merge_cells("A1:B1")
        for col, h in enumerate(["Ism Familiya", "Telefon raqami"], 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E86AB")
            cell.alignment = Alignment(horizontal="center")
        for idx, std in enumerate(stds, 1):
            ws.cell(row=2+idx, column=1, value=std["student_name"])
            ws.cell(row=2+idx, column=2, value=std["student_phone"])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"oquvchilar_{grp['group_name']}.xlsx"
        return web.Response(
            body=buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={fname}'}
        )
    except Exception as e:
        return web.Response(status=500, text=str(e))

async def admin_api_branch_delete(request):
    """Filialni o'chirish"""
    import json as _json
    try:
        data = await request.json()
        idx = int(data['index'])
        if 0 <= idx < len(LOCATIONS):
            removed = LOCATIONS.pop(idx)
            # Delete from database
            async with db.pool.acquire() as conn:
                await conn.execute("DELETE FROM branches WHERE name = $1", removed['name'])
            return web.Response(text=_json.dumps({'ok': True, 'name': removed['name']}), content_type='application/json')
        return web.Response(text=_json.dumps({'ok': False, 'error': 'Index xato'}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_branch_update(request):
    """Filialni yangilash"""
    import json as _json
    try:
        data = await request.json()
        idx = int(data['index'])
        name = data['name'].strip()
        lat = float(data['lat'])
        lon = float(data['lon'])
        if not name or not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Noto\'g\'ri ma\'lumot'}), content_type='application/json')
        if 0 <= idx < len(LOCATIONS):
            old_name = LOCATIONS[idx]['name']
            LOCATIONS[idx] = {'name': name, 'lat': lat, 'lon': lon}
            # Update in database
            async with db.pool.acquire() as conn:
                await conn.execute("UPDATE branches SET name=$1, lat=$2, lon=$3 WHERE name=$4", name, lat, lon, old_name)
            await update_all_keyboards()
            return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
        return web.Response(text=_json.dumps({'ok': False, 'error': 'Index xato'}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')


async def admin_api_salary_excel(request):
    """Maosh hisoblash natijasini Excel - har filial alohida sheet + Umumiy sheet"""
    import json as _json
    try:
        data = await request.json()
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        teacher_name = data['teacher_name']
        specialty    = data['specialty']
        results      = data['results']
        total_gross  = float(data['total_gross'])
        tax          = float(data['tax'])
        net          = float(data['net'])
        now_uzb      = datetime.now(UZB_TZ)
        month_str    = now_uzb.strftime('%Y-%m')

        # Ranglar
        COLOR_HDR   = "1a3a5c"   # sarlavha
        COLOR_COL   = "2E86AB"   # ustun nomi
        COLOR_TOTAL = "4472C4"
        COLOR_TAX   = "C00000"
        COLOR_NET   = "006100"
        COLOR_ROW1  = "EBF4FA"
        COLOR_ROW2  = "FFFFFF"

        thin   = Side(border_style="thin", color="AAAAAA")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        def hdr_cell(ws, row, col, val, bg=COLOR_COL, fg="FFFFFF", sz=11, bold=True):
            c = ws.cell(row=row, column=col, value=val)
            c.fill  = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.font  = Font(bold=bold, size=sz, color=fg)
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            return c

        def data_cell(ws, row, col, val, bold=False, color="000000", align="center", bg=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font   = Font(bold=bold, size=10, color=color)
            c.border = border
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            if bg:
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            return c

        def write_summary_rows(ws, cur_row, total_gross, tax, net, ncols):
            """Jami qatorlarni yozish"""
            ws.append([])
            cur_row += 1
            for label, val, col in [
                ("JAMI (soliqsiz)", total_gross, COLOR_TOTAL),
                ("Soliq (7.5%)",    tax,         COLOR_TAX),
                ("QO'LGA TEGADI",  net,         COLOR_NET),
            ]:
                r = ws.max_row + 1
                ws.append([""] * (ncols - 2) + [label, f"{int(val):,} so'm".replace(',', ' ')])
                for ci in range(1, ncols + 1):
                    c = ws.cell(row=r, column=ci)
                    c.border = border
                c = ws.cell(row=r, column=ncols - 1)
                c.font = Font(bold=True, size=11)
                c.alignment = Alignment(horizontal="right")
                c = ws.cell(row=r, column=ncols)
                c.font = Font(bold=True, size=12, color=col)
                c.alignment = Alignment(horizontal="center")
                c.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        wb = Workbook()

        # ═══════════════════════════════════════════════
        # 1. UMUMIY SHEET — barcha filiallar
        # ═══════════════════════════════════════════════
        ws0 = wb.active
        ws0.title = "Umumiy"
        ws0.row_dimensions[1].height = 28

        is_it = (specialty == 'IT')
        if is_it:
            headers = ["Guruh", "Filial", "250k o'quvchi", "250k dars", "400k o'quvchi", "400k dars",
                       "Imtixon %", "Jarima", "Bonus", "Hisoblangan"]
            col_w   = [22, 20, 13, 10, 13, 10, 10, 10, 12, 16]
        else:
            headers = ["Guruh", "Filial", "O'quvchilar", "Darslar", "Imtixon %",
                       "Jarima (so'm)", "Bonus", "Hisoblangan"]
            col_w   = [22, 20, 12, 10, 10, 15, 12, 16]

        ncols = len(headers)

        # Sarlavha
        ws0.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        t = ws0.cell(row=1, column=1, value=f"OYLIK HISOBOT — {teacher_name}  ({specialty})  {month_str}")
        t.font      = Font(bold=True, size=13, color="FFFFFF")
        t.fill      = PatternFill(start_color=COLOR_HDR, end_color=COLOR_HDR, fill_type="solid")
        t.alignment = Alignment(horizontal="center", vertical="center")
        t.border    = border

        # Ustun nomlari
        for ci, h in enumerate(headers, 1):
            hdr_cell(ws0, 2, ci, h)

        # Qatorlar
        for ri, r in enumerate(results):
            bg = COLOR_ROW1 if ri % 2 == 0 else COLOR_ROW2
            cur = ws0.max_row + 1
            if is_it:
                vals = [r.get('group_name', r['branch']), r['branch'],
                        r.get('s250',0), r.get('l250',0),
                        r.get('s400',0), r.get('l400',0),
                        f"{r['perc']}%", str(r.get('penalty','—')),
                        f"{int(r.get('bonus',0)):,} so'm".replace(',',' ') if r.get('bonus',0) else '—',
                        f"{int(r['gross']):,} so'm".replace(',',' ')]
            else:
                vals = [r.get('group_name', r['branch']), r['branch'],
                        r.get('students',0), r.get('lessons',0),
                        f"{r['perc']}%", str(r.get('penalty','—')),
                        f"{int(r.get('bonus',0)):,} so'm".replace(',',' ') if r.get('bonus',0) else '—',
                        f"{int(r['gross']):,} so'm".replace(',',' ')]
            ws0.append(vals)
            for ci in range(1, ncols + 1):
                c = ws0.cell(row=cur, column=ci)
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                if ci == 1:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.font = Font(bold=True, size=10)

        write_summary_rows(ws0, ws0.max_row, total_gross, tax, net, ncols)

        for ci, w in enumerate(col_w, 1):
            ws0.column_dimensions[get_column_letter(ci)].width = w

        # ═══════════════════════════════════════════════
        # 2. HAR BIR FILIAL UCHUN ALOHIDA SHEET
        # ═══════════════════════════════════════════════
        for r in results:
            br_name = r['branch']
            grp_name = r.get('group_name', br_name)
            safe_title = grp_name[:31].replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
            ws = wb.create_sheet(title=safe_title)
            ws.row_dimensions[1].height = 28

            # Sarlavha
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            t = ws.cell(row=1, column=1, value=f"{grp_name}  ({br_name})  —  {teacher_name}  ({specialty})  {month_str}")
            t.font      = Font(bold=True, size=13, color="FFFFFF")
            t.fill      = PatternFill(start_color=COLOR_HDR, end_color=COLOR_HDR, fill_type="solid")
            t.alignment = Alignment(horizontal="center", vertical="center")
            t.border    = border

            # Ustun nomlari
            for ci, h in enumerate(headers, 1):
                hdr_cell(ws, 2, ci, h)

            # Faqat shu filial qatori
            if is_it:
                vals = [r.get('group_name', r['branch']), r['branch'],
                        r.get('s250',0), r.get('l250',0),
                        r.get('s400',0), r.get('l400',0),
                        f"{r['perc']}%", str(r.get('penalty','—')),
                        f"{int(r.get('bonus',0)):,} so'm".replace(',',' ') if r.get('bonus',0) else '—',
                        f"{int(r['gross']):,} so'm".replace(',',' ')]
            else:
                vals = [r.get('group_name', r['branch']), r['branch'],
                        r.get('students',0), r.get('lessons',0),
                        f"{r['perc']}%", str(r.get('penalty','—')),
                        f"{int(r.get('bonus',0)):,} so'm".replace(',',' ') if r.get('bonus',0) else '—',
                        f"{int(r['gross']):,} so'm".replace(',',' ')]
            ws.append(vals)
            for ci in range(1, ncols + 1):
                c = ws.cell(row=3, column=ci)
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = PatternFill(start_color=COLOR_ROW1, end_color=COLOR_ROW1, fill_type="solid")
                if ci == 1:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.font = Font(bold=True, size=10)

            # Faqat shu filial uchun gross/tax/net
            br_gross = float(r['gross'])
            br_tax   = br_gross * 0.075
            br_net   = br_gross - br_tax
            write_summary_rows(ws, ws.max_row, br_gross, br_tax, br_net, ncols)

            for ci, w in enumerate(col_w, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"Oylik_{teacher_name}_{month_str}.xlsx"
        return web.Response(
            body=buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'}
        )
    except Exception as e:
        logging.error(f"salary_excel error: {e}", exc_info=True)
        return web.Response(status=500, text=str(e))

async def admin_api_monthly_excel(request):
    """Oylik Excel hisobot - oy tanlash bilan"""
    try:
        year = int(request.rel_url.query.get('year', datetime.now(UZB_TZ).year))
        month = int(request.rel_url.query.get('month', datetime.now(UZB_TZ).month))
        buf = await create_monthly_excel(year, month)
        fname = f"Davomat_{year}_{str(month).zfill(2)}.xlsx"
        return web.Response(
            body=buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname}"'}
        )
    except Exception as e:
        logging.error(f"monthly_excel error: {e}")
        return web.Response(status=500, text=str(e))

async def admin_api_daily_pdf(request):
    """Kunlik PDF - sana tanlash bilan"""
    try:
        from datetime import date as _date
        date_str = request.rel_url.query.get('date', datetime.now(UZB_TZ).strftime('%Y-%m-%d'))
        d = _date.fromisoformat(date_str)
        buf = await get_combined_report_pdf(d)
        return web.Response(
            body=buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="davomat_{date_str}.xlsx"'}
        )
    except Exception as e:
        logging.error(f"daily_pdf error: {e}")
        return web.Response(status=500, text=str(e))

async def admin_api_branch_groups(request):
    """Filial bo'yicha guruhlar va ularning o'quvchilar davomati"""
    import json as _json
    branch = request.rel_url.query.get('branch', '')
    month_str = request.rel_url.query.get('month', datetime.now(UZB_TZ).strftime('%Y-%m'))
    if not branch:
        return web.Response(text=_json.dumps({'ok': False, 'error': 'branch kerak'}), content_type='application/json')
    branch_groups = [{'id': gid, 'group_name': gdata['group_name'], 'lesson_type': gdata['lesson_type'],
                      'teacher_id': gdata.get('teacher_id'), 'teacher_name': user_names.get(gdata.get('teacher_id'),'—'),
                      'day_times': gdata.get('day_times',{})}
                     for gid, gdata in groups.items() if gdata.get('branch') == branch]
    return web.Response(text=_json.dumps({'ok': True, 'groups': branch_groups}, ensure_ascii=False, default=str),
                        content_type='application/json', charset='utf-8')

async def admin_api_schedule_view(request):
    """Barcha faol dars jadvallarini qaytarish"""
    import json as _json
    branch = request.rel_url.query.get('branch', '')
    result = []
    for gid, gdata in groups.items():
        if branch and gdata.get('branch') != branch:
            continue
        teacher_name = user_names.get(gdata.get('teacher_id'), '—')
        result.append({
            'id': gid,
            'group_name': gdata.get('group_name',''),
            'branch': gdata.get('branch',''),
            'lesson_type': gdata.get('lesson_type',''),
            'teacher_name': teacher_name,
            'day_times': gdata.get('day_times',{}),
            'days': gdata.get('days',[]),
            'time_text': gdata.get('time_text',''),
        })
    result.sort(key=lambda x: (x['branch'], x['group_name']))
    return web.Response(text=_json.dumps({'ok': True, 'schedules': result}, ensure_ascii=False, default=str),
                        content_type='application/json', charset='utf-8')





# HISOBOT API FUNKSIYALARI
async def admin_api_reports_attendance(request):
    """O'qituvchi davomati hisoboti Excel formatda - har o'qituvchi alohida sheet"""
    import json as _json
    from datetime import datetime as dt, timedelta
    
    # Admin tekshiruvi
    if not _check_admin_session(request):
        return web.json_response({'error': 'Unauthorized'}, status=401)
    
    start_date = request.rel_url.query.get('start_date')
    end_date = request.rel_url.query.get('end_date')
    
    if not start_date or not end_date:
        return web.json_response({'error': 'start_date va end_date kerak'}, status=400)
    
    try:
        from datetime import date as _date_cls
        start_date_obj = _date_cls.fromisoformat(start_date)
        end_date_obj = _date_cls.fromisoformat(end_date)
        
        # Ma'lumotlar
        async with db.pool.acquire() as conn:
            records = await conn.fetch("""
                SELECT a.*, u.full_name, u.specialty 
                FROM attendance a 
                JOIN users u ON a.user_id = u.user_id 
                WHERE a.date BETWEEN $1 AND $2 
                ORDER BY a.date, a.time
            """, start_date_obj, end_date_obj)
        
        all_groups = dict(groups)
        
        # O'qituvchilar bo'yicha guruhlash
        teachers = {}
        for record in records:
            teacher_name = record['full_name'] or 'Noma\'lum'
            if teacher_name not in teachers:
                teachers[teacher_name] = {
                    'specialty': record['specialty'] or '',
                    'records': [],
                    'total_late_minutes': 0
                }
            teachers[teacher_name]['records'].append(record)
        
        # Excel fayl yaratish
        wb = Workbook()
        
        # Border stili
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['#', 'Filial', 'Sana', 'Kun', 'Dars vaqti', 'Keldi', 'Kechikish (daq)', 'Holat']
        
        # Excel'da ruxsat etilmagan belgilarni tozalash
        invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
        
        for idx, (teacher_name, data) in enumerate(teachers.items()):
            # Sheet nomini tozalash
            clean_name = teacher_name
            for char in invalid_chars:
                clean_name = clean_name.replace(char, '')
            sheet_name = clean_name[:31] if len(clean_name) > 31 else clean_name
            if not sheet_name:
                sheet_name = f'O\'qituvchi {idx+1}'
            
            if idx == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            # Sarlavha qatorini yozish
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            # Ma'lumotlar
            row_idx = 2
            total_late_minutes = 0
            
            for record in data['records']:
                att_date = record['date']
                att_time = str(record['time'])
                branch = record['branch']
                
                # Hafta kuni
                day_names = ['Dushanba', 'Seshanba', 'Chorshanba', 'Payshanba', 'Juma', 'Shanba', 'Yakshanba']
                day_name = day_names[att_date.weekday()]
                
                # Dars vaqtini topish va kechikishni hisoblash
                lesson_time = ''
                late_minutes = 0
                status = ''
                
                for gid, gdata in all_groups.items():
                    if gdata.get('teacher_id') == record['user_id'] and gdata.get('branch') == branch:
                        day_times = gdata.get('day_times', {})
                        if day_name in day_times:
                            lesson_time = day_times[day_name]
                            att_parts = list(map(int, att_time.split(':')[:2]))
                            les_parts = list(map(int, lesson_time.split(':')))
                            att_minutes = att_parts[0] * 60 + att_parts[1]
                            les_minutes = les_parts[0] * 60 + les_parts[1]
                            diff = att_minutes - les_minutes
                            
                            if diff <= 0:
                                status = 'Vaqtida' if diff == 0 else f'{abs(diff)} daq erta'
                            else:
                                late_minutes = diff
                                total_late_minutes += diff
                                status = f'{diff} daq kech'
                            break
                
                if not lesson_time:
                    status = 'Dars topilmadi'
                
                # Ma'lumotlarni yozish
                ws.cell(row=row_idx, column=1, value=row_idx-1).border = thin_border
                ws.cell(row=row_idx, column=2, value=branch or '').border = thin_border
                ws.cell(row=row_idx, column=3, value=str(att_date)).border = thin_border
                ws.cell(row=row_idx, column=4, value=day_name).border = thin_border
                ws.cell(row=row_idx, column=5, value=lesson_time).border = thin_border
                ws.cell(row=row_idx, column=6, value=att_time).border = thin_border
                
                late_cell = ws.cell(row=row_idx, column=7, value=late_minutes)
                late_cell.border = thin_border
                if late_minutes > 0:
                    late_cell.font = Font(color="FF0000", bold=True)
                elif late_minutes == 0 and status and 'erta' not in status:
                    late_cell.font = Font(color="006600")
                
                status_cell = ws.cell(row=row_idx, column=8, value=status)
                status_cell.border = thin_border
                if 'kech' in status:
                    status_cell.font = Font(color="FF0000")
                elif 'erta' in status:
                    status_cell.font = Font(color="228B22")
                elif 'Vaqtida' in status:
                    status_cell.font = Font(color="0066CC")
                
                row_idx += 1
            
            # JAMI qatori
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="JAMI:").font = Font(bold=True, size=12)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            ws.cell(row=row_idx, column=7, value=total_late_minutes).font = Font(bold=True, color="FF0000", size=12)
            ws.cell(row=row_idx, column=8, value=f"Bu oy jami {total_late_minutes} daqiqa kech qoldi").font = Font(bold=True, size=11)
            
            # Chegarani qo'shish
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).border = thin_border
            
            # Ustun kengliklari
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 20
            
            # Muzlatish
            ws.freeze_panes = "A2"
        
        # Excel faylni bytes ga aylantirish
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return web.Response(
            body=output.read(),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="davomat_hisoboti_{start_date}_{end_date}.xlsx"'
            }
        )
        
    except Exception as e:
        logging.error(f"Attendance report error: {e}")
        return web.json_response({'error': str(e)}, status=500)

async def admin_api_reports_students(request):
    """O'quvchilar davomati hisoboti Excel formatda"""
    import json as _json
    
    # Admin tekshiruvi
    if not _check_admin_session(request):
        return web.json_response({'error': 'Unauthorized'}, status=401)
    
    month = request.rel_url.query.get('month')
    if not month:
        return web.json_response({'error': 'month parametri kerak'}, status=400)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "O'quvchilar Davomati Hisoboti"
        
        # Sarlavha
        headers = ['#', 'Guruh', 'O\'quvchi', 'Telefon', 'Kelgan kunlar', 'Kelmagan kunlar', 'Foiz']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ma'lumotlar
        async with db.pool.acquire() as conn:
            records = await conn.fetch("""
                SELECT gs.group_id, g.group_name, gs.student_name, gs.student_phone,
                       COUNT(sa.id) FILTER (WHERE sa.status = 'Kelgan') as present_count,
                       COUNT(sa.id) as total_count
                FROM group_students gs
                JOIN groups g ON gs.group_id = g.id
                LEFT JOIN student_attendance sa ON gs.group_id = sa.group_id 
                    AND gs.student_name = sa.student_name 
                    AND DATE_PART('year', sa.lesson_date) = $1
                    AND DATE_PART('month', sa.lesson_date) = $2
                GROUP BY gs.group_id, g.group_name, gs.student_name, gs.student_phone
                ORDER BY g.group_name, gs.student_name
            """, int(month.split('-')[0]), int(month.split('-')[1]))
        
        for row_idx, record in enumerate(records, 2):
            present = record['present_count'] or 0
            total = record['total_count'] or 0
            percentage = round((present / total * 100) if total > 0 else 0, 1)
            
            ws.cell(row=row_idx, column=1, value=row_idx-1)
            ws.cell(row=row_idx, column=2, value=record['group_name'])
            ws.cell(row=row_idx, column=3, value=record['student_name'])
            ws.cell(row=row_idx, column=4, value=record['student_phone'])
            ws.cell(row=row_idx, column=5, value=present)
            ws.cell(row=row_idx, column=6, value=total - present)
            ws.cell(row=row_idx, column=7, value=f"{percentage}%")
            
            # Har bir qator uchun moslikni sozlash
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="left", vertical="center")
        
        # Ustun kengliklarini avtomatik sozlash
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Qatilarni muzlatish (harakatlanmaydigan qatir)
        ws.freeze_panes = "A2"
        
        # Qo'shimcha ma'lumot qo'shish (statistika)
        if records:
            summary_row = len(records) + 4
            ws.cell(row=summary_row, column=1, value="JAMI:").font = Font(bold=True)
            ws.cell(row=summary_row, column=2, value=len(records)).font = Font(bold=True)
        
        # Excel faylni bytes ga aylantirish
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return web.Response(
            body=output.read(),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="oquvchilar_hisoboti_{month}.xlsx"'
            }
        )
        
    except Exception as e:
        return web.json_response({'error': str(e)}, status=500)

async def admin_api_reports_groups(request):
    """Guruhlar hisoboti Excel formatda"""
    import json as _json
    
    # Admin tekshiruvi
    if not _check_admin_session(request):
        return web.json_response({'error': 'Unauthorized'}, status=401)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Guruhlar Hisoboti"
        
        # Sarlavha
        headers = ['#', 'Guruh nomi', 'Filial', 'Fan turi', 'O\'qituvchi', 'O\'quvchilar soni', 'Dars kunlari', 'Dars vaqti']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ma'lumotlar
        row_idx = 2
        for group_id, group_data in groups.items():
            teacher_name = user_names.get(group_data.get('teacher_id'), 'Noma\'lum')
            students_count = len(group_students.get(group_id, []))
            days = ', '.join(group_data.get('days', []))
            day_times = group_data.get('day_times', {})
            times_str = ', '.join([f"{day}: {time}" for day, time in day_times.items()]) if day_times else ''
            
            ws.cell(row=row_idx, column=1, value=row_idx-1)
            ws.cell(row=row_idx, column=2, value=group_data.get('group_name', ''))
            ws.cell(row=row_idx, column=3, value=group_data.get('branch', ''))
            ws.cell(row=row_idx, column=4, value=group_data.get('lesson_type', ''))
            ws.cell(row=row_idx, column=5, value=teacher_name)
            ws.cell(row=row_idx, column=6, value=students_count)
            ws.cell(row=row_idx, column=7, value=days)
            ws.cell(row=row_idx, column=8, value=times_str)
            
            # Har bir qator uchun moslikni sozlash
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="left", vertical="center")
            
            row_idx += 1
        
        # Ustun kengliklarini avtomatik sozlash
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Qatilarni muzlatish (harakatlanmaydigan qatir)
        ws.freeze_panes = "A2"
        
        # Qo'shimcha ma'lumot qo'shish (statistika)
        if len(groups) > 0:
            summary_row = len(list(groups.items())) + 4
            ws.cell(row=summary_row, column=1, value="JAMI GURUHLAR:").font = Font(bold=True)
            ws.cell(row=summary_row, column=2, value=len(groups)).font = Font(bold=True)
            total_students = sum(len(group_students.get(gid, [])) for gid in groups)
            ws.cell(row=summary_row, column=6, value=f"JAMI O'QUVCHILAR: {total_students}").font = Font(bold=True)
        
        # Excel faylni bytes ga aylantirish
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return web.Response(
            body=output.read(),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="guruhlar_hisoboti_{datetime.now(UZB_TZ).strftime("%Y%m%d")}.xlsx"'
            }
        )
        
    except Exception as e:
        return web.json_response({'error': str(e)}, status=500)

async def admin_api_reports_payments(request):
    """To'lovlar hisoboti Excel formatda"""
    import json as _json
    
    # Admin tekshiruvi
    if not _check_admin_session(request):
        return web.json_response({'error': 'Unauthorized'}, status=401)
    
    month = request.rel_url.query.get('month')
    if not month:
        return web.json_response({'error': 'month parametri kerak'}, status=400)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "To'lovlar"
        
        # Sarlavha
        headers = ['#', 'Guruh', 'O\'quvchi', 'Telefon', 'To\'lov holati', 'Summa', 'Izoh']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Ma'lumotlar
        async with db.pool.acquire() as conn:
            records = await conn.fetch("""
                SELECT g.group_name, gs.student_name, gs.student_phone,
                       COALESCE(sp.paid, false) as paid,
                       COALESCE(sp.amount, 0) as amount,
                       sp.note
                FROM group_students gs
                JOIN groups g ON gs.group_id = g.id
                LEFT JOIN student_payments sp ON gs.group_id = sp.group_id 
                    AND gs.student_name = sp.student_name 
                    AND sp.month = $1
                ORDER BY g.group_name, gs.student_name
            """, month)
        
        for row_idx, record in enumerate(records, 2):
            status = "To'ladi" if record['paid'] else "To'lamadi"
            amount = record['amount'] if record['paid'] else 0
            
            ws.cell(row=row_idx, column=1, value=row_idx-1)
            ws.cell(row=row_idx, column=2, value=record['group_name'])
            ws.cell(row=row_idx, column=3, value=record['student_name'])
            ws.cell(row=row_idx, column=4, value=record['student_phone'])
            ws.cell(row=row_idx, column=5, value=status)
            ws.cell(row=row_idx, column=6, value=amount)
            ws.cell(row=row_idx, column=7, value=record['note'] or '')
        
        # Excel faylni bytes ga aylantirish
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return web.Response(
            body=output.read(),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="tolovlar_hisoboti_{month}.xlsx"'
            }
        )
        
    except Exception as e:
        return web.json_response({'error': str(e)}, status=500)

async def admin_api_reports_branches(request):
    """Filiallar statistikasi hisoboti Excel formatda"""
    import json as _json
    
    # Admin tekshiruvi
    if not _check_admin_session(request):
        return web.json_response({'error': 'Unauthorized'}, status=401)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Filiallar statistikasi"
        
        # Sarlavha
        headers = ['#', 'Filial', 'Guruhlar soni', 'O\'quvchilar soni', 'O\'qituvchilar soni', 'IT guruhlar', 'Koreys guruhlar']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Filiallar bo'yicha statistika hisoblash
        branch_stats = {}
        for loc in LOCATIONS:
            branch_name = loc['name']
            branch_stats[branch_name] = {
                'groups': 0,
                'students': 0,
                'teachers': set(),
                'it_groups': 0,
                'korean_groups': 0
            }
        
        # Guruhlar statistikasi
        for group_id, group_data in groups.items():
            branch = group_data.get('branch', '')
            if branch in branch_stats:
                branch_stats[branch]['groups'] += 1
                branch_stats[branch]['students'] += len(group_students.get(group_id, []))
                if group_data.get('teacher_id'):
                    branch_stats[branch]['teachers'].add(group_data['teacher_id'])
                
                lesson_type = group_data.get('lesson_type', '')
                if lesson_type == 'IT':
                    branch_stats[branch]['it_groups'] += 1
                elif lesson_type == 'Koreys tili':
                    branch_stats[branch]['korean_groups'] += 1
        
        # Ma'lumotlarni Excel ga yozish
        row_idx = 2
        for branch_name, stats in branch_stats.items():
            if stats['groups'] > 0:  # Faqat faol filiallar
                ws.cell(row=row_idx, column=1, value=row_idx-1)
                ws.cell(row=row_idx, column=2, value=branch_name)
                ws.cell(row=row_idx, column=3, value=stats['groups'])
                ws.cell(row=row_idx, column=4, value=stats['students'])
                ws.cell(row=row_idx, column=5, value=len(stats['teachers']))
                ws.cell(row=row_idx, column=6, value=stats['it_groups'])
                ws.cell(row=row_idx, column=7, value=stats['korean_groups'])
                
                row_idx += 1
        
        # Excel faylni bytes ga aylantirish
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return web.Response(
            body=output.read(),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="filiallar_hisoboti_{datetime.now(UZB_TZ).strftime("%Y%m%d")}.xlsx"'
            }
        )
        
    except Exception as e:
        return web.json_response({'error': str(e)}, status=500)

async def admin_api_reports_general(request):
    """Umumiy hisobot Excel formatda"""
    import json as _json
    
    # Admin tekshiruvi
    if not _check_admin_session(request):
        return web.json_response({'error': 'Unauthorized'}, status=401)
    
    report_type = request.rel_url.query.get('type', 'monthly')
    
    try:
        wb = Workbook()
        
        # 1. Umumiy statistika
        ws1 = wb.active
        ws1.title = "Umumiy statistika"
        
        # Asosiy ma'lumotlar
        total_users = len(user_ids)
        active_users = sum(1 for status in user_status.values() if status != 'blocked')
        total_groups = len(groups)
        total_students = sum(len(students) for students in group_students.values())
        
        stats_data = [
            ['Ko\'rsatkichlar', 'Qiymat'],
            ['Jami foydalanuvchilar', total_users],
            ['Faol foydalanuvchilar', active_users],
            ['Jami guruhlar', total_groups],
            ['Jami o\'quvchilar', total_students],
            ['IT o\'qituvchilari', sum(1 for spec in user_specialty.values() if spec == 'IT')],
            ['Koreys tili o\'qituvchilari', sum(1 for spec in user_specialty.values() if spec == 'Koreys tili')],
        ]
        
        for row_idx, (label, value) in enumerate(stats_data, 1):
            ws1.cell(row=row_idx, column=1, value=label)
            ws1.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                ws1.cell(row=row_idx, column=1).font = Font(bold=True)
                ws1.cell(row=row_idx, column=2).font = Font(bold=True)
        
        # 2. Guruhlar
        ws2 = wb.create_sheet("Guruhlar")
        headers = ['Guruh nomi', 'Filial', 'Fan turi', 'O\'qituvchi', 'O\'quvchilar']
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header)
            ws2.cell(row=1, column=col).font = Font(bold=True)
        
        row_idx = 2
        for group_id, group_data in groups.items():
            teacher_name = user_names.get(group_data.get('teacher_id'), 'Noma\'lum')
            students_count = len(group_students.get(group_id, []))
            
            ws2.cell(row=row_idx, column=1, value=group_data.get('group_name', ''))
            ws2.cell(row=row_idx, column=2, value=group_data.get('branch', ''))
            ws2.cell(row=row_idx, column=3, value=group_data.get('lesson_type', ''))
            ws2.cell(row=row_idx, column=4, value=teacher_name)
            ws2.cell(row=row_idx, column=5, value=students_count)
            
            row_idx += 1
        
        # Excel faylni bytes ga aylantirish
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return web.Response(
            body=output.read(),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="umumiy_hisobot_{report_type}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            }
        )
        
    except Exception as e:
        return web.json_response({'error': str(e)}, status=500)

async def admin_api_business_report(request):
    """Korean style business report - students, revenue, expenses, profit"""
    import json as _json
    try:
        month = request.query.get('month', datetime.now(UZB_TZ).strftime('%Y-%m'))
        
        # Get branches from groups
        branch_data = {}
        async with db.pool.acquire() as conn:
            # Get all groups with their branch and lesson type
            groups_data = await conn.fetch("""
                SELECT g.branch, g.lesson_type, g.id, g.group_name
                FROM groups g
                WHERE g.branch IS NOT NULL AND g.branch != ''
            """)
            
            # Get student counts by branch and type
            for g in groups_data:
                branch = g['branch']
                lesson_type = g['lesson_type'] or 'Boshqa'
                if branch not in branch_data:
                    branch_data[branch] = {'IT': 0, 'Koreys tili': 0, 'total': 0}
                
                # Count students in this group
                student_count = await conn.fetchval(
                    "SELECT COUNT(*) FROM group_students WHERE group_id=$1",
                    g['id']
                )
                student_count = student_count or 0
                
                if lesson_type in ['IT', 'Koreys tili']:
                    branch_data[branch][lesson_type] += student_count
                branch_data[branch]['total'] += student_count
        
        # Get payment data for this month
        payment_data = {'total': 0, 'paid': 0, 'unpaid': 0, 'amount': 0}
        async with db.pool.acquire() as conn:
            payments = await conn.fetch("""
                SELECT sp.paid, sp.amount, sp.group_id, g.branch, g.lesson_type
                FROM student_payments sp
                JOIN groups g ON sp.group_id = g.id
                WHERE sp.month = $1
            """, month)
            
            for p in payments:
                payment_data['total'] += 1
                if p['paid']:
                    payment_data['paid'] += 1
                    payment_data['amount'] += p['amount'] or 0
                else:
                    payment_data['unpaid'] += 1
        
        # Get salary data for this month
        salary_data = {'teacher_kr': 0, 'teacher_it': 0, 'office': 0, 'total': 0}
        
        # Teacher attendance data for this month
        teacher_attendance = []
        
        # Calculate teacher salaries and collect attendance
        for uid in user_ids:
            spec = user_specialty.get(uid, '')
            status = user_status.get(uid, 'active')
            name = user_names.get(uid, '')
            if status == 'deleted' or status == 'blocked':
                continue
            
            # Count attendance for this month
            month_atts = [a for a in daily_attendance_log 
                         if a[0] == uid and a[2].startswith(month)]
            month_att_count = len(month_atts)
            
            # Calculate late minutes for this month
            total_late_minutes = 0
            for a in month_atts:
                # a = (uid, branch, date, time, late_minutes)
                if len(a) > 4:
                    total_late_minutes += a[4] or 0
            
            if spec in ['Koreys tili', 'IT', 'Ofis xodimi']:
                teacher_attendance.append({
                    'name': name,
                    'specialty': spec,
                    'attendance_count': month_att_count,
                    'late_minutes': total_late_minutes
                })
            
            if spec == 'Koreys tili':
                salary_data['teacher_kr'] += 1800000
            elif spec == 'IT':
                salary_data['teacher_it'] += 1500000
            elif spec == 'Ofis xodimi':
                salary_data['office'] += 1000000
        
        # Get manual expenses from database
        manual_expenses = {'rent': 0, 'utilities': 0, 'accounting': 0, 'marketing': 0, 'other': 0}
        try:
            async with db.pool.acquire() as conn:
                expenses = await conn.fetch(
                    "SELECT expense_type, amount FROM business_expenses WHERE month=$1",
                    month
                )
                for e in expenses:
                    if e['expense_type'] in manual_expenses:
                        manual_expenses[e['expense_type']] = e['amount'] or 0
        except:
            pass
        
        salary_data['total'] = salary_data['teacher_kr'] + salary_data['teacher_it'] + salary_data['office']
        
        # Calculate totals
        total_revenue = payment_data['amount']
        total_expenses = salary_data['total'] + sum(manual_expenses.values())
        operating_profit = total_revenue - total_expenses
        
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'month': month,
                'branches': branch_data,
                'payments': payment_data,
                'salaries': salary_data,
                'teacher_attendance': teacher_attendance,
                'manual_expenses': manual_expenses,
                'summary': {
                    'total_revenue': total_revenue,
                    'total_expenses': total_expenses,
                    'operating_profit': operating_profit
                }
            }, ensure_ascii=False, default=str),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        logging.error(f"business_report error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_business_expenses_save(request):
    """Save manual expenses for business report"""
    import json as _json
    try:
        data = await request.json()
        month = data.get('month', datetime.now(UZB_TZ).strftime('%Y-%m'))
        expenses = data.get('expenses', {})
        
        async with db.pool.acquire() as conn:
            for expense_type, amount in expenses.items():
                await conn.execute("""
                    INSERT INTO business_expenses (month, expense_type, amount)
                    VALUES ($1, $2, $3)
                    ON CONFLICT (month, expense_type) 
                    DO UPDATE SET amount = EXCLUDED.amount
                """, month, expense_type, int(amount))
        
        return web.Response(
            text=_json.dumps({'ok': True}),
            content_type='application/json'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_salary_configs_get(request):
    """Get salary configurations"""
    import json as _json
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM salary_configs ORDER BY category, bin_key")
            
            # If no configs exist, seed with defaults
            if not rows:
                salary_defaults = [
                    ('soeup', '수습', 'bin_1', '1호봉', 7500000),
                    ('sawon', '사원', 'bin_1', '1호봉', 8500000),
                    ('sawon', '사원', 'bin_2', '2호봉', 9500000),
                    ('sawon', '사원', 'bin_3', '3호봉', 10500000),
                    ('daeri', '대리', 'bin_1', '1호봉', 11500000),
                    ('daeri', '대리', 'bin_2', '2호봉', 13000000),
                    ('daeri', '대리', 'bin_3', '3호봉', 14500000),
                    ('gwallija', '관리자', 'bin_1', '1호봉', 16000000),
                    ('gwallija', '관리자', 'bin_2', '2호봉', 17500000),
                    ('gwallija', '관리자', 'bin_3', '3호봉', 19000000),
                ]
                for cat, cat_kr, bin_key, bin_name, amount in salary_defaults:
                    await conn.execute("""
                        INSERT INTO salary_configs (category, category_kr, bin_key, bin_name, amount)
                        VALUES ($1, $2, $3, $4, $5)
                    """, cat, cat_kr, bin_key, bin_name, amount)
                rows = await conn.fetch("SELECT * FROM salary_configs ORDER BY category, bin_key")
        
        configs = {}
        for row in rows:
            cat = row['category']
            if cat not in configs:
                configs[cat] = {
                    'category': cat,
                    'category_kr': row['category_kr'],
                    'bins': {}
                }
            configs[cat]['bins'][row['bin_key']] = {
                'name': row['bin_name'],
                'amount': row['amount']
            }
        
        return web.Response(
            text=_json.dumps({'ok': True, 'configs': list(configs.values())}, ensure_ascii=False),
            content_type='application/json'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_salary_configs_save(request):
    """Save salary configurations"""
    import json as _json
    try:
        data = await request.json()
        configs = data.get('configs', [])
        reset = data.get('reset', False)
        
        # If reset requested, delete all and reseed
        if reset:
            async with db.pool.acquire() as conn:
                await conn.execute("DELETE FROM salary_configs")
                salary_defaults = [
                    ('soeup', '수습', 'bin_1', '1호봉', 7500000),
                    ('sawon', '사원', 'bin_1', '1호봉', 8500000),
                    ('sawon', '사원', 'bin_2', '2호봉', 9500000),
                    ('sawon', '사원', 'bin_3', '3호봉', 10500000),
                    ('daeri', '대리', 'bin_1', '1호봉', 11500000),
                    ('daeri', '대리', 'bin_2', '2호봉', 13000000),
                    ('daeri', '대리', 'bin_3', '3호봉', 14500000),
                    ('gwallija', '관리자', 'bin_1', '1호봉', 16000000),
                    ('gwallija', '관리자', 'bin_2', '2호봉', 17500000),
                    ('gwallija', '관리자', 'bin_3', '3호봉', 19000000),
                ]
                for cat, cat_kr, bin_key, bin_name, amount in salary_defaults:
                    await conn.execute("""
                        INSERT INTO salary_configs (category, category_kr, bin_key, bin_name, amount)
                        VALUES ($1, $2, $3, $4, $5)
                    """, cat, cat_kr, bin_key, bin_name, amount)
        
        async with db.pool.acquire() as conn:
            for config in configs:
                category = config.get('category', '')
                category_kr = config.get('category_kr', '')
                bins = config.get('bins', {})
                
                for bin_key, bin_data in bins.items():
                    bin_name = bin_data.get('name', '')
                    amount = int(bin_data.get('amount', 0))
                    
                    await conn.execute("""
                        INSERT INTO salary_configs (category, category_kr, bin_key, bin_name, amount)
                        VALUES ($1, $2, $3, $4, $5)
                        ON CONFLICT (category, bin_key) 
                        DO UPDATE SET category_kr = EXCLUDED.category_kr, bin_name = EXCLUDED.bin_name, amount = EXCLUDED.amount, updated_at = NOW()
                    """, category, category_kr, bin_key, bin_name, amount)
        
        return web.Response(
            text=_json.dumps({'ok': True}),
            content_type='application/json'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def _cache_all_photos():
    """Barcha foydalanuvchilar Telegram rasmini background da cache qilish"""
    import asyncio as _as
    await _as.sleep(15)
    for uid in list(user_ids):
        try:
            photos = await bot.get_user_profile_photos(uid, limit=1)
            if photos.total_count > 0:
                file_id = photos.photos[0][-1].file_id
                file_obj = await bot.get_file(file_id)
                user_photo_cache[uid] = f"https://api.telegram.org/file/bot{TOKEN}/{file_obj.file_path}"
        except Exception:
            pass
        await _as.sleep(0.3)
    logging.info(f"Photo cache: {len(user_photo_cache)} ta")

async def miniapp_get_profile_photo(request):
    """Foydalanuvchi Telegram profil rasmini URL qaytarish"""
    import json as _json
    try:
        uid = int(request.query.get('user_id', 0))
        if not uid:
            return web.Response(text=_json.dumps({'ok': False, 'url': None}), content_type='application/json')
        try:
            photos = await bot.get_user_profile_photos(uid, limit=1)
            if photos.total_count > 0:
                file_id = photos.photos[0][-1].file_id
                file = await bot.get_file(file_id)
                url = f"https://api.telegram.org/file/bot{TOKEN}/{file.file_path}"
                return web.Response(
                    text=_json.dumps({'ok': True, 'url': url}),
                    content_type='application/json'
                )
        except Exception as e:
            logging.error(f"get profile photo error: {e}")
        return web.Response(text=_json.dumps({'ok': True, 'url': None}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'url': None}), content_type='application/json')

async def miniapp_update_profile(request):
    """Profil yangilash - ism va mutaxassislik"""
    import json as _json
    try:
        data = await request.json()
        uid = int(data['user_id'])
        new_name = data.get('name', '').strip()
        new_specialty = data.get('specialty', '').strip()
        logging.info(f"miniapp_update_profile: uid={uid}, name={new_name}, specialty={new_specialty}")
        
        if new_name and len(new_name) >= 3:
            user_names[uid] = new_name
            logging.info(f"miniapp_update_profile: updated user_names[{uid}] = {new_name}")
        if new_specialty in ['IT', 'Koreys tili', 'Ofis xodimi']:
            user_specialty[uid] = new_specialty
            logging.info(f"miniapp_update_profile: updated user_specialty[{uid}] = {new_specialty}")
        
        try:
            user = await db.get_user(uid)
            if user:
                await db.save_user(
                    user_id=uid,
                    full_name=user_names.get(uid, user['full_name']),
                    specialty=user_specialty.get(uid, user['specialty']),
                    language=user.get('language', 'uz')
                )
                logging.info(f"miniapp_update_profile: saved to database")
        except Exception as e:
            logging.error(f"profile update db error: {e}")
        
        return web.Response(
            text=_json.dumps({'ok': True, 'name': user_names.get(uid,''), 'specialty': user_specialty.get(uid,'')}),
            content_type='application/json'
        )
    except Exception as e:
        logging.error(f"miniapp_update_profile error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_change_lang(request):
    """Til o'zgartirish"""
    import json as _json
    try:
        data = await request.json()
        uid = int(data['user_id'])
        lang = data.get('lang', 'uz')
        if lang not in ['uz', 'ru']:
            lang = 'uz'
        user_languages[uid] = lang
        try:
            user = await db.get_user(uid)
            if user:
                await db.save_user(
                    user_id=uid,
                    full_name=user_names.get(uid, user['full_name']),
                    specialty=user_specialty.get(uid, user['specialty']),
                    language=lang
                )
        except Exception as e:
            logging.error(f"lang change db error: {e}")
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_teacher_page(request):
    """O'qituvchi Mini App sahifasi"""
    try:
        with open('teacher_app.html', 'r', encoding='utf-8') as f:
            html = f.read()
        return web.Response(text=html, content_type='text/html', charset='utf-8')
    except FileNotFoundError:
        return web.Response(text='<h1>teacher_app.html topilmadi</h1>', content_type='text/html')

async def miniapp_teacher_data(request):
    """O'qituvchi asosiy ma'lumotlari - statistika, jadval, guruhlar"""
    import json as _json
    try:
        uid = int(request.query.get('user_id', 0))
        if not uid:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'user_id kerak'}), content_type='application/json')
        
        now_uzb = datetime.now(UZB_TZ)
        today = now_uzb.strftime('%Y-%m-%d')
        this_month = now_uzb.strftime('%Y-%m')
        today_day = WEEKDAYS_UZ[now_uzb.weekday()]
        
        # Statistika
        month_att = sum(1 for k in daily_attendance_log if k[0]==uid and k[2].startswith(this_month))
        total_att = sum(1 for k in daily_attendance_log if k[0]==uid)
        today_att = [(k[1], k[3]) for k in daily_attendance_log if k[0]==uid and k[2]==today]
        
        # Guruhlar
        my_groups = []
        for gid, gdata in groups.items():
            if gdata.get('teacher_id') != uid:
                continue
            studs = group_students.get(gid, [])
            is_today = today_day in gdata.get('days', [])
            my_groups.append({
                'id': gid,
                'group_name': gdata.get('group_name',''),
                'branch': gdata.get('branch',''),
                'lesson_type': gdata.get('lesson_type',''),
                'days': gdata.get('days',[]),
                'day_times': gdata.get('day_times',{}),
                'is_today': is_today,
                'student_count': len(studs),
            })
        my_groups.sort(key=lambda g:(0 if g['is_today'] else 1, g['group_name']))
        
        # Filiallar
        brs = [{'name':b['name'],'lat':b.get('lat',0),'lon':b.get('lon',0)} for b in LOCATIONS]
        
        # Profil
        name = user_names.get(uid,'')
        specialty = user_specialty.get(uid,'')
        status = user_status.get(uid,'active')
        language = user_languages.get(uid,'uz')
        logging.info(f"miniapp_teacher_data: uid={uid}, name='{name}', specialty='{specialty}', in_user_ids={uid in user_ids}")
        
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'user': {'id': uid, 'name': name, 'specialty': specialty, 'status': status, 'language': language},
                'stats': {
                    'today': len(today_att),
                    'today_branch': today_att[0][0] if today_att else '',
                    'today_time': today_att[0][1] if today_att else '',
                    'month': month_att,
                    'total': total_att,
                    'groups': len(my_groups),
                },
                'groups': my_groups,
                'branches': brs,
                'today_day': today_day,
                'today_date': today,
            }, ensure_ascii=False, default=str),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        logging.error(f"miniapp_teacher_data error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_group_students(request):
    """Guruh o'quvchilari + to'lov ma'lumotlari"""
    import json as _json
    try:
        gid = int(request.query.get('group_id', 0))
        month = request.query.get('month', datetime.now(UZB_TZ).strftime('%Y-%m'))
        if not gid or gid not in groups:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Guruh topilmadi'}), content_type='application/json')
        studs = group_students.get(gid, [])
        # To'lov ma'lumotlari
        payments = {}
        try:
            async with db.pool.acquire() as conn:
                rows = await conn.fetch(
                    "SELECT student_name, paid, amount, note FROM student_payments WHERE group_id=$1 AND month=$2",
                    gid, month
                )
                for r in rows:
                    payments[r['student_name']] = {'paid': r['paid'], 'amount': r['amount'], 'note': r['note'] or ''}
        except Exception as e:
            logging.error(f"payments fetch error: {e}")
        
        students_data = []
        for i, s in enumerate(studs):
            pay = payments.get(s['name'], {'paid': False, 'amount': 0, 'note': ''})
            students_data.append({
                'index': i,
                'id': s.get('id', i),
                'name': s['name'],
                'phone': s['phone'],
                'paid': pay['paid'],
                'amount': pay['amount'],
                'note': pay['note'],
            })
        
        gdata = groups[gid]
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'group_name': gdata.get('group_name',''),
                'branch': gdata.get('branch',''),
                'lesson_type': gdata.get('lesson_type',''),
                'students': students_data,
                'month': month,
                'paid_count': sum(1 for s in students_data if s['paid']),
                'total_amount': sum(s['amount'] for s in students_data if s['paid']),
            }, ensure_ascii=False),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_save_payment(request):
    """O'quvchi to'lovini saqlash"""
    import json as _json
    try:
        data = await request.json()
        gid = int(data['group_id'])
        student_name = data['student_name']
        student_phone = data.get('student_phone', '')
        month = data.get('month', datetime.now(UZB_TZ).strftime('%Y-%m'))
        paid = bool(data.get('paid', False))
        amount = int(data.get('amount', 0))
        note = data.get('note', '')

        async with db.pool.acquire() as conn:
            # Maktab to'lovi bo'lsa — barcha o'quvchilarni ham to'landi qilamiz
            if student_name == '__school__' and note == 'school':
                # Avval maktab to'lovini saqlaymiz
                await conn.execute("""
                    INSERT INTO student_payments (group_id, student_name, student_phone, month, paid, amount, note, updated_at)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,NOW())
                    ON CONFLICT (group_id, student_name, month)
                    DO UPDATE SET paid=$5, amount=$6, note=$7, updated_at=NOW()
                """, gid, student_name, student_phone, month, True, amount, note)
                # Barcha o'quvchilarni to'landi qilamiz (summasiz, note='school_auto')
                studs = group_students.get(gid, [])
                for s in studs:
                    await conn.execute("""
                        INSERT INTO student_payments (group_id, student_name, student_phone, month, paid, amount, note, updated_at)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,NOW())
                        ON CONFLICT (group_id, student_name, month)
                        DO UPDATE SET paid=$5, note=$7, updated_at=NOW()
                    """, gid, s['name'], s['phone'], month, True, 0, 'school_auto')
            else:
                await conn.execute("""
                    INSERT INTO student_payments (group_id, student_name, student_phone, month, paid, amount, note, updated_at)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,NOW())
                    ON CONFLICT (group_id, student_name, month)
                    DO UPDATE SET paid=$5, amount=$6, note=$7, updated_at=NOW()
                """, gid, student_name, student_phone, month, paid, amount, note)

        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        logging.error(f"save_payment error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_add_student(request):
    """Guruhga o'quvchi qo'shish (teacher mini app)"""
    import json as _json
    try:
        data = await request.json()
        gid = int(data['group_id'])
        name = data['name'].strip()
        phone = data['phone'].strip()
        if not name or not phone:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Ism va telefon kerak'}), content_type='application/json')
        async with db.pool.acquire() as conn:
            sid = await conn.fetchval(
                "INSERT INTO group_students(group_id, student_name, student_phone) VALUES($1,$2,$3) RETURNING id",
                gid, name, phone
            )
        if gid not in group_students:
            group_students[gid] = []
        group_students[gid].append({'id': sid, 'name': name, 'phone': phone})
        return web.Response(text=_json.dumps({'ok': True, 'id': sid}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_del_student(request):
    """Guruhdan o'quvchi o'chirish (teacher mini app)"""
    import json as _json
    try:
        data = await request.json()
        gid = int(data['group_id'])
        student_name = data['student_name']
        async with db.pool.acquire() as conn:
            await conn.execute(
                "DELETE FROM group_students WHERE group_id=$1 AND student_name=$2",
                gid, student_name
            )
        if gid in group_students:
            group_students[gid] = [s for s in group_students[gid] if s['name'] != student_name]
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_group_att_history(request):
    """Guruh o'quvchilar davomat tarixi (oylik)"""
    import json as _json
    try:
        gid = int(request.rel_url.query.get('group_id', 0))
        month = request.rel_url.query.get('month', datetime.now(UZB_TZ).strftime('%Y-%m'))
        async with db.pool.acquire() as conn:
            rows = await conn.fetch(
                """SELECT student_name, lesson_date, status FROM student_attendance
                   WHERE group_id=$1 AND TO_CHAR(lesson_date,'YYYY-MM')=$2
                   ORDER BY lesson_date, student_name""",
                gid, month
            )
        by_student = {}
        for r in rows:
            n = r['student_name']
            if n not in by_student: by_student[n] = []
            by_student[n].append({'date': str(r['lesson_date']), 'status': r['status']})
        return web.Response(
            text=_json.dumps({'ok': True, 'data': by_student}, ensure_ascii=False),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_page(request):
    """Mini App HTML sahifasi"""
    try:
        with open('miniapp.html', 'r', encoding='utf-8') as f:
            html = f.read()
        return web.Response(text=html, content_type='text/html', charset='utf-8')
    except FileNotFoundError:
        return web.Response(text='<h1>miniapp.html topilmadi</h1>', content_type='text/html')

async def miniapp_api_init(request):
    """Mini App init - foydalanuvchi va guruhlarini qaytarish"""
    import json as _json
    try:
        uid = int(request.rel_url.query.get('user_id', 0))
        if not uid:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'user_id kerak'}), content_type='application/json')
        
        now_uzb = datetime.now(UZB_TZ)
        today_day = WEEKDAYS_UZ[now_uzb.weekday()]
        today_str = now_uzb.date().isoformat()
        
        # O'qituvchi ma'lumoti
        name = user_names.get(uid, '')
        specialty = user_specialty.get(uid, '')
        status = user_status.get(uid, 'active')
        
        if status == 'blocked':
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Siz bloklangansiz'}), content_type='application/json')
        
        # Bu o'qituvchining barcha guruhlari
        my_groups = []
        for gid, gdata in groups.items():
            if gdata.get('teacher_id') != uid:
                continue
            studs = group_students.get(gid, [])
            is_today = today_day in gdata.get('days', [])
            today_time = gdata.get('day_times', {}).get(today_day, '')
            
            # Bugun davomat qilinganmi?
            already_done = False
            try:
                async with db.pool.acquire() as conn:
                    row = await conn.fetchval(
                        "SELECT COUNT(*) FROM student_attendance WHERE group_id=$1 AND lesson_date=$2",
                        gid, now_uzb.date()
                    )
                    already_done = (row or 0) > 0
            except:
                pass
            
            my_groups.append({
                'id': gid,
                'group_name': gdata.get('group_name', ''),
                'branch': gdata.get('branch', ''),
                'lesson_type': gdata.get('lesson_type', ''),
                'days': gdata.get('days', []),
                'day_times': gdata.get('day_times', {}),
                'today_time': today_time,
                'is_today': is_today,
                'student_count': len(studs),
                'already_done': already_done,
            })
        
        # Bugungi guruhlar avval
        my_groups.sort(key=lambda g: (0 if g['is_today'] else 1, g['group_name']))
        
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'user': {'id': uid, 'name': name, 'specialty': specialty},
                'groups': my_groups,
                'today_day': today_day,
                'today_date': today_str,
            }, ensure_ascii=False, default=str),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        logging.error(f"miniapp_init error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_api_students(request):
    """Guruh o'quvchilari va bugungi davomat holati"""
    import json as _json
    try:
        gid = int(request.rel_url.query.get('group_id', 0))
        if not gid or gid not in groups:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Guruh topilmadi'}), content_type='application/json')
        
        now_uzb = datetime.now(UZB_TZ)
        studs = group_students.get(gid, [])
        
        # Bugungi mavjud davomat
        existing = {}
        try:
            async with db.pool.acquire() as conn:
                rows = await conn.fetch(
                    "SELECT student_name, status FROM student_attendance WHERE group_id=$1 AND lesson_date=$2",
                    gid, now_uzb.date()
                )
                for row in rows:
                    existing[row['student_name']] = row['status']
        except:
            pass
        
        students_data = []
        for i, s in enumerate(studs):
            prev_status = existing.get(s['name'], None)
            students_data.append({
                'index': i,
                'name': s['name'],
                'phone': s['phone'],
                'prev_status': prev_status,  # 'Kelgan', 'Kelmagan', None
            })
        
        gdata = groups[gid]
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'group_name': gdata.get('group_name', ''),
                'branch': gdata.get('branch', ''),
                'lesson_type': gdata.get('lesson_type', ''),
                'students': students_data,
                'already_done': len(existing) > 0,
            }, ensure_ascii=False),
            content_type='application/json', charset='utf-8'
        )
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def miniapp_api_submit(request):
    """Davomat saqlash - Mini App dan"""
    import json as _json
    try:
        data = await request.json()
        uid = int(data['user_id'])
        gid = int(data['group_id'])
        present_indices = data.get('present', [])  # kelgan o'quvchilar indekslari
        
        if gid not in groups:
            return web.Response(text=_json.dumps({'ok': False, 'error': 'Guruh topilmadi'}), content_type='application/json')
        
        now_uzb = datetime.now(UZB_TZ)
        studs = group_students.get(gid, [])
        gdata = groups[gid]
        group_name = gdata.get('group_name', '')
        branch = gdata.get('branch', '')
        today_day = WEEKDAYS_UZ[now_uzb.weekday()]
        lesson_date = now_uzb.date()
        current_date = now_uzb.strftime('%d.%m.%Y')
        
        # DB ga saqlash
        async with db.pool.acquire() as conn:
            await conn.execute(
                "DELETE FROM student_attendance WHERE group_id=$1 AND lesson_date=$2",
                gid, lesson_date
            )
            for i, s in enumerate(studs):
                status = "Kelgan" if i in present_indices else "Kelmagan"
                await conn.execute(
                    "INSERT INTO student_attendance (group_id, student_name, student_phone, lesson_date, status) VALUES ($1,$2,$3,$4,$5)",
                    gid, s['name'], s['phone'], lesson_date, status
                )
        
        # Excel yangilash
        try:
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            if gid in group_attendance_files:
                buf_in = io.BytesIO(group_attendance_files[gid])
                wb = load_workbook(buf_in)
                ws = wb.active
                new_col = ws.max_column + 1
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Guruh Davomati"
                for ci, h in enumerate(['№', "O'quvchi", 'Telefon'], 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
                    c.alignment = Alignment(horizontal="center"); c.border = border
                for i, s in enumerate(studs):
                    ws.cell(row=i+2, column=1, value=i+1).border = border
                    ws.cell(row=i+2, column=2, value=s['name']).border = border
                    ws.cell(row=i+2, column=3, value=s['phone']).border = border
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 28
                ws.column_dimensions['C'].width = 16
                new_col = 4
            
            col_letter = get_column_letter(new_col)
            hc = ws.cell(row=1, column=new_col, value=f"{current_date}\n{today_day}")
            hc.font = Font(bold=True, color="FFFFFF"); hc.fill = header_fill
            hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); hc.border = border
            ws.column_dimensions[col_letter].width = 14; ws.row_dimensions[1].height = 30
            
            for i, s in enumerate(studs):
                status = "Kelgan" if i in present_indices else "Kelmagan"
                dc = ws.cell(row=i+2, column=new_col, value=status)
                dc.alignment = Alignment(horizontal="center"); dc.border = border
                dc.fill = green_fill if i in present_indices else red_fill
            
            buf_out = io.BytesIO(); wb.save(buf_out); buf_out.seek(0)
            file_bytes = buf_out.read()
            group_attendance_files[gid] = file_bytes
            
            async with db.pool.acquire() as conn:
                await conn.execute(
                    """INSERT INTO group_excel_files (group_id, file_data, updated_at) VALUES ($1,$2,NOW())
                       ON CONFLICT (group_id) DO UPDATE SET file_data=EXCLUDED.file_data, updated_at=NOW()""",
                    gid, file_bytes
                )
        except Exception as ex:
            logging.error(f"miniapp excel error: {ex}")
        
        # Admin guruhiga oddiy xabar (Excel siz)
        kelgan = [studs[i]['name'] for i in present_indices if i < len(studs)]
        kelmagan = [s['name'] for i, s in enumerate(studs) if i not in present_indices]
        teacher_name = user_names.get(uid, "Noma'lum")
        lesson_time = gdata.get('day_times', {}).get(today_day, gdata.get('time', '—'))
        
        msg = (
            f"📋 *Guruh Davomati*\n\n"
            f"👥 Guruh: *{group_name}*\n"
            f"🏢 Filial: {branch}\n"
            f"👨‍🏫 O'qituvchi: {teacher_name}\n"
            f"📅 {current_date} ({today_day}) | ⏰ {lesson_time}\n\n"
            f"✅ Keldi: *{len(kelgan)} ta*"
        )
        if kelgan:
            msg += "\n" + "\n".join(f"  • {n}" for n in kelgan)
        if kelmagan:
            msg += f"\n\n❌ Kelmadi: *{len(kelmagan)} ta*"
            msg += "\n" + "\n".join(f"  • {n}" for n in kelmagan[:8])
            if len(kelmagan) > 8:
                msg += f"\n  ...+{len(kelmagan)-8} ta"
        
        try:
            await bot.send_message(ADMIN_GROUP_ID, msg, parse_mode="Markdown")
        except Exception as ex:
            logging.error(f"miniapp admin xabar xato: {ex}")
        
        return web.Response(
            text=_json.dumps({
                'ok': True,
                'kelgan': len(kelgan),
                'kelmagan': len(kelmagan),
                'total': len(studs),
            }),
            content_type='application/json'
        )
    except Exception as e:
        logging.error(f"miniapp_submit error: {e}")
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

async def admin_api_group_delete(request):
    """Guruhni o'chirish"""
    import json as _json
    try:
        data = await request.json()
        gid = int(data['group_id'])
        async with db.pool.acquire() as conn:
            await conn.execute("DELETE FROM groups WHERE id = $1", gid)
        groups.pop(gid, None)
        group_students.pop(gid, None)
        return web.Response(text=_json.dumps({'ok': True}), content_type='application/json')
    except Exception as e:
        return web.Response(text=_json.dumps({'ok': False, 'error': str(e)}), content_type='application/json')

def _check_admin_session(request):
    return _check_admin_request(request)

async def admin_api_data(request):
    """Admin Mini App uchun JSON API — barcha ma'lumotlar"""
    import json as _json
    from datetime import date as _date, timedelta
    if not _check_admin_session(request):
        return web.Response(text=_json.dumps({'ok': False, 'error': 'Ruxsat yoq'}), status=401, content_type='application/json')
    try:
        now_uzb = datetime.now(UZB_TZ)
        today_str = now_uzb.strftime('%Y-%m-%d')
        this_month = now_uzb.strftime('%Y-%m')

        # Foydalanuvchilar (arxivdagilar ham)
        users_data = []
        skipped_users = []
        logging.info(f"admin_api_data: user_ids count = {len(user_ids)}")
        for uid in user_ids:
            name = user_names.get(uid, '')
            if not name:
                skipped_users.append({'uid': uid, 'name': name, 'reason': 'no name'})
                continue
            # Arxivlangan foydalanuvchini aniqlash
            is_archived = '[ARXIV]' in name
            clean_name = name.replace('[ARXIV]', '').strip()
            month_att = len([
                a for a in daily_attendance_log
                if a[0] == uid and a[2].startswith(this_month)
            ])
            users_data.append({
                'user_id': uid,
                'name': clean_name,
                'specialty': user_specialty.get(uid, ''),
                'language': user_languages.get(uid, 'uz'),
                'status': 'archived' if is_archived else user_status.get(uid, 'active'),
                'attendance_count': month_att,
                'photo_url': user_photo_cache.get(uid),
            })
        logging.info(f"admin_api_data: returning {len(users_data)} users, skipped {len(skipped_users)} users")
        if skipped_users:
            logging.info(f"admin_api_data: skipped users: {skipped_users}")
        users_data.sort(key=lambda x: x['name'])

        # Guruhlar
        groups_data = []
        for gid, gdata in groups.items():
            students = group_students.get(gid, [])
            tid = gdata.get('teacher_id')
            groups_data.append({
                'id': gid,
                'group_name': gdata.get('group_name', ''),
                'branch': gdata.get('branch', ''),
                'lesson_type': gdata.get('lesson_type', ''),
                'teacher_id': tid,
                'teacher_name': user_names.get(tid, '—') if tid else '—',
                'days': gdata.get('days', []),
                'day_times': gdata.get('day_times', {}),
                'students': students,
                'student_count': len(students),
            })

        # Bugungi davomat
        today_att = []
        for a in daily_attendance_log:
            if a[2] == today_str:
                today_att.append({
                    'user_id': a[0],
                    'branch': a[1],
                    'time': a[3] if len(a) > 3 else '',
                })

        # Haftalik statistika (oxirgi 7 kun)
        weekly = []
        weekday_names = ['Du', 'Se', 'Ch', 'Pa', 'Ju', 'Sh', 'Ya']
        for i in range(6, -1, -1):
            d = (now_uzb - timedelta(days=i)).strftime('%Y-%m-%d')
            count = len([a for a in daily_attendance_log if a[2] == d])
            dow = (now_uzb - timedelta(days=i)).weekday()
            weekly.append({'day': weekday_names[dow], 'count': count, 'date': d})

        # Live class data - dars jadvali va hozirgi darslar
        live_classes = {'classes_now': [], 'today_classes': [], 'teachers_status': [], 'upcoming_classes': []}
        try:
            current_time = now_uzb.time()
            current_hour = now_uzb.hour
            current_minute = now_uzb.minute
            current_day_name = now_uzb.strftime('%A')
            
            # Convert English day names to Uzbek
            day_map = {
                'Monday': 'Dushanba', 'Tuesday': 'Seshanba', 'Wednesday': 'Chorshanba',
                'Thursday': 'Payshanba', 'Friday': 'Juma', 'Saturday': 'Shanba', 'Sunday': 'Yakshanba'
            }
            uz_day = day_map.get(current_day_name, current_day_name)
            
            # Jami darslar (bugun uchun)
            all_classes = []
            for gid, gdata in groups.items():
                day_times = gdata.get('day_times', {})
                if uz_day in day_times:
                    time_str = day_times[uz_day]
                    teacher_id = gdata.get('teacher_id')
                    teacher_name = user_names.get(teacher_id, '—') if teacher_id else '—'
                    
                    # O'qituvchi davomat tekshirish
                    teacher_attended = False
                    teacher_attend_time = ''
                    for a in today_att:
                        if a['user_id'] == teacher_id:
                            teacher_attended = True
                            teacher_attend_time = a.get('time', '')
                            break
                    
                    class_info = {
                        'group_id': gid,
                        'group_name': gdata.get('group_name', ''),
                        'branch': gdata.get('branch', ''),
                        'lesson_type': gdata.get('lesson_type', ''),
                        'teacher_id': teacher_id,
                        'teacher_name': teacher_name,
                        'time': time_str,
                        'student_count': len(group_students.get(gid, [])),
                        'teacher_attended': teacher_attended,
                        'teacher_attend_time': teacher_attend_time
                    }
                    all_classes.append(class_info)
            
            # Vaqt bo'yicha tartiblash
            all_classes.sort(key=lambda x: x['time'])
            
            # Hozirgi darslar (hozir vaqt + - 30 daqiqa)
            for cls in all_classes:
                try:
                    cls_hour, cls_minute = map(int, cls['time'].split(':'))
                    cls_minutes = cls_hour * 60 + cls_minute
                    now_minutes = current_hour * 60 + current_minute
                    
                    # 30 daqiqa oldin yoki keyin
                    if abs(cls_minutes - now_minutes) <= 30:
                        live_classes['classes_now'].append(cls)
                except:
                    pass
            
            # Barcha bugungi darslar
            live_classes['today_classes'] = all_classes
            
            # O'qituvchilar holati
            teachers_status = {}
            for cls in all_classes:
                tid = cls['teacher_id']
                if tid and tid not in teachers_status:
                    teachers_status[tid] = {
                        'teacher_id': tid,
                        'teacher_name': cls['teacher_name'],
                        'specialty': user_specialty.get(tid, ''),
                        'classes_count': 0,
                        'attended_count': 0,
                        'classes': []
                    }
                if tid:
                    teachers_status[tid]['classes_count'] += 1
                    teachers_status[tid]['classes'].append({
                        'group_name': cls['group_name'],
                        'time': cls['time'],
                        'attended': cls['teacher_attended']
                    })
                    if cls['teacher_attended']:
                        teachers_status[tid]['attended_count'] += 1
            
            live_classes['teachers_status'] = list(teachers_status.values())
            
            # Keyingi darslar (keyingi 2 soat ichida)
            upcoming = []
            for cls in all_classes:
                try:
                    cls_hour, cls_minute = map(int, cls['time'].split(':'))
                    cls_minutes = cls_hour * 60 + cls_minute
                    now_minutes = current_hour * 60 + current_minute
                    
                    # Hozir va keyingi 120 daqiqa ichida
                    if cls_minutes > now_minutes and cls_minutes - now_minutes <= 120:
                        upcoming.append(cls)
                except:
                    pass
            
            live_classes['upcoming_classes'] = upcoming
            
        except Exception as e:
            logging.error(f"live_classes error: {e}")

        # Broadcast tarixi
        bc_hist = []
        try:
            async with db.pool.acquire() as conn:
                bc_rows = await conn.fetch(
                    "SELECT message_text, sent_count, failed_count, specialty, created_at FROM broadcast_history ORDER BY created_at DESC LIMIT 20"
                )
                for r in bc_rows:
                    bc_hist.append({
                        'text': (r['message_text'] or '')[:120],
                        'sent_count': r['sent_count'],
                        'failed_count': r['failed_count'],
                        'specialty': r['specialty'] or '',
                        'date': r['created_at'].strftime('%d.%m %H:%M') if r['created_at'] else '',
                    })
        except Exception as e:
            logging.error(f"broadcast_history fetch: {e}")

        # Dashboard: faol guruhlar o'quvchilar soni (total) vs to'laganlar (paid)
        pay_stats = {'paid': 0, 'total': 0, 'amount': 0}
        try:
            # Jami o'quvchilar soni - faol guruhlardagi
            total_studs = sum(len(group_students.get(gid, [])) for gid in groups)
            async with db.pool.acquire() as conn:
                ps = await conn.fetchrow(
                    "SELECT "
                    "COALESCE(SUM(CASE WHEN paid THEN 1 ELSE 0 END),0) as paid, "
                    "COALESCE(SUM(CASE WHEN paid THEN amount ELSE 0 END),0) as amount "
                    "FROM student_payments WHERE month=$1", this_month
                )
                if ps:
                    pay_stats = {
                        'paid': int(ps['paid'] or 0),
                        'total': total_studs,
                        'amount': int(ps['amount'] or 0)
                    }
        except Exception as e:
            logging.error(f"pay_stats: {e}")

        return web.Response(
            text=_json.dumps({
                'users': users_data,
                'groups': groups_data,
                'today_attendance': today_att,
                'weekly_stats': weekly,
                'branches': LOCATIONS,
                'broadcast_history': bc_hist,
                'pay_stats': pay_stats,
                'live_classes': live_classes,
                'meta': {
                    'total_users': len(users_data),
                    'total_groups': len(groups_data),
                    'today_present': len(today_att),
                    'generated_at': now_uzb.strftime('%Y-%m-%d %H:%M:%S'),
                }
            }, ensure_ascii=False, default=str),
            content_type='application/json',
            charset='utf-8'
        )
    except Exception as e:
        logging.error(f"admin_api_data error: {e}", exc_info=True)
        return web.Response(
            text=_json.dumps({'ok': False, 'error': str(e), 'users': [], 'groups': [], 'today_attendance': [], 'weekly_stats': [], 'branches': [], 'broadcast_history': [], 'meta': {}}),
            content_type='application/json', charset='utf-8'
        )

async def health_check(request):
    now_uzb = datetime.now(UZB_TZ)
    logging.info(f"Cron-job.org tomonidan tekshirildi: {now_uzb.strftime('%Y-%m-%d %H:%M:%S')}")
    return web.Response(text=f"Bot healthy - {now_uzb.strftime('%H:%M:%S')}", status=200)

async def start_web_server():
    """Bu funksiya webhook rejimida ishlatilmaydi — main() ichida to'g'ridan webhook server ishga tushadi"""
    pass

@dp.message(CommandStart())
async def cmd_start(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return

    # Arxivlangan foydalanuvchi qayta kirsa — arxivdan chiqaramiz
    current_name = user_names.get(user_id, '')
    if '[ARXIV]' in current_name:
        clean_name = current_name.replace('[ARXIV]', '').strip()
        user_names[user_id] = clean_name
        user_status[user_id] = 'active'
        try:
            async with db.pool.acquire() as conn:
                await conn.execute(
                    "UPDATE users SET full_name=$1, status='active' WHERE user_id=$2",
                    clean_name, user_id
                )
        except Exception as e:
            logging.error(f"Arxivdan chiqarish xato: {e}")

    if user_id not in user_names or not user_names.get(user_id):
        if user_id not in user_languages:
            keyboard = await language_selection_keyboard()
            await message.answer(
                "Iltimos, tilni tanlang:\nПожалуйста, выберите язык:\n언어를 선택하세요:",
                reply_markup=keyboard
            )
            return
        await state.set_state(Registration.waiting_for_name)
        await message.answer(get_text(user_id, 'ask_name'))
        return

    user_ids.add(user_id)
    keyboard = await main_keyboard(user_id)
    name = user_names.get(user_id, message.from_user.full_name)
    specialty = user_specialty.get(user_id, '')

    welcome_text = get_text(user_id, 'welcome', name=name)
    if specialty:
        specialty_display = get_specialty_display(specialty, user_languages.get(user_id, 'uz'))
        welcome_text += f"\n\n{specialty_display}"

    await message.answer(
        welcome_text,
        reply_markup=keyboard
    )

@dp.message(Registration.waiting_for_name)
async def process_name(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    full_name = message.text.strip()
    
    user_names[user_id] = full_name
    user_ids.add(user_id)
    logging.info(f"process_name: added user {user_id} ({full_name}) to user_ids. Total: {len(user_ids)}")
    
    lang = user_languages.get(user_id, 'uz')
    await db.save_user(user_id, full_name, None, lang)
    
    await state.update_data(name=full_name)
    
    keyboard = await specialty_keyboard(user_id)
    await state.set_state(Registration.waiting_for_specialty)
    await message.answer(
        get_text(user_id, 'ask_specialty'),
        reply_markup=keyboard
    )

@dp.message(Registration.waiting_for_specialty)
async def process_specialty(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text
    lang = user_languages.get(user_id, 'uz')
    
    if text == TRANSLATIONS[lang]['specialty_it']:
        specialty = 'IT'
    elif text == TRANSLATIONS[lang]['specialty_korean']:
        specialty = 'Koreys tili'
    elif text == TRANSLATIONS[lang]['specialty_office']:
        specialty = 'Ofis xodimi'
    else:
        await message.answer("❌ Noto'g'ri tanlov. Qaytadan urinib ko'ring.")
        return
    
    user_specialty[user_id] = specialty
    user_status[user_id] = 'active'
    
    await db.save_user(user_id, user_names[user_id], specialty, lang)
    
    await state.clear()
    
    keyboard = await main_keyboard(user_id)
    name = user_names.get(user_id)
    specialty_display = get_specialty_display(specialty, lang)
    
    welcome_text = get_text(user_id, 'welcome', name=name) + f"\n\n{specialty_display}"
    
    await message.answer(
        welcome_text,
        reply_markup=keyboard
    )

@dp.callback_query(F.data.startswith("lang_"))
async def set_initial_language(callback: types.CallbackQuery, state: FSMContext):
    try:
        user_id = callback.from_user.id
        lang = callback.data.split("_")[1]
        
        user_languages[user_id] = lang
        
        await callback.answer()
        await state.set_state(Registration.waiting_for_name)
        try:
            pass  # message.delete() webhook da ishlamas
        except:
            pass
        await bot.send_message(user_id, get_text(user_id, 'ask_name'))
    except Exception as e:
        logging.error(f"set_initial_language error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.message(F.text.in_({'\U0001F310 Til', '\U0001F310 Язык', '\U0001F310 언어'}))
async def change_language(message: types.Message):
    user_id = message.from_user.id
    keyboard = await language_selection_keyboard()
    lang = user_languages.get(user_id, 'uz')
    if lang == 'ru':
        await message.answer("Выберите язык / Tilni tanlang:", reply_markup=keyboard)
    else:
        await message.answer("Tilni tanlang / Выберите язык:", reply_markup=keyboard)

@dp.callback_query(F.data.startswith("change_lang_"))
async def set_changed_language(callback: types.CallbackQuery):
    try:
        user_id = callback.from_user.id
        lang = callback.data.split("_")[2]
        user_languages[user_id] = lang
        
        user = await db.get_user(user_id)
        if user:
            await db.save_user(user_id, user['full_name'], user['specialty'], lang)
        
        await callback.answer()
        keyboard = await main_keyboard(user_id)
        try:
            pass  # message.delete() webhook da ishlamas
        except:
            pass
        await bot.send_message(user_id, get_text(user_id, 'language_changed'), reply_markup=keyboard)
    except Exception as e:
        logging.error(f"set_changed_language error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.message(F.text == "👤 Mening profilim")
async def show_profile(message: types.Message):
    user_id = message.from_user.id
    lang = user_languages.get(user_id, 'uz')
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    name = user_names.get(user_id, "Noma'lum")
    specialty = user_specialty.get(user_id, "Ko'rsatilmagan")
    
    spec_display = get_specialty_display(specialty, lang)
    lang_display = "O'zbekcha"
    
    profile_text = get_text(user_id, 'profile_info', 
                           name=name, 
                           specialty=spec_display, 
                           lang=lang_display)
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text=get_text(user_id, 'edit_name'), callback_data="edit_name"))
    builder.row(InlineKeyboardButton(text=get_text(user_id, 'edit_my_specialty'), callback_data="edit_my_specialty"))
    builder.row(InlineKeyboardButton(text=get_text(user_id, 'back_to_menu'), callback_data="back_to_main"))
    
    await message.answer(
        profile_text,
        reply_markup=builder.as_markup()
    )

@dp.callback_query(F.data == "edit_my_specialty")
async def edit_my_specialty_start(callback: types.CallbackQuery):
    uid = callback.from_user.id
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="💻 IT", callback_data="save_spec_IT"))
    builder.row(InlineKeyboardButton(text="🇰🇷 Koreys tili", callback_data="save_spec_Koreys tili"))
    builder.row(InlineKeyboardButton(text="🏢 Ofis xodimi", callback_data="save_spec_Ofis xodimi"))
    builder.row(InlineKeyboardButton(text=get_text(uid, 'back_btn'), callback_data="back_to_profile_view"))
    
    await callback.message.edit_text(
        get_text(uid, 'select_new_spec'),
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("save_spec_"))
async def save_new_specialty(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    new_spec = callback.data.replace("save_spec_", "")
    
    user_specialty[user_id] = new_spec
    
    try:
        user = await db.get_user(user_id)
        if user:
            await db.save_user(
                user_id=user_id,
                full_name=user['full_name'],
                specialty=new_spec,
                language=user['language']
            )
        await callback.answer(get_text(user_id, 'spec_updated'), show_alert=True)
    except Exception as e:
        logging.error(f"Spec update error: {e}")
        await callback.answer("Xatolik yuz berdi")

    try:
        pass  # message.delete() webhook da ishlamas
    except:
        pass
    await show_profile(callback.message)

@dp.callback_query(F.data == "back_to_profile_view")
async def back_to_profile_view(callback: types.CallbackQuery):
    await callback.answer()
    try:
        pass  # message.delete() webhook da ishlamas
    except:
        pass
    await show_profile(callback.message)

@dp.callback_query(F.data == "edit_name")
async def edit_name_start(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    
    await callback.answer()
    
    await state.set_state(ProfileEdit.waiting_for_new_name)
    await callback.message.edit_text(
        get_text(user_id, 'enter_new_name')
    )

@dp.message(ProfileEdit.waiting_for_new_name)
async def process_new_name(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    new_name = message.text.strip()
    
    if len(new_name) < 3:
        await message.answer("❌ Ism juda qisqa. Iltimos, qaytadan kiriting:")
        return
    
    user_names[user_id] = new_name
    
    try:
        user = await db.get_user(user_id)
        if user:
            await db.save_user(
                user_id=user_id,
                full_name=new_name,
                specialty=user['specialty'],
                language=user['language']
            )
    except Exception as e:
        logging.error(f"PostgreSQL da ism yangilashda xatolik: {e}")
    
    await state.clear()
    
    await message.answer(
        get_text(user_id, 'name_updated')
    )
    
    await show_profile(message)

@dp.callback_query(F.data == "back_to_main")
async def back_to_main_menu(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    
    await callback.answer()
    keyboard = await main_keyboard(user_id)
    try:
        pass  # message.delete() webhook da ishlamas
    except:
        pass
    await bot.send_message(user_id, "🏠 Asosiy menyu", reply_markup=keyboard)

@dp.message(F.text == "📅 Dars jadvalim")
async def view_my_schedule_pdf(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    teacher_groups = [(gid, gdata) for gid, gdata in groups.items() if gdata['teacher_id'] == user_id]
    if not teacher_groups:
        await message.answer(get_text(user_id, 'no_schedules'))
        return
    
    try:
        buf = await create_schedule_pdf(user_id)
        clean_name = user_names.get(user_id, 'user').replace(' ', '_')
        await message.answer_document(
            types.BufferedInputFile(buf.getvalue(),
                                    filename=f"Dars_Jadvali_{clean_name}.xlsx"),
            caption=get_text(user_id, 'my_schedule')
        )
    except Exception as e:
        logging.error(f"view_my_schedule error: {e}")
        await message.answer("❌ Fayl yaratishda xatolik yuz berdi")

# ============================================================
# O'QITUVCHI GURUH STATISTIKASI VA O'QUVCHI CRUD
# ============================================================
@dp.message(F.text == "👥 Guruhlarim")
async def my_groups_handler(message: types.Message):
    user_id = message.from_user.id
    if user_status.get(user_id) == 'blocked':
        return
    teacher_groups = [(gid, gd) for gid, gd in groups.items() if gd.get('teacher_id') == user_id]
    if not teacher_groups:
        await message.answer("Sizga hali guruh biriktirilmagan.")
        return
    builder = InlineKeyboardBuilder()
    for gid, gd in teacher_groups:
        stud_count = len(group_students.get(gid, []))
        builder.row(InlineKeyboardButton(
            text=f"👥 {gd['group_name']} ({stud_count} o'q)",
            callback_data=f"tgrp_{gid}"
        ))
    await message.answer("👥 Guruhlaringiz:", reply_markup=builder.as_markup())

@dp.callback_query(F.data.startswith("tgrp_"),
                   ~F.data.startswith("tgrp_add_"),
                   ~F.data.startswith("tgrp_del"),
                   ~F.data.startswith("tgrp_back"))
async def teacher_group_detail(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    raw = callback.data.replace("tgrp_", "")
    try:
        gid = int(raw)
    except ValueError:
        await callback.answer()
        return
    grp = groups.get(gid)
    if not grp or grp.get('teacher_id') != user_id:
        await callback.answer("Ruxsat yo'q!", show_alert=True)
        return
    await callback.answer()
    students = group_students.get(gid, [])

    # Davomat statistikasi (oxirgi 30 kun)
    from datetime import date as d_date2, timedelta
    today = datetime.now(UZB_TZ).date()
    month_start = today.replace(day=1)
    
    async with db.pool.acquire() as conn:
        att_rows = await conn.fetch(
            "SELECT lesson_date, status, student_name FROM student_attendance "
            "WHERE group_id=$1 AND lesson_date >= $2 ORDER BY lesson_date DESC",
            gid, month_start
        )

    # O'quvchi bo'yicha statistika
    student_stats = {}
    for row in att_rows:
        name = row['student_name']
        if name not in student_stats:
            student_stats[name] = {'kelgan': 0, 'kelmagan': 0}
        if row['status'] == 'Kelgan':
            student_stats[name]['kelgan'] += 1
        else:
            student_stats[name]['kelmagan'] += 1

    day_times = grp.get('day_times', {})
    days_str = "\n".join([f"  📅 {d}: {t}" for d, t in day_times.items()]) if day_times else grp.get('time_text', '—')

    text = (
        f"👥 *{grp['group_name']}*\n"
        f"🏢 {grp['branch']}  |  📚 {grp['lesson_type']}\n\n"
        f"⏰ Dars vaqtlari:\n{days_str}\n\n"
        f"🧑‍🎓 O'quvchilar ({len(students)} ta) — bu oy statistikasi:\n"
    )
    for i, s in enumerate(students, 1):
        st = student_stats.get(s['name'], {})
        kelgan = st.get('kelgan', 0)
        kelmagan = st.get('kelmagan', 0)
        total = kelgan + kelmagan
        pct = f"{int(kelgan/total*100)}%" if total > 0 else "ma'lumot yo'q"
        text += f"{i}. {s['name']} — ✅{kelgan} ❌{kelmagan} ({pct})\n"

    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="➕ O'quvchi qo'shish", callback_data=f"tgrp_add_{gid}"),
        InlineKeyboardButton(text="➖ O'quvchi o'chirish", callback_data=f"tgrp_del_{gid}")
    )
    builder.row(InlineKeyboardButton(text="🔙 Guruhlarim", callback_data="tgrp_back"))
    await callback.message.edit_text(text, reply_markup=builder.as_markup(), parse_mode="Markdown")

@dp.callback_query(F.data == "tgrp_back")
async def teacher_groups_back(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    await callback.answer()
    teacher_grps = [(gid, gd) for gid, gd in groups.items() if gd.get('teacher_id') == user_id]
    if not teacher_grps:
        await callback.message.edit_text("Guruh topilmadi.")
        return
    builder = InlineKeyboardBuilder()
    for gid, gd in teacher_grps:
        stud_count = len(group_students.get(gid, []))
        builder.row(InlineKeyboardButton(
            text=f"👥 {gd['group_name']} ({stud_count} o'q)",
            callback_data=f"tgrp_{gid}"
        ))
    await callback.message.edit_text("👥 Guruhlaringiz:", reply_markup=builder.as_markup())

# --- O'QUVCHI QO'SHISH ---
@dp.callback_query(F.data.startswith("tgrp_add_"))
async def teacher_add_student_start(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    try:

        gid = int(callback.data.replace("tgrp_add_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    grp = groups.get(gid)
    if not grp or grp.get('teacher_id') != user_id:
        await callback.answer("Ruxsat yo'q!", show_alert=True)
        return
    await callback.answer()
    await state.update_data(tadd_gid=gid)
    await state.set_state(TeacherAddStudent.entering_name)
    await callback.message.edit_text(
        f"➕ *{grp['group_name']}* ga yangi o'quvchi qo'shish\n\nO'quvchi ismini kiriting:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"tgrp_{gid}")
        ]])
    )

@dp.message(TeacherAddStudent.entering_name)
async def teacher_add_student_name(message: types.Message, state: FSMContext):
    await state.update_data(tadd_name=message.text.strip())
    await state.set_state(TeacherAddStudent.entering_phone)
    await message.answer("📞 Telefon raqamini kiriting (+998XXXXXXXXX):")

@dp.message(TeacherAddStudent.entering_phone)
async def teacher_add_student_phone(message: types.Message, state: FSMContext):
    data = await state.get_data()
    gid = data['tadd_gid']
    name = data['tadd_name']
    phone = message.text.strip()
    grp = groups.get(gid, {})
    try:
        async with db.pool.acquire() as conn:
            await conn.execute(
                "INSERT INTO group_students (group_id, student_name, student_phone) VALUES ($1,$2,$3)",
                gid, name, phone
            )
        group_students[gid].append({'name': name, 'phone': phone})
        await state.clear()
        await message.answer(
            f"✅ *{name}* guruhga qo'shildi!\n"
            f"📞 {phone}\n"
            f"Guruh: {grp.get('group_name', '')}",
            parse_mode="Markdown"
        )
    except Exception as e:
        await state.clear()
        await message.answer(f"❌ Xatolik: {e}")

# --- O'QUVCHI O'CHIRISH ---
@dp.callback_query(F.data.startswith("tgrp_del_"))
async def teacher_del_student_list(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    try:

        gid = int(callback.data.replace("tgrp_del_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    grp = groups.get(gid)
    if not grp or grp.get('teacher_id') != user_id:
        await callback.answer("Ruxsat yo'q!", show_alert=True)
        return
    await callback.answer()
    students = group_students.get(gid, [])
    if not students:
        await callback.answer("Guruhda o'quvchi yo'q!", show_alert=True)
        return
    builder = InlineKeyboardBuilder()
    for i, s in enumerate(students):
        builder.row(InlineKeyboardButton(
            text=f"🗑 {s['name']}",
            callback_data=f"tgrp_delstd_{gid}_{i}"
        ))
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data=f"tgrp_{gid}"))
    await callback.message.edit_text(
        f"🗑 O'chirish uchun o'quvchini tanlang ({grp['group_name']}):",
        reply_markup=builder.as_markup()
    )

@dp.callback_query(F.data.startswith("tgrp_delstd_"))
async def teacher_del_student_confirm(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    parts = callback.data.replace("tgrp_delstd_", "").split("_")
    gid = int(parts[0])
    idx = int(parts[1])
    grp = groups.get(gid)
    if not grp or grp.get('teacher_id') != user_id:
        await callback.answer("Ruxsat yo'q!", show_alert=True)
        return
    students = group_students.get(gid, [])
    if idx >= len(students):
        await callback.answer("O'quvchi topilmadi!", show_alert=True)
        return
    student = students[idx]
    builder = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Ha, o'chirish", callback_data=f"tgrp_delconfirm_{gid}_{idx}"),
            InlineKeyboardButton(text="❌ Yo'q", callback_data=f"tgrp_del_{gid}")
        ]
    ])
    await callback.answer()
    await callback.message.edit_text(
        f"⚠️ *{student['name']}* ni guruhdan o'chirishni tasdiqlaysizmi?",
        reply_markup=builder, parse_mode="Markdown"
    )

@dp.callback_query(F.data.startswith("tgrp_delconfirm_"))
async def teacher_del_student_do(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    parts = callback.data.replace("tgrp_delconfirm_", "").split("_")
    gid = int(parts[0])
    idx = int(parts[1])
    grp = groups.get(gid)
    if not grp or grp.get('teacher_id') != user_id:
        await callback.answer("Ruxsat yo'q!", show_alert=True)
        return
    students = group_students.get(gid, [])
    if idx >= len(students):
        await callback.answer("O'quvchi topilmadi!", show_alert=True)
        return
    student = students[idx]
    try:
        async with db.pool.acquire() as conn:
            await conn.execute(
                "DELETE FROM group_students WHERE group_id=$1 AND student_name=$2 AND student_phone=$3",
                gid, student['name'], student['phone']
            )
        group_students[gid].pop(idx)
        await callback.answer(f"✅ {student['name']} o'chirildi!")
        # Guruh sahifasiga qaytish
        students_new = group_students.get(gid, [])
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="➕ O'quvchi qo'shish", callback_data=f"tgrp_add_{gid}"),
            InlineKeyboardButton(text="➖ O'quvchi o'chirish", callback_data=f"tgrp_del_{gid}")
        )
        builder.row(InlineKeyboardButton(text="🔙 Guruhlarim", callback_data="tgrp_back"))
        await callback.message.edit_text(
            f"✅ *{student['name']}* guruhdan o'chirildi.\n\n"
            f"Guruh: {grp['group_name']} ({len(students_new)} ta o'quvchi)",
            reply_markup=builder.as_markup(), parse_mode="Markdown"
        )
    except Exception as e:
        await callback.answer(f"Xatolik: {e}", show_alert=True)

# ============================================================

@dp.message(F.text == "📊 Mening statistikam")
async def my_stats(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    now_uzb = datetime.now(UZB_TZ)
    current_month = now_uzb.strftime("%Y-%m")
    
    user_attendances = defaultdict(list)
    
    for (uid, branch, date, time) in daily_attendance_log:
        if uid == user_id:
            user_attendances[branch].append((date, time))
    
    if not user_attendances:
        await message.answer(get_text(user_id, 'no_stats'))
        return
    
    month_names_uz = {
        "01": "Yanvar", "02": "Fevral", "03": "Mart", "04": "Aprel",
        "05": "May", "06": "Iyun", "07": "Iyul", "08": "Avgust",
        "09": "Sentabr", "10": "Oktabr", "11": "Noyabr", "12": "Dekabr"
    }
    
    lang = user_languages.get(user_id, 'uz')
    month_names = month_names_uz
    weekdays = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
    current_month_text = "(joriy oy)"
    
    text = get_text(user_id, 'stats') + "\n\n"
    
    for branch, date_time_list in user_attendances.items():
        text += f"🏢 {branch}\n"
        
        dates_by_month = defaultdict(list)
        for date_str, time_str in date_time_list:
            year_month = date_str[:7]
            dates_by_month[year_month].append((date_str, time_str))
        
        for year_month, month_data in sorted(dates_by_month.items(), reverse=True):
            year, month = year_month.split('-')
            month_name = month_names.get(month, month)
            
            month_display = f"{month_name} {year}"
            if year_month == current_month:
                month_display += f" {current_month_text}"
            
            text += f"   📅 {month_display}\n"
            
            for date_str, time_str in sorted(month_data, reverse=True):
                date_parts = date_str.split('-')
                year, month, day = date_parts
                
                date_obj = datetime(int(year), int(month), int(day), tzinfo=UZB_TZ)
                weekday = date_obj.weekday()
                weekday_name = weekdays[weekday]
                formatted_date = f"{int(day):02d}.{int(month):02d}.{year}"
                text += f"      • {formatted_date} ({weekday_name}) - ⏰ {time_str}\n"
            
            text += "\n"
        
        text += "\n"
    
    await message.answer(text)

@dp.message(F.text == "🏢 Filiallar")
async def show_branches(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    builder = InlineKeyboardBuilder()
    
    for loc in LOCATIONS:
        maps_link = get_yandex_maps_link(loc['lat'], loc['lon'])
        builder.row(
            InlineKeyboardButton(text=f"📍 {loc['name']}", url=maps_link)
        )
    
    await message.answer(
        "🏢 Mavjud filiallar (lokatsiya uchun bosing):",
        reply_markup=builder.as_markup()
    )

@dp.message(F.text.in_({'\u2753 Yordam', '\u2753 Помощь', '\u2753 도움말'}))
async def help_command(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    await message.answer(
        get_text(user_id, 'help')
    )

@dp.message(F.text.in_({'\U0001F3C6 Hafta topi', '\U0001F3C6 Топ недели', '\U0001F3C6 주간 TOP'}))
async def weekly_top(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    now_uzb = datetime.now(UZB_TZ)
    week_ago = now_uzb - timedelta(days=7)
    week_ago_str = week_ago.strftime("%Y-%m-%d")
    
    weekly_stats = defaultdict(int)
    
    for (uid, branch, date, time) in daily_attendance_log:
        if date >= week_ago_str:
            weekly_stats[uid] += 1
    
    if not weekly_stats:
        no_data_msg = "💭 Bu hafta hali davomat yo'q"
        
        await message.answer(no_data_msg)
        return
    
    top_users = sorted(weekly_stats.items(), key=lambda x: x[1], reverse=True)[:10]
    
    top_list = ""
    for i, (uid, count) in enumerate(top_users, 1):
        try:
            name = user_names.get(uid, f"ID: {uid}")
            specialty = user_specialty.get(uid, '')
            specialty_display = f" [{specialty}]" if specialty else ""
        except:
            name = f"Foydalanuvchi {uid}"
            specialty_display = ""
        
        medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
        top_list += f"{medal} {name}{specialty_display}: {count} marta\n"
    
    await message.answer(
        get_text(user_id, 'weekly_top', top_list=top_list)
    )

# --- BARQAROR LOKATSIYA HANDLERI (FAQAT OCHIQ FORWARD TEKSHIRILADI) ---
@dp.message(F.location)
async def handle_location(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    user_ids.add(user_id)
    
    # 1. BLOKLANGAN FOYDALANUVCHINI TEKSHIRISH
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return

    # 2. FAQAT OCHIQ FORWARDNI TEKSHIRAMIZ
    if message.forward_origin is not None:
        user_warning = (
            "⚠️ DIQQAT: SOXTA DAVOMATGA URINISH!\n\n"
            "Siz boshqa foydalanuvchidan uzatilgan (forward) lokatsiyani yuborish orqali "
            "yolg'on davomat qilishga urundingiz.\n\n"
            "🚫 Ushbu harakatingiz soxtakorlik sifatida qayd etildi va adminlarga yuborildi!"
        )
        await message.answer(user_warning, parse_mode="Markdown")

        t_name = user_names.get(user_id, message.from_user.full_name)
        t_spec = user_specialty.get(user_id, 'Noma\'lum')
        admin_alert = (
            f"🚨 SOXTA DAVOMATGA URINISH!\n\n"
            f"👤 Xodim: {t_name}\n"
            f"📚 Soha: {t_spec}\n"
            f"🆔 ID: `{user_id}`\n"
            f"📍 Holat: Forward qilingan lokatsiya yubordi.\n"
            f"🕒 Vaqt: {datetime.now(UZB_TZ).strftime('%H:%M:%S')}"
        )
        await bot.send_message(ADMIN_GROUP_ID, admin_alert, parse_mode="Markdown")
        return

    # 3. MASOFANI O'LCHASH
    now_uzb = datetime.now(UZB_TZ)
    today_date = now_uzb.strftime("%Y-%m-%d")
    current_month = now_uzb.strftime("%Y-%m")
    now_time = now_uzb.strftime("%H:%M:%S")

    user_coords = (message.location.latitude, message.location.longitude)
    found_branch = None
    min_distance = float('inf')
    
    for branch in LOCATIONS:
        dist = geodesic((branch["lat"], branch["lon"]), user_coords).meters
        if dist <= ALLOWED_DISTANCE:
            if dist < min_distance:
                min_distance = dist
                found_branch = branch["name"]
    
    # 4. NATIJAGA QARAB JAVOB BERISH
    if found_branch:
        # Avval davomat qilganmi?
        already_attended = any(k[0] == user_id and k[1] == found_branch and k[2] == today_date for k in daily_attendance_log)
        if already_attended:
            await message.answer(get_text(user_id, 'already_attended', branch=found_branch), parse_mode="Markdown")
            return

        # PostgreSQL va RAM ga saqlash
        await db.save_attendance(user_id, found_branch, today_date, now_time)
        daily_attendance_log.add((user_id, found_branch, today_date, now_time))
        
        counter_key = (user_id, found_branch, current_month)
        attendance_counter[counter_key] = attendance_counter.get(counter_key, 0) + 1
        visit_number = attendance_counter[counter_key]
        
        # Adminga hisobot
        full_name = user_names.get(user_id, message.from_user.full_name)
        specialty = user_specialty.get(user_id, '')
        specialty_display = f" [{specialty}]" if specialty else ""
        
        report = (
            f"✅ Yangi Davomat\n\n"
            f"👤 O'qituvchi: {full_name}{specialty_display}\n"
            f"📍 Manzil: {found_branch}\n"
            f"📅 Sana: {today_date}\n"
            f"⏰ Vaqt: {now_time}\n"
            f"📊 Shu oydagi tashrif: {visit_number}-marta\n"
            f"📍 Masofa: {min_distance:.1f} metr"
        )
        await bot.send_message(chat_id=ADMIN_GROUP_ID, text=report, parse_mode="Markdown")

        # Foydalanuvchiga muvaffaqiyat xabari
        success_text = get_text(
            user_id, 
            'attendance_success', 
            branch=found_branch, 
            date=today_date, 
            time=now_time, 
            count=visit_number, 
            distance=min_distance
        )
        
        # Keyboard yangilash (yangi tugmalar bilan)
        new_kb = await main_keyboard(user_id)
        await message.answer(success_text, parse_mode="Markdown", reply_markup=new_kb)

        # --- O'QUVCHILAR DAVOMATI TIZIMI (faqat o'qituvchilar uchun) ---
        if user_specialty.get(user_id) != 'Ofis xodimi':
            today_day = WEEKDAYS_UZ[now_uzb.weekday()]
            now_hour = now_uzb.hour
            now_minute = now_uzb.minute
            now_total = now_hour * 60 + now_minute  # daqiqalarda

            # Bugun bu filialdagi o'qituvchi guruhlarini topamiz
            # Vaqtga mos (±90 daqiqa) guruhlarni ustuvor qilamiz
            today_groups = []
            for g_id, g_data in groups.items():
                if g_data.get('teacher_id') != user_id:
                    continue
                days_data = g_data.get('days', [])
                if today_day not in days_data:
                    continue
                # Guruh vaqtini tekshiramiz
                day_times = g_data.get('day_times', {})
                grp_time_str = day_times.get(today_day, '')
                grp_minutes = None
                if grp_time_str:
                    try:
                        h, m = grp_time_str.strip().split(':')
                        grp_minutes = int(h) * 60 + int(m)
                    except Exception:
                        pass
                # Vaqt farqi: ±90 daqiqa ichida bo'lsa "joriy dars"
                is_current = False
                if grp_minutes is not None:
                    diff = abs(now_total - grp_minutes)
                    is_current = diff <= 90
                today_groups.append((g_id, g_data, grp_time_str, is_current))

            if today_groups:
                # Joriy (vaqtga mos) guruhlar ustuvor
                current_groups = [g for g in today_groups if g[3]]
                show_groups = current_groups if current_groups else today_groups

                if len(show_groups) == 1:
                    g_id, g_data, g_time, _ = show_groups[0]
                    studs = group_students.get(g_id, [])
                    # Miniapp URL — to'g'ridan guruh sahifasiga
                    webapp_url = (
                        f"{BASE_URL}/miniapp"
                        f"?user_id={user_id}&group_id={g_id}&date={today_date}"
                    )
                    grp_name = g_data.get('group_name', 'Guruh')
                    time_txt = f" · ⏰ {g_time}" if g_time else ""
                    stud_txt = f"🧑‍🎓 {len(studs)} ta o'quvchi" if studs else "⚠️ O'quvchi yo'q"
                    mini_kb = InlineKeyboardMarkup(inline_keyboard=[[
                        InlineKeyboardButton(
                            text=f"📋 {grp_name} — davomatni belgilash",
                            web_app=types.WebAppInfo(url=webapp_url)
                        )
                    ]])
                    await message.answer(
                        f"📚 *{grp_name}*{time_txt}\n"
                        f"{stud_txt}\n\n"
                        f"👇 Tugmani bosib o'quvchilar davomatini belgilang:",
                        reply_markup=mini_kb,
                        parse_mode="Markdown"
                    )
                else:
                    # Bir nechta guruh — har biri uchun alohida tugma
                    webapp_url_base = f"{BASE_URL}/miniapp?user_id={user_id}&date={today_date}"
                    buttons = []
                    grp_info_lines = []
                    for g_id, g_data, g_time, is_cur in show_groups:
                        grp_name = g_data.get('group_name', 'Guruh')
                        studs = group_students.get(g_id, [])
                        cur_mark = " ⚡" if is_cur else ""
                        time_txt = f" {g_time}" if g_time else ""
                        buttons.append([InlineKeyboardButton(
                            text=f"📋 {grp_name}{cur_mark}",
                            web_app=types.WebAppInfo(
                                url=f"{webapp_url_base}&group_id={g_id}"
                            )
                        )])
                        grp_info_lines.append(
                            f"  • *{grp_name}*{time_txt} — {len(studs)} o'quvchi{cur_mark}"
                        )
                    multi_kb = InlineKeyboardMarkup(inline_keyboard=buttons)
                    grp_list = "\n".join(grp_info_lines)
                    await message.answer(
                        f"📚 Bugun *{len(show_groups)} ta* guruhingiz:\n{grp_list}\n\n"
                        f"👇 Guruhni tanlang va davomatni belgilang:",
                        reply_markup=multi_kb,
                        parse_mode="Markdown"
                    )
            else:
                # Bugun dars yo'q - faqat teacher app tugmasi
                teacher_url = f"{BASE_URL}/teacher?user_id={user_id}&section=groups"
                no_lesson_kb = InlineKeyboardMarkup(inline_keyboard=[[
                    InlineKeyboardButton(
                        text="📱 HANCOM Teacher",
                        web_app=types.WebAppInfo(url=teacher_url)
                    )
                ]])
                await message.answer(
                    f"ℹ️ Bugun ({today_day}) jadvalda dars yo'q.",
                    reply_markup=no_lesson_kb,
                    parse_mode="Markdown"
                )
    else:
        # FAQAT HUDUDDAN TASHQARIDA BO'LSA
        await message.answer(get_text(user_id, 'not_in_area'), parse_mode="Markdown")

@dp.message(lambda m: m.text in [
    '📊 Mening statistikam',
    '📋 Dars jadvalim (PDF)',
    '👥 Guruhlarim',
    '👤 Mening profilim',
    '🏢 Filiallar',
])
async def open_teacher_miniapp_section(message: types.Message):
    uid = message.from_user.id
    if uid not in user_ids or user_status.get(uid) == 'blocked':
        return
    section_map = {
        '📊 Mening statistikam': 'stats',
        '📋 Dars jadvalim (PDF)': 'schedule',
        '👥 Guruhlarim': 'groups',
        '👤 Mening profilim': 'profile',
        '🏢 Filiallar': 'branches',
    }
    section = section_map.get(message.text, 'stats')
    section_names = {
        'stats': '📊 Statistika', 'schedule': '📅 Dars jadvali',
        'groups': '📚 Guruhlarim', 'profile': '👤 Profil', 'branches': '🏢 Filiallar'
    }
    webapp_url = f"{BASE_URL}/teacher?user_id={uid}&section={section}"
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text=f"{section_names.get(section,'📱 Ochish')} →",
            web_app=types.WebAppInfo(url=webapp_url)
        )
    ]])
    await message.answer(f"{section_names.get(section,'📱')} bo'limini ochish uchun tugmani bosing:", reply_markup=kb)

@dp.message(F.text == "🧑‍🎓 O'quvchilar davomati")
async def open_student_att_miniapp(message: types.Message):
    uid = message.from_user.id
    if uid not in user_ids or user_status.get(uid) == 'blocked':
        return
    if user_specialty.get(uid) == 'Ofis xodimi':
        await message.answer("Bu funksiya sizda mavjud emas.")
        return
    webapp_url = f"{BASE_URL}/miniapp?user_id={uid}"
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="🧑‍🎓 Davomatni belgilash",
            web_app=types.WebAppInfo(url=webapp_url)
        )
    ]])
    await message.answer(
        "📋 *O'quvchilar davomati*\n\nQuyidagi tugmani bosib Mini App ni oching:",
        reply_markup=kb, parse_mode="Markdown"
    )

# --- O'QUVCHILAR DAVOMATI HANDLERLARI ---
@dp.callback_query(StudentAttendance.selecting_students, F.data.startswith("std_check_"))
async def std_check_callback(callback: types.CallbackQuery, state: FSMContext):
    try:

        idx = int(callback.data.replace("std_check_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    data = await state.get_data()
    selected = data.get('selected_stds', [])
    
    if idx in selected:
        selected.remove(idx)
    else:
        selected.append(idx)
    
    await state.update_data(selected_stds=selected)
    await callback.message.edit_reply_markup(
        reply_markup=await get_student_attendance_kb(data['current_group_id'], selected)
    )
    await callback.answer()

@dp.callback_query(StudentAttendance.selecting_students, F.data == "std_submit")
async def std_submit_callback(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer("⏳ Davomat saqlanmoqda...")
    data = await state.get_data()
    g_id = data['current_group_id']
    selected = data.get('selected_stds', [])
    students = group_students.get(g_id, [])
    user_id = callback.from_user.id

    now_uzb = datetime.now(UZB_TZ)
    group_name = groups[g_id]['group_name']
    teacher_name = user_names.get(user_id, "Noma'lum")
    current_date = now_uzb.strftime('%d.%m.%Y')
    current_day = WEEKDAYS_UZ[now_uzb.weekday()]
    col_header = f"{current_date}\n{current_day}"

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Excel yaratish
    if g_id in group_attendance_files:
        buf_in = io.BytesIO(group_attendance_files[g_id])
        wb = load_workbook(buf_in)
        ws = wb.active
        new_col = ws.max_column + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Guruh Davomati"
        headers = ['№', "O'quvchi", 'Telefon']
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for i, s in enumerate(students):
            ws.cell(row=i+2, column=1, value=i+1).border = border
            ws.cell(row=i+2, column=2, value=s['name']).border = border
            ws.cell(row=i+2, column=3, value=s['phone']).border = border
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 16
        new_col = 4

    col_letter = get_column_letter(new_col)
    hcell = ws.cell(row=1, column=new_col, value=col_header)
    hcell.font = Font(bold=True, color="FFFFFF")
    hcell.fill = header_fill
    hcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hcell.border = border
    ws.row_dimensions[1].height = 30
    ws.column_dimensions[col_letter].width = 14

    for i, s in enumerate(students):
        status = "Kelgan" if i in selected else "Kelmagan"
        dcell = ws.cell(row=i+2, column=new_col, value=status)
        dcell.alignment = Alignment(horizontal="center")
        dcell.border = border
        dcell.fill = green_fill if i in selected else red_fill

    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    file_bytes = buf_out.read()
    group_attendance_files[g_id] = file_bytes

    # Excel ni DB ga ham saqlash (restart uchun)
    try:
        async with db.pool.acquire() as conn:
            await conn.execute(
                """INSERT INTO group_excel_files (group_id, file_data, updated_at)
                   VALUES ($1, $2, NOW())
                   ON CONFLICT (group_id) DO UPDATE
                   SET file_data=EXCLUDED.file_data, updated_at=NOW()""",
                g_id, file_bytes
            )
    except Exception as e:
        logging.error(f"Excel DB saqlash xato: {e}")

    # student_attendance DB ga saqlash
    lesson_date = now_uzb.date()
    try:
        async with db.pool.acquire() as conn:
            await conn.execute(
                "DELETE FROM student_attendance WHERE group_id=$1 AND lesson_date=$2",
                g_id, lesson_date
            )
            for i, s in enumerate(students):
                status = "Kelgan" if i in selected else "Kelmagan"
                await conn.execute(
                    "INSERT INTO student_attendance (group_id, student_name, student_phone, lesson_date, status) VALUES ($1,$2,$3,$4,$5)",
                    g_id, s['name'], s['phone'], lesson_date, status
                )
    except Exception as e:
        logging.error(f"student_attendance DB saqlash xatosi: {e}")

    # Adminga Excel yuborish
    filename = f"Davomat_{group_name}.xlsx"
    ginfo = groups.get(g_id, {})
    branch = ginfo.get('branch', '—')
    lesson_type = ginfo.get('lesson_type', '—')
    day_times_g = ginfo.get('day_times', {})
    lesson_time = day_times_g.get(current_day) or ginfo.get('time', '—')
    caption = (
        f"📋 *Guruh Davomati*\n\n"
        f"👥 Guruh: {group_name}\n"
        f"👤 O'qituvchi: {teacher_name}\n"
        f"🏢 Filial: {branch}\n"
        f"📚 Fan: {lesson_type}\n"
        f"⏰ Vaqt: {lesson_time}\n"
        f"📅 {current_date} — {current_day}\n\n"
        f"✅ Kelganlar: {len(selected)}/{len(students)}\n"
        f"❌ Kelmagan: {len(students) - len(selected)}/{len(students)}"
    )
    try:
        await bot.send_document(
            ADMIN_GROUP_ID,
            types.BufferedInputFile(file_bytes, filename=filename),
            caption=caption,
            parse_mode="Markdown"
        )
    except Exception as e:
        logging.error(f"Admin ga davomat yuborishda xato: {e}")

    # State tozalash
    await state.clear()

    # O'qituvchiga — ro'yxatni edit qilib natija ko'rsatamiz
    kelganlar = [students[i]['name'] for i in selected]
    kelmagan = [students[i]['name'] for i in range(len(students)) if i not in selected]

    result_text = (
        f"✅ *Davomat yuborildi!*\n\n"
        f"👥 {group_name}\n"
        f"📅 {current_date} — {current_day}\n\n"
        f"✅ Kelganlar ({len(selected)}):\n"
    )
    for name in kelganlar:
        result_text += f"  • {name}\n"
    if kelmagan:
        result_text += f"\n❌ Kelmagan ({len(kelmagan)}):\n"
        for name in kelmagan:
            result_text += f"  • {name}\n"

    late_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="🕐 Kech qolgan o'quvchini belgilash",
            callback_data=f"std_late_{g_id}"
        )
    ]])

    try:
        await callback.message.edit_text(
            result_text,
            reply_markup=late_kb,
            parse_mode="Markdown"
        )
    except Exception:
        # edit ishlamasa yangi xabar
        try:
            pass  # message.delete() webhook da ishlamas
        except:
            pass
        await bot.send_message(
            callback.from_user.id,
            result_text,
            reply_markup=late_kb,
            parse_mode="Markdown"
        )


@dp.callback_query(F.data.startswith("std_late_"))
async def std_late_start(callback: types.CallbackQuery, state: FSMContext):
    try:

        g_id = int(callback.data.replace("std_late_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    students = group_students.get(g_id, [])
    if not students:
        await callback.answer("❌ O'quvchilar topilmadi!", show_alert=True)
        return

    await callback.answer()
    # Bugungi davomatni DB dan o'qiymiz
    lesson_date = datetime.now(UZB_TZ).date()
    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch(
                "SELECT student_name, status FROM student_attendance WHERE group_id=$1 AND lesson_date=$2",
                g_id, lesson_date
            )
        status_map = {r['student_name']: r['status'] for r in rows}
    except Exception as e:
        logging.error(f"std_late DB read error: {e}")
        status_map = {}

    # Allaqachon "Kelgan" bo'lganlarni selected qilamiz
    pre_selected = [
        i for i, s in enumerate(students)
        if status_map.get(s['name'], 'Kelmagan') == 'Kelgan'
    ]

    await state.update_data(late_group_id=g_id, late_selected=pre_selected)
    await state.set_state(StudentAttendance.late_students)

    kb = await get_student_attendance_kb(g_id, pre_selected)
    await callback.message.edit_text(
        f"🕐 Kech qolgan o'quvchilarni ham belgilang\n\n"
        f"Allaqachon belgilangan o'quvchilar — avval kelganlar.\n"
        f"Kech qolganlarni ham belgilab, *Davomatni yakunlash* tugmasini bosing:",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(StudentAttendance.late_students, F.data.startswith("std_check_"))
async def std_late_check(callback: types.CallbackQuery, state: FSMContext):
    try:

        idx = int(callback.data.replace("std_check_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    data = await state.get_data()
    g_id = data['late_group_id']
    selected = data.get('late_selected', [])
    if idx in selected:
        selected.remove(idx)
    else:
        selected.append(idx)
    await state.update_data(late_selected=selected)
    await callback.message.edit_reply_markup(
        reply_markup=await get_student_attendance_kb(g_id, selected)
    )
    await callback.answer()

@dp.callback_query(StudentAttendance.late_students, F.data == "std_submit")
async def std_late_submit(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer("⏳ Saqlanmoqda...")
    data = await state.get_data()
    g_id = data['late_group_id']
    selected = data.get('late_selected', [])
    students = group_students.get(g_id, [])
    user_id = callback.from_user.id

    now_uzb = datetime.now(UZB_TZ)
    group_name = groups[g_id]['group_name']
    teacher_name = user_names.get(user_id, "Noma'lum")
    current_date = now_uzb.strftime('%d.%m.%Y')
    current_day = WEEKDAYS_UZ[now_uzb.weekday()]
    lesson_date = now_uzb.date()

    # DB yangilash
    try:
        async with db.pool.acquire() as conn:
            await conn.execute(
                "DELETE FROM student_attendance WHERE group_id=$1 AND lesson_date=$2",
                g_id, lesson_date
            )
            for i, s in enumerate(students):
                status = "Kelgan" if i in selected else "Kelmagan"
                await conn.execute(
                    "INSERT INTO student_attendance (group_id, student_name, student_phone, lesson_date, status) VALUES ($1,$2,$3,$4,$5)",
                    g_id, s['name'], s['phone'], lesson_date, status
                )
    except Exception as e:
        logging.error(f"std_late DB update error: {e}")

    # Excel yangilash
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if g_id in group_attendance_files:
        buf_in = io.BytesIO(group_attendance_files[g_id])
        wb = load_workbook(buf_in)
        ws = wb.active
        last_col = ws.max_column
        # O'sha kunda yaratilgan oxirgi ustunni yangilaymiz
        for i, s in enumerate(students):
            status = "Kelgan" if i in selected else "Kelmagan"
            dcell = ws.cell(row=i+2, column=last_col, value=status)
            dcell.alignment = Alignment(horizontal="center")
            dcell.border = border
            dcell.fill = green_fill if i in selected else red_fill
        buf_out = io.BytesIO()
        wb.save(buf_out)
        buf_out.seek(0)
        file_bytes = buf_out.read()
        group_attendance_files[g_id] = file_bytes
        # Excel ni DB ga saqlash
        try:
            async with db.pool.acquire() as conn:
                await conn.execute(
                    """INSERT INTO group_excel_files (group_id, file_data, updated_at)
                       VALUES ($1, $2, NOW())
                       ON CONFLICT (group_id) DO UPDATE
                       SET file_data=EXCLUDED.file_data, updated_at=NOW()""",
                    g_id, file_bytes
                )
        except Exception as e:
            logging.error(f"Excel DB saqlash xato (late): {e}")
    else:
        file_bytes = None

    # Adminga yangilangan fayl yuborish
    ginfo2 = groups.get(g_id, {})
    branch2 = ginfo2.get('branch', '—')
    lesson_type2 = ginfo2.get('lesson_type', '—')
    day_times_g2 = ginfo2.get('day_times', {})
    lesson_time2 = day_times_g2.get(current_day) or ginfo2.get('time', '—')
    kech_qolgan = len(selected)
    caption = (
        f"🕐 *Davomat yangilandi (kech qolganlar)*\n\n"
        f"👥 Guruh: {group_name}\n"
        f"👤 O'qituvchi: {teacher_name}\n"
        f"🏢 Filial: {branch2}\n"
        f"📚 Fan: {lesson_type2}\n"
        f"⏰ Vaqt: {lesson_time2}\n"
        f"📅 {current_date} — {current_day}\n\n"
        f"🕐 Kech qolganlar bilan kelganlar: {kech_qolgan}/{len(students)}"
    )
    if file_bytes:
        await bot.send_document(
            ADMIN_GROUP_ID,
            types.BufferedInputFile(file_bytes, filename=f"Davomat_{group_name}.xlsx"),
            caption=caption,
            parse_mode="Markdown"
        )

    await state.clear()

    # O'qituvchiga natija ko'rsatish
    kelganlar = [students[i]['name'] for i in selected]
    kelmagan = [students[i]['name'] for i in range(len(students)) if i not in selected]

    result_text = (
        f"✅ *Davomat yangilandi!*\n\n"
        f"👥 {group_name}\n"
        f"📅 {current_date} — {current_day}\n\n"
        f"✅ Kelganlar ({len(selected)}):\n"
    )
    for name in kelganlar:
        result_text += f"  • {name}\n"
    if kelmagan:
        result_text += f"\n❌ Kelmagan ({len(kelmagan)}):\n"
        for name in kelmagan:
            result_text += f"  • {name}\n"

    # "Kech qolgan" tugmasini yana ko'rsatamiz — o'qituvchi qayta yangilay olsin
    late_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="🕐 Kech qolganlarni qayta belgilash",
            callback_data=f"std_late_{g_id}"
        )
    ]])

    try:
        await callback.message.edit_text(result_text, reply_markup=late_kb, parse_mode="Markdown")
    except Exception:
        try:
            pass  # message.delete() webhook da ishlamas
        except:
            pass
        await bot.send_message(callback.from_user.id, result_text,
                               reply_markup=late_kb, parse_mode="Markdown")


async def get_calendar_keyboard(year: int, month: int, lang: str):
    builder = InlineKeyboardBuilder()
    
    month_names = ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun", "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]
    wd_names = ["Du", "Se", "Ch", "Pa", "Ju", "Sha", "Ya"]

    m_name = month_names[month-1]
    header_text = f"{m_name} {year}"
    
    builder.row(
        InlineKeyboardButton(text="⬅️", callback_data=f"cal_nav_prev_{year}_{month}"),
        InlineKeyboardButton(text=header_text, callback_data="ignore"),
        InlineKeyboardButton(text="➡️", callback_data=f"cal_nav_next_{year}_{month}")
    )

    header_days = [InlineKeyboardButton(text=day, callback_data="ignore") for day in wd_names]
    builder.row(*header_days)

    month_calendar = calendar.monthcalendar(year, month)
    for week in month_calendar:
        row_btns = []
        for day in week:
            if day == 0:
                row_btns.append(InlineKeyboardButton(text=" ", callback_data="ignore"))
            else:
                row_btns.append(InlineKeyboardButton(
                    text=str(day), 
                    callback_data=f"cal_set_{year}-{month:02d}-{day:02d}")
                )
        builder.row(*row_btns)

    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    return builder.as_markup()

@dp.callback_query(F.data == "admin_pdf_report")
async def admin_pdf_report_start(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    now = datetime.now(UZB_TZ)
    lang = user_languages.get(callback.from_user.id, 'uz')
    
    keyboard = await get_calendar_keyboard(now.year, now.month, lang)
    await callback.message.edit_text(
        "📅 Hisobot sanasini kalendardan tanlang:",
        reply_markup=keyboard
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("cal_nav_"))
async def process_calendar_navigation(callback: types.CallbackQuery):
    parts = callback.data.split("_")
    action = parts[2]
    year = int(parts[3])
    month = int(parts[4])
    
    if action == "prev":
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    else:
        month += 1
        if month == 13:
            month = 1
            year += 1
            
    lang = user_languages.get(callback.from_user.id, 'uz')
    keyboard = await get_calendar_keyboard(year, month, lang)
    
    await callback.message.edit_reply_markup(reply_markup=keyboard)
    await callback.answer()

@dp.callback_query(F.data.startswith("cal_set_"))
async def process_calendar_selection(callback: types.CallbackQuery):
    date_str = callback.data.replace("cal_set_", "")
    report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    
    await callback.message.answer(f"⏳ {date_str} davomat hisoboti tayyorlanmoqda...")
    
    try:
        buf = await get_combined_report_pdf(report_date)
        await callback.message.answer_document(
            types.BufferedInputFile(buf.read(), filename=f"Davomat_{date_str}.xlsx"),
            caption=f"📊 {date_str} hisoboti (Excel formatda)."
        )
    except Exception as e:
        logging.error(f"Calendar Excel error: {e}")
        await callback.message.answer("❌ Fayl yaratishda xatolik yuz berdi.")
        
    await callback.answer()

@dp.callback_query(F.data == "ignore")
async def process_ignore_callback(callback: types.CallbackQuery):
    await callback.answer()

@dp.message(Command("admin"))
async def admin_panel(message: types.Message):
    if not check_admin(message.chat.id):
        return
    
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="💰 Oylik hisoblash", callback_data="admin_salary_calc")
        )
        builder.row(
            InlineKeyboardButton(text="📚 Guruhlar", callback_data="admin_groups_menu")
        )
        builder.row(
            InlineKeyboardButton(text="👥 Foydalanuvchilar", callback_data="admin_users_main"),
            InlineKeyboardButton(text="📢 Xabar yuborish", callback_data="admin_broadcast")
        )
        builder.row(
            InlineKeyboardButton(text="🏢 Filiallar", callback_data="admin_locations_main"),
            InlineKeyboardButton(text="📅 Dars jadvallari", callback_data="admin_schedules_main")
        )
        builder.row(
            InlineKeyboardButton(text="📊 Oylik hisobot (Excel)", callback_data="admin_excel_menu"),
            InlineKeyboardButton(text="📊 Kunlik PDF", callback_data="admin_pdf_report")
        )
        builder.row(
            InlineKeyboardButton(text="🧑‍🎓 O'quvchilar davomati", callback_data="admin_student_att_branches")
        )
        
        await message.answer(
            "👨‍💼 Admin Panel\n\nKerakli bo'limni tanlang:",
            reply_markup=builder.as_markup()
        )
    except Exception as e:
        logging.error(f"admin_panel error: {e}")
        await message.answer("❌ Admin panelni ochishda xatolik yuz berdi")

# --- GURUHLAR MENYU ---
@dp.callback_query(F.data == "admin_groups_menu")
async def admin_groups_menu(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="➕ Yangi Guruh", callback_data="admin_create_group"))
    builder.row(InlineKeyboardButton(text="👥 Faol Guruhlar", callback_data="admin_active_groups"))
    builder.row(InlineKeyboardButton(text="📋 PDF Dars Jadval", callback_data="admin_schedules_pdf"))
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    await callback.message.edit_text(
        "📚 Guruhlar\n\nKerakli amalni tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()

# --- O'QUVCHILAR DAVOMATI (ADMIN) ---
@dp.callback_query(F.data == "admin_student_att_branches")
async def admin_student_att_branches(callback: types.CallbackQuery):
    await callback.answer()
    builder = InlineKeyboardBuilder()
    for i, loc in enumerate(LOCATIONS):
        builder.row(InlineKeyboardButton(
            text=f"🏢 {loc['name']}",
            callback_data=f"stdatt_br_{i}"
        ))
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    await callback.message.edit_text(
        "🧑‍🎓 O'quvchilar davomati\n\nFilial tanlang:",
        reply_markup=builder.as_markup()
    )

@dp.callback_query(F.data.startswith("stdatt_br_"))
async def admin_student_att_months(callback: types.CallbackQuery):
    await callback.answer()
    try:

        idx = int(callback.data.replace("stdatt_br_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    if idx < 0 or idx >= len(LOCATIONS):
        await callback.message.edit_text("❌ Filial topilmadi!")
        return
    branch = LOCATIONS[idx]['name']

    await callback.message.edit_text(f"🏢 {branch}\n\n⏳ Ma'lumotlar yuklanmoqda...")

    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT DISTINCT TO_CHAR(sa.lesson_date, 'YYYY-MM') as ym,
                       TO_CHAR(sa.lesson_date, 'MM') as month_num,
                       TO_CHAR(sa.lesson_date, 'YYYY') as year
                FROM student_attendance sa
                JOIN groups g ON sa.group_id = g.id
                WHERE g.branch = $1
                ORDER BY ym DESC
                LIMIT 12
            """, branch)
    except Exception as e:
        logging.error(f"stdatt months error: {e}")
        await callback.message.edit_text(
            f"❌ Xatolik yuz berdi: {e}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_student_att_branches")
            ]])
        )
        return

    MONTHS_UZ = {
        '01': 'Yanvar', '02': 'Fevral', '03': 'Mart', '04': 'Aprel',
        '05': 'May', '06': 'Iyun', '07': 'Iyul', '08': 'Avgust',
        '09': 'Sentabr', '10': 'Oktabr', '11': 'Noyabr', '12': 'Dekabr'
    }

    if not rows:
        await callback.message.edit_text(
            f"🏢 {branch}\n\n📭 Hozircha o'quvchilar davomati yo'q.\n\nBu filialda davomat qilingan guruh topilmadi.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_student_att_branches")
            ]])
        )
        return

    branch_idx = idx
    builder = InlineKeyboardBuilder()
    for row in rows:
        label = f"{MONTHS_UZ.get(row['month_num'], row['month_num'])} {row['year']}"
        builder.row(InlineKeyboardButton(
            text=f"📅 {label}",
            callback_data=f"stdatt_month_{row['ym']}_{branch_idx}"
        ))
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_student_att_branches"))
    await callback.message.edit_text(
        f"🏢 {branch}\n\nOy tanlang:",
        reply_markup=builder.as_markup()
    )

@dp.callback_query(F.data.startswith("stdatt_month_"))
async def admin_student_att_excel(callback: types.CallbackQuery):
    parts = callback.data.replace("stdatt_month_", "").split("_", 1)
    ym = parts[0]   # "2025-03"
    branch_idx = int(parts[1]) if len(parts) > 1 else 0
    branch = LOCATIONS[branch_idx]['name'] if branch_idx < len(LOCATIONS) else ""
    year, month = ym.split("-")

    MONTHS_UZ = {
        '01': 'Yanvar', '02': 'Fevral', '03': 'Mart', '04': 'Aprel',
        '05': 'May', '06': 'Iyun', '07': 'Iyul', '08': 'Avgust',
        '09': 'Sentabr', '10': 'Oktabr', '11': 'Noyabr', '12': 'Dekabr'
    }
    month_name = MONTHS_UZ.get(month, month)

    try:
        async with db.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT sa.student_name, sa.student_phone,
                       sa.lesson_date, sa.status,
                       g.group_name
                FROM student_attendance sa
                JOIN groups g ON sa.group_id = g.id
                WHERE g.branch = $1
                  AND TO_CHAR(sa.lesson_date, 'YYYY-MM') = $2
                ORDER BY g.group_name, sa.student_name, sa.lesson_date
            """, branch, ym)
    except Exception as e:
        logging.error(f"stdatt excel error: {e}")
        await callback.answer("❌ Xatolik yuz berdi", show_alert=True)
        return

    if not rows:
        await callback.answer("📭 Bu oy uchun davomat yo'q", show_alert=True)
        return

    await callback.answer("⏳ Excel tayyorlanmoqda...")

    # Guruh bo'yicha ma'lumotlarni tartiblaymiz
    from collections import defaultdict as _dd
    # {group_name: {student: {date: status}}}
    groups_data = _dd(lambda: _dd(lambda: _dd(str)))
    groups_students = _dd(set)  # guruh -> (name, phone) set
    all_dates = set()

    for r in rows:
        gn = r['group_name']
        key = (r['student_name'], r['student_phone'])
        groups_students[gn].add(key)
        groups_data[gn][key][r['lesson_date']] = r['status']
        all_dates.add(r['lesson_date'])

    all_dates = sorted(all_dates)
    WEEKDAYS_SHORT = ['Du', 'Se', 'Ch', 'Pa', 'Ju', 'Sh', 'Ya']

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    group_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for gn, students_dict in groups_data.items():
        ws = wb.create_sheet(title=gn[:30])

        # 1-qator: Guruh nomi va oy
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(all_dates))
        title_cell = ws.cell(row=1, column=1, value=f"{gn} — {month_name} {year}")
        title_cell.font = Font(bold=True, size=12)
        title_cell.fill = group_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # 2-qator: sarlavhalar
        headers = ['№', "O'quvchi", 'Telefon']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

        for di, d in enumerate(all_dates):
            day_label = f"{d.strftime('%d.%m')}\n{WEEKDAYS_SHORT[d.weekday()]}"
            c = ws.cell(row=2, column=4 + di, value=day_label)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
            ws.column_dimensions[get_column_letter(4 + di)].width = 8

        ws.row_dimensions[2].height = 28
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 16

        # O'quvchilar
        for ri, (key, dates_dict) in enumerate(sorted(students_dict.items()), 1):
            name, phone = key
            ws.cell(row=ri+2, column=1, value=ri).border = border
            nc = ws.cell(row=ri+2, column=2, value=name)
            nc.border = border
            pc = ws.cell(row=ri+2, column=3, value=phone)
            pc.border = border

            kelgan = 0
            for di, d in enumerate(all_dates):
                status = dates_dict.get(d, "—")
                sc = ws.cell(row=ri+2, column=4+di, value=status)
                sc.alignment = Alignment(horizontal="center")
                sc.border = border
                if status == "Kelgan":
                    sc.fill = green_fill
                    kelgan += 1
                elif status == "Kelmagan":
                    sc.fill = red_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Oquvchilar_Davomat_{branch}_{month_name}_{year}.xlsx"
    caption = (
        f"🧑‍🎓 O'quvchilar davomati\n"
        f"🏢 Filial: {branch}\n"
        f"📅 {month_name} {year}\n"
        f"📊 Jami darslar: {len(all_dates)} ta"
    )
    await bot.send_document(
        ADMIN_GROUP_ID,
        types.BufferedInputFile(buf.read(), filename=filename),
        caption=caption
    )

# --- GURUH YARATISH HANDLERLARI (YANGI) - TUZATILGAN VERSIYA---
@dp.callback_query(F.data == "admin_create_group")
async def start_group_creation(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    builder = InlineKeyboardBuilder()
    for loc in LOCATIONS:
        builder.row(InlineKeyboardButton(text=loc['name'], callback_data=f"grp_br_{loc['name']}"))
    builder.row(InlineKeyboardButton(text="🔙 Bekor qilish", callback_data="admin_back"))
    await callback.message.edit_text("Qaysi filialga guruh qo'shmoqchisiz?", reply_markup=builder.as_markup())
    await state.set_state(CreateGroup.selecting_branch)

@dp.callback_query(CreateGroup.selecting_branch, F.data.startswith("grp_br_"))
async def grp_branch_selected(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    await state.update_data(branch=callback.data.replace("grp_br_", ""))
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="💻 IT", callback_data="grp_type_IT"),
                InlineKeyboardButton(text="🇰🇷 Koreys tili", callback_data="grp_type_Koreys tili"))
    await callback.message.edit_text("Dars turini tanlang:", reply_markup=builder.as_markup())
    await state.set_state(CreateGroup.selecting_type)

@dp.callback_query(CreateGroup.selecting_type, F.data.startswith("grp_type_"))
async def grp_type_selected(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    await state.update_data(type=callback.data.replace("grp_type_", ""))
    builder = InlineKeyboardBuilder()
    for uid, name in user_names.items():
        if user_specialty.get(uid) != 'Ofis xodimi' and user_specialty.get(uid) is not None:
            builder.row(InlineKeyboardButton(text=name, callback_data=f"grp_tchr_{uid}"))
    await callback.message.edit_text("O'qituvchini tanlang:", reply_markup=builder.as_markup())
    await state.set_state(CreateGroup.selecting_teacher)

@dp.callback_query(CreateGroup.selecting_teacher, F.data.startswith("grp_tchr_"))
async def grp_teacher_selected(callback: types.CallbackQuery, state: FSMContext):
    try:

        teacher_id = int(callback.data.replace("grp_tchr_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    await state.update_data(teacher_id=teacher_id, selected_days=[])
    await state.set_state(CreateGroup.selecting_days)
    await grp_show_days(callback.message, [])
    await callback.answer()

async def grp_show_days(message: types.Message, selected):
    builder = InlineKeyboardBuilder()
    for day in WEEKDAYS_UZ:
        text = f"✅ {day}" if day in selected else f"⬜ {day}"
        builder.row(InlineKeyboardButton(text=text, callback_data=f"grp_day_{day}"))
    builder.row(InlineKeyboardButton(text="➡️ Keyingisi", callback_data="grp_days_next"))
    await message.edit_text("Dars kunlarini tanlang (bir nechta tanlashingiz mumkin):", reply_markup=builder.as_markup())

@dp.callback_query(CreateGroup.selecting_days, F.data.startswith("grp_day_"))
async def grp_toggle_day(callback: types.CallbackQuery, state: FSMContext):
    day = callback.data.replace("grp_day_", "")
    data = await state.get_data()
    selected = data.get('selected_days', [])
    if day in selected:
        selected.remove(day)
    else:
        selected.append(day)
    await state.update_data(selected_days=selected)
    await grp_show_days(callback.message, selected)
    await callback.answer()

@dp.callback_query(CreateGroup.selecting_days, F.data == "grp_days_next")
async def grp_days_next(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected_days = data.get('selected_days', [])
    if not selected_days:
        await callback.answer("Hech bo'lmaganda 1 kun tanlang!", show_alert=True)
        return
    # Har kun uchun alohida vaqt so'rash — birinchi kundan boshlaymiz
    await state.update_data(day_times={}, current_day_idx=0)
    await state.set_state(CreateGroup.entering_day_times)
    first_day = selected_days[0]
    await callback.message.edit_text(
        f"⏰ *{first_day}* kuni dars vaqtini kiriting\n"
        f"_(masalan: 14:00)_\n\n"
        f"📅 Kunlar: {len(selected_days)} ta | 1/{len(selected_days)}",
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.message(CreateGroup.entering_day_times)
async def grp_day_time_entered(message: types.Message, state: FSMContext):
    time_pattern = re.compile(r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$')
    if not time_pattern.match(message.text.strip()):
        await message.answer("❌ Noto'g'ri format! HH:MM ko'rinishida kiriting (masalan: 14:00):")
        return

    data = await state.get_data()
    selected_days = data.get('selected_days', [])
    day_times = data.get('day_times', {})
    idx = data.get('current_day_idx', 0)

    # Joriy kun vaqtini saqlaymiz
    current_day = selected_days[idx]
    day_times[current_day] = message.text.strip()
    idx += 1

    await state.update_data(day_times=day_times, current_day_idx=idx)

    if idx < len(selected_days):
        # Keyingi kun vaqtini so'raymiz
        next_day = selected_days[idx]
        filled = "\n".join([f"  ✅ {d}: {t}" for d, t in day_times.items()])
        await message.answer(
            f"{filled}\n\n"
            f"⏰ *{next_day}* kuni dars vaqtini kiriting\n"
            f"_(masalan: 16:00)_\n\n"
            f"📅 {idx + 1}/{len(selected_days)}",
            parse_mode="Markdown"
        )
    else:
        # Barcha kunlar vaqti kiritildi — guruh nomini so'raymiz
        filled = "\n".join([f"  ✅ {d}: {t}" for d, t in day_times.items()])
        await message.answer(
            f"✅ Barcha kunlar vaqti kiritildi:\n{filled}\n\n"
            f"Guruh uchun nom kiriting (masalan: Koreys-1 yoki IT-A):",
            parse_mode="Markdown"
        )
        await state.set_state(CreateGroup.entering_group_name)

@dp.message(CreateGroup.entering_group_name)
async def grp_name_entered(message: types.Message, state: FSMContext):
    await state.update_data(group_name=message.text.strip())
    # Shablon Excel yuboramiz
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"
    ws["A1"] = "O'QUVCHILAR RO'YXATI"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="4472C4")
    ws.merge_cells("A1:B1")
    for col, h in enumerate(["Ism Familiya", "Telefon raqami"], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E86AB")
        cell.alignment = Alignment(horizontal="center")
    # Namuna qatorlar
    for i, (name, phone) in enumerate([("Ali Karimov", "+998901234567"), ("Barno Qosimova", "+998911234567")], 1):
        ws.cell(row=2+i, column=1, value=name)
        ws.cell(row=2+i, column=2, value=phone)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    note_row = 5
    ws.cell(row=note_row, column=1, value="ESLATMA: 3-qatordan boshlab o'quvchilarni kiriting. Namuna qatorlarni o'chiring.")
    ws.merge_cells(f"A{note_row}:B{note_row}")
    ws.cell(row=note_row, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await message.answer_document(
        types.BufferedInputFile(buf.read(), filename="oquvchilar_shablon.xlsx"),
        caption=(
            "📋 O'quvchilar ro'yxati uchun shablon\n\n"
            "1. Faylni yuklab oling\n"
            "2. 3-qatordan boshlab o'quvchi ismi va telefon raqamini kiriting\n"
            "3. To'ldirilgan faylni yuboring — guruh yaratiladi"
        )
    )
    await state.set_state(CreateGroup.waiting_excel)

@dp.message(CreateGroup.waiting_excel, F.document)
async def grp_excel_students_received(message: types.Message, state: FSMContext):
    doc = message.document
    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await message.answer("Faqat .xlsx yoki .xls fayl yuboring!")
        return
    try:
        file = await message.bot.get_file(doc.file_id)
        file_buf = io.BytesIO()
        await message.bot.download_file(file.file_path, file_buf)
        file_buf.seek(0)
        wb = openpyxl.load_workbook(file_buf)
        ws = wb.active
        students = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            vals = (list(row) + [None, None])[:2]
            name, phone = vals
            if name and str(name).strip():
                students.append({
                    "name": str(name).strip(),
                    "phone": str(phone).strip() if phone else "-"
                })
        if not students:
            await message.answer("❌ O'quvchilar topilmadi! 3-qatordan boshlab kiriting.")
            return
        data = await state.get_data()
        # day_times: {kun: vaqt} — masalan {"Dushanba": "14:00", "Chorshanba": "16:00"}
        day_times = data.get('day_times', {})
        # Bitta umumiy vaqt olish (display uchun)
        times_display = "\n".join([f"  📅 {d}: {t}" for d, t in day_times.items()])
        # days list
        days_list = list(day_times.keys())

        async with db.pool.acquire() as conn:
            group_id = await conn.fetchval("""
                INSERT INTO groups (group_name, branch, lesson_type, teacher_id, days_data, time_text)
                VALUES ($1, $2, $3, $4, $5::jsonb, $6) RETURNING id
            """, data['group_name'], data['branch'], data['type'], data['teacher_id'],
               json.dumps(day_times),
               ", ".join([f"{d} {t}" for d, t in day_times.items()]))
            for std in students:
                await conn.execute(
                    "INSERT INTO group_students (group_id, student_name, student_phone) VALUES ($1,$2,$3)",
                    group_id, std["name"], std["phone"]
                )
        groups[group_id] = {
            'group_name': data['group_name'],
            'branch': data['branch'],
            'lesson_type': data['type'],
            'teacher_id': data['teacher_id'],
            'days': days_list,
            'day_times': day_times,
            'time': list(day_times.values())[0] if day_times else '',
            'time_text': list(day_times.values())[0] if day_times else '',
            'created_at': datetime.now(UZB_TZ)
        }
        group_students[group_id] = students
        teacher_msg = (
            f"🆕 *Yangi guruh biriktirildi!*\n\n"
            f"👥 Guruh: {data['group_name']}\n"
            f"🏢 Filial: {data['branch']}\n"
            f"📚 Fan: {data['type']}\n"
            f"{times_display}\n"
            f"🧑‍🎓 O'quvchilar soni: {len(students)} ta\n\n"
            f"📍 Botda davomat qilganingizda ushbu o'quvchilar ro'yxati chiqadi."
        )
        try:
            await bot.send_message(data['teacher_id'], teacher_msg, parse_mode="Markdown")
        except Exception as e:
            logging.error(f"Failed to notify teacher: {e}")
        await state.clear()
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="Guruhni ko'rish", callback_data=f"grp_view_{group_id}"))
        builder.row(InlineKeyboardButton(text="Admin Panel", callback_data="admin_back"))
        await message.answer(
            f"✅ Guruh *{data['group_name']}* yaratildi!\n\n"
            f"🏢 {data['branch']}\n"
            f"📚 {data['type']}\n"
            f"👤 O'qituvchi: {user_names.get(data['teacher_id'], '?')}\n"
            f"{times_display}\n"
            f"👥 O'quvchilar: {len(students)} ta",
            reply_markup=builder.as_markup(),
            parse_mode="Markdown"
        )
    except Exception as e:
        logging.error(f"grp_excel_students_received error: {e}")
        await message.answer(f"❌ Faylni o'qishda xatolik: {e}")
        await state.clear()

@dp.message(CreateGroup.waiting_excel)
async def grp_excel_wrong(message: types.Message, state: FSMContext):
    await message.answer("Iltimos, .xlsx formatidagi Excel fayl yuboring!")

async def create_visual_timetable_img(branch_name: str):
    days = ['Dushanba', 'Seshanba', 'Chorshanba', 'Payshanba', 'Juma', 'Shanba', 'Yakshanba']
    time_slots = ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00', '18:00', '19:00', '20:00']
    
    plt.figure(figsize=(14, 8))
    ax = plt.gca()
    
    colors_map = {
        'IT': '#E3F2FD',
        'Koreys tili': '#E8F5E9',
        'Ofis xodimi': '#FFF3E0'
    }
    border_map = {
        'IT': '#1565C0', 
        'Koreys tili': '#2E7D32', 
        'Ofis xodimi': '#EF6C00'
    }

    for i in range(len(days) + 1):
        plt.axvline(i, color='gray', linestyle='--', alpha=0.3)
    for i in range(len(time_slots) + 1):
        plt.axhline(i, color='gray', linestyle='--', alpha=0.3)

    found_any = False
    for sid, data in schedules.items():
        if data['branch'] == branch_name:
            found_any = True
            uid = data['user_id']
            spec = user_specialty.get(uid, 'IT')
            t_name = user_names.get(uid, "Noma'lum")
            
            for day, t_val in data['days'].items():
                if day in days:
                    day_idx = days.index(day)
                    try:
                        h, m = map(int, t_val.split(':'))
                        start_y = h + (m/60)
                        y_pos = len(time_slots) - (start_y - 8)
                        
                        rect = plt.Rectangle((day_idx + 0.05, y_pos - 0.9), 0.9, 0.8, 
                                            facecolor=colors_map.get(spec, '#F5F5F5'),
                                            edgecolor=border_map.get(spec, 'gray'),
                                            linewidth=1.5, alpha=0.9, zorder=3)
                        ax.add_patch(rect)
                        
                        plt.text(day_idx + 0.5, y_pos - 0.5, f"{t_name}\n({t_val})\n{spec}", 
                                 ha='center', va='center', fontsize=8, fontweight='bold', zorder=4)
                    except: continue

    plt.xticks(np.arange(0.5, len(days), 1), days, fontweight='bold')
    plt.yticks(np.arange(0.5, len(time_slots), 1), time_slots[::-1], fontweight='bold')
    
    plt.title(f"🏢 {branch_name} - Haftalik Bandlik Jadvali", fontsize=16, pad=20, fontweight='bold', color='#1A237E')
    plt.xlim(0, len(days))
    plt.ylim(0, len(time_slots))
    
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], marker='s', color='w', label='IT Bo\'limi', markerfacecolor='#E3F2FD', markersize=15, markeredgecolor='#1565C0'),
        Line2D([0], [0], marker='s', color='w', label='Koreys tili', markerfacecolor='#E8F5E9', markersize=15, markeredgecolor='#2E7D32'),
        Line2D([0], [0], marker='s', color='w', label='Ofis xodimi', markerfacecolor='#FFF3E0', markersize=15, markeredgecolor='#EF6C00')
    ]
    ax.legend(handles=legend_elements, loc='upper center', bbox_to_anchor=(0.5, -0.05), ncol=3)

    img_buf = io.BytesIO()
    plt.savefig(img_buf, format='png', bbox_inches='tight', dpi=150)
    img_buf.seek(0)
    plt.close()
    return img_buf, found_any

@dp.callback_query(F.data == "admin_visual_schedule")
async def visual_schedule_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    builder = InlineKeyboardBuilder()
    for loc in LOCATIONS:
        builder.row(InlineKeyboardButton(text=loc['name'], callback_data=f"v_br_{loc['name']}"))
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    
    await callback.message.edit_text(
        "Qaysi filialning vizual bandlik jadvalini ko'rmoqchisiz?",
        reply_markup=builder.as_markup()
    )
    await state.set_state(VisualSchedule.selecting_branch)
    await callback.answer()

@dp.callback_query(VisualSchedule.selecting_branch, F.data.startswith("v_br_"))
async def visual_schedule_process(callback: types.CallbackQuery, state: FSMContext):
    branch_name = callback.data.replace("v_br_", "")
    await callback.message.answer(f"⏳ {branch_name} uchun vizual jadval tayyorlanmoqda...")
    
    try:
        img_buf, found = await create_visual_timetable_img(branch_name)
        if not found:
            await callback.message.answer(f"📭 {branch_name} filialida hali darslar belgilanmagan.")
        else:
            await callback.message.answer_photo(
                types.BufferedInputFile(img_buf.read(), filename="timetable.png"),
                caption=f"🖼 {branch_name} filialining haftalik bandlik xaritasi."
            )
    except Exception as e:
        logging.error(f"Visual schedule error: {e}")
        traceback.print_exc()
        await callback.message.answer("❌ Jadvalni chizishda xatolik yuz berdi.")
    
    await state.clear()
    await callback.answer()

# --- OYLIK KALKULYATOR HANDLERS (TUZATILGAN VERSIYA)---
@dp.callback_query(F.data == "admin_salary_calc")
async def salary_calc_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="💻 IT", callback_data="sal_spec_IT"),
        InlineKeyboardButton(text="🇰🇷 Koreys tili", callback_data="sal_spec_Koreys tili")
    )
    builder.row(
        InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
    )
    await callback.message.edit_text(
        "Qaysi fan o'qituvchisiga oylik hisoblamoqchisiz?",
        reply_markup=builder.as_markup()
    )
    await state.set_state(SalaryCalc.selecting_specialty)
    await callback.answer()

@dp.callback_query(SalaryCalc.selecting_specialty, F.data.startswith("sal_spec_"))
async def salary_calc_spec(callback: types.CallbackQuery, state: FSMContext):
    spec = callback.data.replace("sal_spec_", "")
    await state.update_data(specialty=spec)
    
    builder = InlineKeyboardBuilder()
    for uid, name in user_names.items():
        if user_specialty.get(uid) == spec:
            builder.row(InlineKeyboardButton(text=str(name), callback_data=f"sal_teacher_{uid}"))
    
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_salary_calc"))
    await callback.message.edit_text(
        f"{spec} o'qituvchisini tanlang:",
        reply_markup=builder.as_markup()
    )
    await state.set_state(SalaryCalc.selecting_teacher)
    await callback.answer()

@dp.callback_query(SalaryCalc.selecting_teacher, F.data.startswith("sal_teacher_"))
async def salary_calc_teacher_selected(callback: types.CallbackQuery, state: FSMContext):
    try:

        uid = int(callback.data.replace("sal_teacher_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    
    # O'qituvchining barcha filiallarini groups dan topamiz
    branches = list({gdata['branch'] for gdata in groups.values() if gdata.get('teacher_id') == uid})
    
    if not branches:
        await callback.message.answer("❌ Bu o'qituvchiga hali guruh biriktirilmagan!")
        await state.clear()
        return

    await state.update_data(
        teacher_id=uid,
        teacher_name=user_names.get(uid, f"ID: {uid}"),
        specialty=user_specialty.get(uid, ''),
        all_branches=branches,
        current_branch_idx=0,
        calculated_results=[]
    )
    
    await salary_ask_next_branch(callback.message, state)
    await callback.answer()

async def salary_ask_next_branch(message: types.Message, state: FSMContext):
    data = await state.get_data()
    idx = data['current_branch_idx']
    branches = data['all_branches']
    current_branch = branches[idx]
    
    await message.answer(f"🏢 Filial: {current_branch}\n\nUshbu filialdagi o'quvchilar sonini kiriting:")
    await state.set_state(SalaryCalc.entering_students)

@dp.message(SalaryCalc.entering_students)
async def salary_students_step(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Raqam kiriting!")
        return
    await state.update_data(temp_students=int(message.text))
    await message.answer("Ushbu filialda bu oy necha marta dars o'tdi?")
    await state.set_state(SalaryCalc.entering_lessons)

@dp.message(SalaryCalc.entering_lessons)
async def salary_lessons_step(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Raqam kiriting!")
        return
    await state.update_data(temp_lessons=int(message.text))
    
    data = await state.get_data()
    builder = InlineKeyboardBuilder()
    if data['specialty'] == "IT":
        builder.row(InlineKeyboardButton(text="35%", callback_data="p_it_35"),
                    InlineKeyboardButton(text="45%", callback_data="p_it_45"))
        await message.answer("Imtixon natijasini tanlang:", reply_markup=builder.as_markup())
    else:
        for p in range(10, 101, 10):
            builder.add(InlineKeyboardButton(text=f"{p}%", callback_data=f"p_kr_{p}"))
        builder.adjust(3)
        await message.answer("Imtixon natijasini tanlang (%):", reply_markup=builder.as_markup())
    await state.set_state(SalaryCalc.selecting_percentage)

@dp.callback_query(SalaryCalc.selecting_percentage)
async def salary_perc_step(callback: types.CallbackQuery, state: FSMContext):
    try:

        perc = int(callback.data.split('_')[-1])

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    await state.update_data(temp_perc=perc)
    
    data = await state.get_data()
    if data['specialty'] == "IT":
        await callback.message.edit_text("Ushbu filial uchun jarima FOIZINI kiriting (masalan: 10. Jarima bo'lmasa 0):")
        await state.set_state(SalaryCalc.entering_penalty_it_percent)
    else:
        await callback.message.edit_text("Ushbu filial uchun jarima SUMMASINI kiriting (so'mda, masalan: 50000. Jarima bo'lmasa 0):")
        await state.set_state(SalaryCalc.entering_penalty_kr_sum)
    await callback.answer()

@dp.message(SalaryCalc.entering_penalty_it_percent)
async def salary_it_penalty_percent(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Raqamda kiriting!")
        return
    await state.update_data(temp_penalty_val=int(message.text))
    await message.answer("Ushbu filialdan jami o'quvchilar to'lovini kiriting:")
    await state.set_state(SalaryCalc.entering_payment)

@dp.message(SalaryCalc.entering_penalty_kr_sum)
async def salary_kr_penalty_sum(message: types.Message, state: FSMContext):
    val = message.text.replace(' ', '').replace(',', '')
    if not val.isdigit():
        await message.answer("Raqamda kiriting!")
        return
    await state.update_data(temp_penalty_val=int(val))
    await process_branch_calculation(message, state)

@dp.message(SalaryCalc.entering_payment)
async def salary_payment_it_step(message: types.Message, state: FSMContext):
    val = message.text.replace(' ', '').replace(',', '')
    if not val.isdigit():
        await message.answer("Raqam kiriting!")
        return
    await state.update_data(temp_payment=int(val))
    await process_branch_calculation(message, state)

async def process_branch_calculation(message: types.Message, state: FSMContext):
    data = await state.get_data()
    spec = data['specialty']
    idx = data['current_branch_idx']
    branch_name = data['all_branches'][idx]
    
    gross = 0
    exam_pen = 0
    penalty_disp = ""

    if spec == "IT":
        share_amount = (data['temp_payment'] * data['temp_perc'] / 100)
        penalty_amount = (share_amount * data['temp_penalty_val'] / 100)
        gross = share_amount - penalty_amount
        penalty_disp = f"{data['temp_penalty_val']}%"
    else:
        students = data['temp_students']
        lessons = data['temp_lessons']
        perc = data['temp_perc']
        base = 1800000 + (students * 100000 if students > 10 else 0)
        
        exam_pen = get_kr_exam_penalty(perc)
        mid_total = base - exam_pen
        
        if lessons < 12:
            gross = (mid_total / 12) * lessons
        else:
            gross = mid_total
        
        gross -= data['temp_penalty_val']
        # Formatlashni alohida bajaramiz
        penalty_val = data['temp_penalty_val']
        penalty_disp = f"{penalty_val:,} so'm".replace(',', ' ')

    res = {
        'branch': branch_name,
        'students': data['temp_students'],
        'lessons': data['temp_lessons'],
        'perc': data['temp_perc'],
        'penalty_display': penalty_disp,
        'exam_penalty': exam_pen,
        'payment': data.get('temp_payment', 0),
        'gross': gross
    }
    
    results_list = data['calculated_results']
    results_list.append(res)
    
    new_idx = idx + 1
    if new_idx < len(data['all_branches']):
        await state.update_data(current_branch_idx=new_idx, calculated_results=results_list)
        await salary_ask_next_branch(message, state)
    else:
        await finalize_multi_branch_salary(message, state, results_list)

async def finalize_multi_branch_salary(message: types.Message, state: FSMContext, results):
    data = await state.get_data()
    total_gross = sum(r['gross'] for r in results)
    tax = total_gross * 0.075
    net = total_gross - tax
    
    excel_file = await create_multi_branch_excel(data['teacher_name'], data['specialty'], results, total_gross, tax, net)
    
    # Formatlashni alohida bajaramiz - f-string ichida replace ishlatilmadi
    s_net = "{:,.0f}".format(net).replace(',', ' ')
    s_tax = "{:,.0f}".format(tax).replace(',', ' ')
    s_gross = "{:,.0f}".format(total_gross).replace(',', ' ')
    
    caption = (f"💰 Hisob-kitob yakunlandi\n\n"
               f"👤 Xodim: {data['teacher_name']}\n"
               f"📚 Mutaxassislik: {data['specialty']}\n"
               f"🏢 Filiallar: {len(results)} ta\n"
               f"──────────────────\n"
               f"Jami (soliqsiz): {s_gross} so'm\n"
               f"Soliq (7.5%): {s_tax} so'm\n"
               f"Qo'lga tegadi: {s_net} so'm")
    
    await message.answer_document(
        types.BufferedInputFile(excel_file.read(), filename=f"Oylik_{data['teacher_name']}_{datetime.now(UZB_TZ).strftime('%Y%m')}.xlsx"),
        caption=caption
    )
    await state.clear()
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🔄 Yangi hisoblash", callback_data="admin_salary_calc"))
    builder.row(InlineKeyboardButton(text="🔙 Admin Panel", callback_data="admin_back"))
    await message.answer(
        "Boshqa amalni tanlang:",
        reply_markup=builder.as_markup()
    )

async def create_multi_branch_excel(teacher_name, specialty, results, total_gross, tax, net):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Oylik Hisoboti"
    
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ['Filial', 'O\'quvchilar', 'Darslar', 'Imtixon %', 'Jarima', 'Tushum', 'Hisoblangan oylik']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r in results:
        row = [
            r['branch'],
            r['students'],
            r['lessons'],
            f"{r['perc']}%",
            r['penalty_display'],
            f"{r['payment']:,.0f}" if r['payment'] > 0 else "—",
            f"{r['gross']:,.0f}"
        ]
        ws.append(row)
        
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    ws.append([])
    
    # Formatlashni alohida bajaramiz
    total_gross_fmt = f"{total_gross:,.0f}".replace(',', ' ')
    tax_fmt = f"{tax:,.0f}".replace(',', ' ')
    net_fmt = f"{net:,.0f}".replace(',', ' ')
    
    summary_rows = [
        ['', '', '', '', '', 'JAMI (soliqsiz):', total_gross_fmt],
        ['', '', '', '', '', 'Soliq (7.5%):', tax_fmt],
        ['', '', '', '', '', 'QO\'LGA TEGADI:', net_fmt]
    ]
    
    for s_row in summary_rows:
        ws.append(s_row)
        for cell in ws[ws.max_row]:
            cell.border = border
            if s_row[5] in ['JAMI (soliqsiz):', 'Soliq (7.5%):', 'QO\'LGA TEGADI:']:
                cell.font = Font(bold=True)
                if s_row[5] == 'QO\'LGA TEGADI:':
                    ws.cell(row=ws.max_row, column=7).font = Font(bold=True, color="006100")
                    ws.cell(row=ws.max_row, column=7).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    column_widths = [20, 12, 10, 10, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@dp.callback_query(F.data == "admin_monthly_report")
async def admin_monthly_report_start(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    now = datetime.now(UZB_TZ)
    start_date = datetime(2026, 3, 1, tzinfo=UZB_TZ) 
    
    builder = InlineKeyboardBuilder()
    months_uz = {1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun", 
                 7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"}
    
    temp_date = start_date
    while temp_date <= now:
        month_name = months_uz[temp_date.month]
        btn_text = f"📅 {month_name} {temp_date.year}"
        builder.row(InlineKeyboardButton(text=btn_text, callback_data=f"gen_month_{temp_date.year}_{temp_date.month}"))
        
        if temp_date.month == 12:
            temp_date = datetime(temp_date.year + 1, 1, 1, tzinfo=UZB_TZ)
        else:
            temp_date = datetime(temp_date.year, temp_date.month + 1, 1, tzinfo=UZB_TZ)
            
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    await callback.message.edit_text("Qaysi oy uchun hisobot kerak?", reply_markup=builder.as_markup())
    await callback.answer()

async def create_monthly_grouped_pdf(year: int, month: int) -> io.BytesIO:
    """Oylik davomat hisoboti — Excel format"""
    import calendar
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    months_uz = {1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun",
                 7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"}
    ws.title = f"{months_uz[month]} {year}"
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    # Sarlavha
    ws.merge_cells('A1:I1')
    ws['A1'] = f"OYLIK DAVOMAT HISOBOTI — {months_uz[month].upper()} {year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    _, last_day = calendar.monthrange(year, month)
    headers = ['№', 'Sana', 'Hafta kuni', "O'qituvchi", 'Filial', 'Dars vaqti', 'Kelgan vaqti', 'Holat', 'Kechikish']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    row = 4
    num = 1
    for uid in sorted(user_ids):
        name = user_names.get(uid, f"ID:{uid}")
        for branch, gdata_list in {g['branch']: [] for g in groups.values() if g.get('teacher_id') == uid}.items():
            branch_groups = [gd for gd in groups.values() if gd.get('teacher_id') == uid and gd.get('branch') == branch]
            day_time_map = {}
            for gd in branch_groups:
                dt = gd.get('day_times', {})
                if dt:
                    day_time_map.update(dt)
                else:
                    for day in gd.get('days', []):
                        if day not in day_time_map:
                            day_time_map[day] = gd.get('time', '—')
            for d in range(1, last_day + 1):
                target_date = f"{year}-{month:02d}-{d:02d}"
                d_obj = d_date(year, month, d)
                weekday = WEEKDAYS_UZ[d_obj.weekday()]
                if weekday not in day_time_map:
                    continue
                sch_time = day_time_map[weekday]
                att = next((a for a in daily_attendance_log if a[0] == uid and a[1] == branch and a[2] == target_date), None)
                if att:
                    ontime, mins = calculate_lateness(att[3], sch_time)
                    status_text = "Vaqtida" if ontime else "Kechikkan"
                    att_time = att[3][:5]
                    late_m = 0 if ontime else mins
                else:
                    status_text = "KELMAGAN"
                    att_time = "—"
                    late_m = "—"
                row_vals = [num, target_date, weekday, name, branch, sch_time, att_time, status_text, late_m]
                for col_idx, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    if col_idx == 8:
                        if status_text == "Kechikkan":
                            cell.font = Font(color="FF0000", bold=True)
                        elif status_text == "Vaqtida":
                            cell.font = Font(color="008000", bold=True)
                        elif status_text == "KELMAGAN":
                            cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")
                num += 1
                row += 1
    for i, w in enumerate([5, 12, 13, 22, 20, 10, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@dp.callback_query(F.data.startswith("gen_month_"))
async def process_month_gen(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    _, _, year, month = callback.data.split("_")
    await callback.message.answer(f"⏳ {month}/{year} uchun hisobot tayyorlanmoqda...")
    excel = await create_monthly_grouped_pdf(int(year), int(month))
    await callback.message.answer_document(
        types.BufferedInputFile(excel.read(), filename=f"hisobot_{year}_{month}.xlsx"),
        caption=f"📊 {month}/{year} oylik davomat hisoboti"
    )
    await callback.answer()

# --- PROFESSIONAL EXCEL HISOBOT (FILIALLAR BO'YICHA GURUHLANGAN) ---
async def create_monthly_excel(year: int, month: int) -> io.BytesIO:
    import calendar
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(border_style="thin", color="000000")
    all_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    main_header_fill = PatternFill(start_color="92D050", fill_type="solid") # Yashil sarlavha
    user_header_fill = PatternFill(start_color="D9D9D9", fill_type="solid") # Kulrang o'qituvchi nomi
    branch_header_fill = PatternFill(start_color="FDE9D9", fill_type="solid") # Och jigarrang filial nomi
    table_header_fill = PatternFill(start_color="2E86AB", fill_type="solid") # Ko'k jadval header
    
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    
    months_uz = {1: "YANVAR", 2: "FEVRAL", 3: "MART", 4: "APREL", 5: "MAY", 6: "IYUN", 
                 7: "IYUL", 8: "AVGUST", 9: "SENTABR", 10: "OKTABR", 11: "NOYABR", 12: "DEKABR"}
    
    specs = ["IT", "Koreys tili", "Ofis xodimi"]
    _, last_day = calendar.monthrange(year, month)

    for spec in specs:
        ws = wb.create_sheet(title=spec)
        
        # 1. Asosiy sarlavha
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        main_title = ws.cell(row=1, column=1)
        main_title.value = f"{spec.upper()} OQITUVCHILARI {months_uz[month]} OYI XISOBOTI"
        main_title.fill = main_header_fill
        main_title.font = Font(size=14, bold=True)
        main_title.alignment = Alignment(horizontal="center")
        for col in range(1, 10):
            ws.cell(row=1, column=col).border = all_border
        
        current_row = 3
        teachers = [uid for uid, s in user_specialty.items() if s == spec]
        
        if not teachers:
            continue

        for uid in sorted(teachers, key=lambda x: user_names.get(x, "")):
            # 2. O'qituvchi nomi (Kulrang)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            user_title = ws.cell(row=current_row, column=1)
            user_title.value = f"👤 {user_names.get(uid, '').upper()} {months_uz[month]} OYLIK XISOBOT"
            user_title.fill = user_header_fill
            user_title.font = bold_font
            user_title.alignment = Alignment(horizontal="center")
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).border = all_border
            current_row += 1

            # 3. Ushbu o'qituvchi ishlaydigan filiallarni aniqlash
            teacher_branches = sorted(list({gdata['branch'] for gdata in groups.values() if gdata.get('teacher_id') == uid}))
            
            for branch in teacher_branches:
                # 4. Filial sarlavhasi (Och jigarrang)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
                branch_title = ws.cell(row=current_row, column=1)
                branch_title.value = f"🏢 Filial: {branch}"
                branch_title.fill = branch_header_fill
                branch_title.font = Font(italic=True, bold=True)
                branch_title.alignment = Alignment(horizontal="center")
                for col in range(1, 10):
                    ws.cell(row=current_row, column=col).border = all_border
                current_row += 1

                # 5. Jadval headeri
                headers = ['№', 'Sana', 'Hafta kuni', 'O\'qituvchi', 'Filial', 'Dars vaqti', 'Kelgan vaqti', 'Holat', 'Kechikish']
                for col, text in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = text
                    cell.fill = table_header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = all_border
                current_row += 1

                # 6. Ma'lumotlarni yozish (groups dict dan)
                # Bu filial+o'qituvchiga tegishli guruhlarni topamiz
                branch_groups = [
                    gd for gd in groups.values()
                    if gd.get('teacher_id') == uid and gd.get('branch') == branch
                ]
                # Dars kunlari va vaqtlari: {weekday: time}
                day_time_map = {}
                for gd in branch_groups:
                    dt = gd.get('day_times', {})
                    if dt:
                        day_time_map.update(dt)
                    else:
                        # eski format: days list + time
                        for day in gd.get('days', []):
                            if day not in day_time_map:
                                day_time_map[day] = gd.get('time', '—')

                # Guruh yaratilgan sana (shu oydagi birinchi kun hisoblanadi)
                grp_created = None
                for gd in branch_groups:
                    cat = gd.get('created_at')
                    if cat:
                        if grp_created is None or cat < grp_created:
                            grp_created = cat

                num = 1
                for d in range(1, last_day + 1):
                    target_date = f"{year}-{month:02d}-{d:02d}"
                    d_obj = d_date(year, month, d)
                    weekday = WEEKDAYS_UZ[d_obj.weekday()]

                    if weekday not in day_time_map:
                        continue

                    # Guruh yaratilgan kundan oldingi darslarni o'tkazib yuborish
                    if grp_created is not None:
                        created_date = grp_created.date() if hasattr(grp_created, 'date') else grp_created
                        if d_obj < created_date:
                            continue

                    sch_time = day_time_map[weekday]
                    att = next(
                        (a for a in daily_attendance_log
                         if a[0] == uid and a[1] == branch and a[2] == target_date),
                        None
                    )

                    if att:
                        ontime, mins = calculate_lateness(att[3], sch_time)
                        status_text = "Vaqtida" if ontime else "Kechikkan"
                        att_time = att[3][:5]  # HH:MM
                        late_m = 0 if ontime else mins
                    else:
                        status_text = "KELMAGAN"
                        att_time = "—"
                        late_m = "—"

                    row_vals = [num, target_date, weekday,
                                user_names.get(uid, ''), branch,
                                sch_time, att_time, status_text, late_m]
                    for col_idx, val in enumerate(row_vals, 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = val
                        cell.border = all_border
                        cell.alignment = Alignment(horizontal="center")
                        if col_idx == 8:
                            if status_text == "Kechikkan":
                                cell.font = Font(color="FF0000", bold=True)
                            elif status_text == "Vaqtida":
                                cell.font = Font(color="008000", bold=True)
                            elif status_text == "KELMAGAN":
                                cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")

                    num += 1
                    current_row += 1

                current_row += 1  # Filiallar orasida kichik masofa
            current_row += 2 # O'qituvchilar orasida katta masofa

        # Ustun kengligi
        for i in range(1, 10):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@dp.callback_query(F.data == "admin_excel_menu")
async def admin_excel_report_start(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    now = datetime.now(UZB_TZ)
    start_date = datetime(2026, 3, 1, tzinfo=UZB_TZ) 
    
    builder = InlineKeyboardBuilder()
    months_uz = {1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun", 
                 7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"}
    
    temp_date = start_date
    while temp_date <= now:
        month_name = months_uz[temp_date.month]
        btn_text = f"📊 Excel: {month_name} {temp_date.year}"
        builder.row(InlineKeyboardButton(text=btn_text, callback_data=f"get_excel_{temp_date.year}_{temp_date.month}"))
        
        if temp_date.month == 12:
            temp_date = datetime(temp_date.year + 1, 1, 1, tzinfo=UZB_TZ)
        else:
            temp_date = datetime(temp_date.year, temp_date.month + 1, 1, tzinfo=UZB_TZ)
            
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    await callback.message.edit_text("Qaysi oy uchun Excel hisobot kerak?", reply_markup=builder.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("get_excel_"))
async def process_excel_download(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    _, _, year, month = callback.data.split("_")
    await callback.message.answer(f"⏳ {month}-{year} uchun Excel tayyorlanmoqda...")
    
    try:
        excel_buf = await create_monthly_excel(int(year), int(month))
        
        filename = f"Davomat_{year}_{month}.xlsx"
        await callback.message.answer_document(
            types.BufferedInputFile(excel_buf.read(), filename=filename),
            caption=f"📈 {month}-{year} oyi uchun buxgalteriya hisoboti tayyor."
        )
    except Exception as e:
        logging.error(f"Excel yaratishda xatolik: {e}")
        await callback.message.answer(f"❌ Excel yaratishda xatolik: {e}")
    
    await callback.answer()

@dp.callback_query(F.data == "admin_pdf_menu")
async def admin_pdf_menu(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="📊 Umumiy statistika", callback_data="pdf_general"),
        InlineKeyboardButton(text="🏆 Filiallar reytingi", callback_data="pdf_branches")
    )
    builder.row(
        InlineKeyboardButton(text="👥 O'qituvchilar reytingi", callback_data="pdf_teachers"),
        InlineKeyboardButton(text="📅 Oylik hisobot", callback_data="pdf_monthly")
    )
    builder.row(
        InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
    )
    
    await callback.message.edit_text(
        "📊 Statistika hisobotlari\n\nKerakli hisobot turini tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()

async def create_general_stats_pdf() -> io.BytesIO:
    """Umumiy statistika — Excel format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Umumiy statistika"
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    now_uzb = datetime.now(UZB_TZ)
    today = now_uzb.strftime("%Y-%m-%d")
    current_month = now_uzb.strftime("%Y-%m")
    ws.merge_cells('A1:B1')
    ws['A1'] = f"UMUMIY STATISTIKA — {now_uzb.strftime('%d.%m.%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = hdr_fill
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A2'].border = ws['B2'].border = border
    data = [
        ("Ko'rsatkich", "Qiymat"),
        ("Jami foydalanuvchilar", len(user_ids)),
        ("Faol foydalanuvchilar", len([u for u in user_ids if user_status.get(u) == 'active'])),
        ("Bloklangan", len([u for u in user_ids if user_status.get(u) == 'blocked'])),
        ("Jami davomatlar", len(daily_attendance_log)),
        ("Bugungi davomatlar", len([k for k in daily_attendance_log if k[2] == today])),
        ("Bu oylik davomatlar", len([k for k in daily_attendance_log if k[2].startswith(current_month)])),
        ("IT o'qituvchilar", len([u for u in user_ids if user_specialty.get(u) == 'IT'])),
        ("Koreys tili o'qituvchilari", len([u for u in user_ids if user_specialty.get(u) == 'Koreys tili'])),
        ("Ofis xodimlari", len([u for u in user_ids if user_specialty.get(u) == 'Ofis xodimi'])),
    ]
    for r, (k, v) in enumerate(data, 3):
        ws.cell(row=r, column=1, value=k).border = border
        ws.cell(row=r, column=2, value=v).border = border
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='left')
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        if r == 3:
            ws.cell(row=r, column=1).font = Font(bold=True, color='FFFFFF')
            ws.cell(row=r, column=2).font = Font(bold=True, color='FFFFFF')
            ws.cell(row=r, column=1).fill = hdr_fill
            ws.cell(row=r, column=2).fill = hdr_fill
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

async def create_branches_stats_pdf() -> io.BytesIO:
    """Filiallar reytingi — Excel format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Filiallar reytingi"
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    ws.merge_cells('A1:C1')
    ws['A1'] = f"FILIALLAR REYTINGI — {datetime.now(UZB_TZ).strftime('%d.%m.%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = hdr_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    for col, h in enumerate(['№', 'Filial', 'Davomatlar soni'], 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    branch_stats = defaultdict(int)
    for (uid, branch, date, time) in daily_attendance_log:
        branch_stats[branch] += 1
    for i, (branch, count) in enumerate(sorted(branch_stats.items(), key=lambda x: x[1], reverse=True), 1):
        for col, val in enumerate([i, branch, count], 1):
            cell = ws.cell(row=3+i, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

async def create_teachers_stats_pdf() -> io.BytesIO:
    """O'qituvchilar reytingi — Excel format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "O'qituvchilar reytingi"
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    ws.merge_cells('A1:D1')
    ws['A1'] = f"O'QITUVCHILAR REYTINGI — {datetime.now(UZB_TZ).strftime('%d.%m.%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = hdr_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    for col, h in enumerate(["№", "O'qituvchi", "Mutaxassislik", "Davomatlar"], 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    teacher_stats = defaultdict(int)
    for (uid, branch, date, time) in daily_attendance_log:
        teacher_stats[uid] += 1
    for i, (uid, count) in enumerate(sorted(teacher_stats.items(), key=lambda x: x[1], reverse=True)[:50], 1):
        row_vals = [i, user_names.get(uid, f"ID:{uid}"), user_specialty.get(uid, ''), count]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=3+i, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

async def create_monthly_stats_pdf() -> io.BytesIO:
    """Oylik statistika — Excel format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Oylik statistika"
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    now_uzb = datetime.now(UZB_TZ)
    current_month = now_uzb.strftime("%Y-%m")
    months_uz = {1:"Yanvar",2:"Fevral",3:"Mart",4:"Aprel",5:"May",6:"Iyun",
                 7:"Iyul",8:"Avgust",9:"Sentabr",10:"Oktabr",11:"Noyabr",12:"Dekabr"}
    month_name = f"{months_uz[now_uzb.month]} {now_uzb.year}"
    ws.merge_cells('A1:D1')
    ws['A1'] = f"{month_name.upper()} OYI STATISTIKASI — {now_uzb.strftime('%d.%m.%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = hdr_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    for col, h in enumerate(["№", "Filial", "Jami davomatlar", "O'qituvchilar"], 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    monthly_stats = defaultdict(lambda: defaultdict(int))
    for (uid, branch, date, time) in daily_attendance_log:
        if date.startswith(current_month):
            monthly_stats[branch][uid] += 1
    total_att = 0
    total_teachers = set()
    for i, (branch, users) in enumerate(sorted(monthly_stats.items(), key=lambda x: sum(x[1].values()), reverse=True), 1):
        total = sum(users.values())
        total_att += total
        total_teachers.update(users.keys())
        row_vals = [i, branch, total, len(users)]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=3+i, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    # Jami qator
    last_r = 3 + len(monthly_stats) + 1
    for col, val in enumerate(["", "UMUMIY:", total_att, len(total_teachers)], 1):
        cell = ws.cell(row=last_r, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', fill_type='solid')
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

async def handle_pdf_reports(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    report_type = callback.data.replace("pdf_", "")
    
    await callback.message.edit_text("⏳ Excel hisobot tayyorlanmoqda, biroz kuting...")
    
    try:
        if report_type == "general":
            buf = await create_general_stats_pdf()
            filename = f"umumiy_statistika_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
            caption = "📊 Umumiy statistika hisoboti"
            
        elif report_type == "branches":
            buf = await create_branches_stats_pdf()
            filename = f"filiallar_reytingi_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
            caption = "🏆 Filiallar reytingi"
            
        elif report_type == "teachers":
            buf = await create_teachers_stats_pdf()
            filename = f"oqituvchilar_reytingi_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
            caption = "👥 Eng faol o'qituvchilar"
            
        elif report_type == "monthly":
            buf = await create_monthly_stats_pdf()
            filename = f"oylik_hisobot_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
            caption = f"📅 {datetime.now(UZB_TZ).strftime('%B %Y')} oyi hisoboti"
            
        else:
            await callback.message.edit_text("❌ Noto'g'ri so'rov")
            await callback.answer()
            return
        
        await bot.send_document(
            callback.message.chat.id,
            types.BufferedInputFile(buf.getvalue(), filename=filename),
            caption=caption
        )
        
    except Exception as e:
        logging.error(f"Excel yaratishda xatolik: {e}")
        await callback.message.edit_text(f"❌ Excel yaratishda xatolik: {str(e)}")
    
    await callback.answer()

@dp.callback_query(F.data == "admin_stats_main")
async def admin_stats_main(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="📊 Umumiy statistika", callback_data="admin_stats_general"),
            InlineKeyboardButton(text="🏆 Filiallar reytingi", callback_data="admin_stats_branches")
        )
        builder.row(
            InlineKeyboardButton(text="👥 O'qituvchilar reytingi", callback_data="admin_stats_teachers"),
            InlineKeyboardButton(text="📅 Oylik hisobot", callback_data="admin_monthly")
        )
        builder.row(
            InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
        )
        
        await callback.message.edit_text(
            "📊 Statistika bo'limi\n\nKerakli statistikani tanlang:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_stats_main error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(F.data == "admin_stats_general")
async def admin_stats_general(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        now_uzb = datetime.now(UZB_TZ)
        today = now_uzb.strftime("%Y-%m-%d")
        current_month = now_uzb.strftime("%Y-%m")
        
        total_users = len(user_ids)
        active_users = len([uid for uid in user_ids if user_status.get(uid) == 'active'])
        blocked_users = len([uid for uid in user_ids if user_status.get(uid) == 'blocked'])
        total_attendances = len(daily_attendance_log)
        today_attendances = len([k for k in daily_attendance_log if k[2] == today])
        monthly_attendances = len([k for k in daily_attendance_log if k[2].startswith(current_month)])
        
        it_teachers = len([uid for uid in user_ids if user_specialty.get(uid) == 'IT'])
        korean_teachers = len([uid for uid in user_ids if user_specialty.get(uid) == 'Koreys tili'])
        office_workers = len([uid for uid in user_ids if user_specialty.get(uid) == 'Ofis xodimi'])
        
        branch_stats = defaultdict(int)
        for (uid, branch, date, time) in daily_attendance_log:
            branch_stats[branch] += 1
        top_branch = max(branch_stats.items(), key=lambda x: x[1]) if branch_stats else ("Yo'q", 0)
        
        teacher_stats = defaultdict(int)
        for (uid, branch, date, time) in daily_attendance_log:
            teacher_stats[uid] += 1
        top_teacher_id = max(teacher_stats.items(), key=lambda x: x[1]) if teacher_stats else (None, 0)
        top_teacher_name = user_names.get(top_teacher_id[0], "Noma'lum") if top_teacher_id[0] else "Yo'q"
        top_teacher_specialty = user_specialty.get(top_teacher_id[0], '')
        top_teacher_display = f"{top_teacher_name}[{top_teacher_specialty}]" if top_teacher_specialty else top_teacher_name
        
        text = f"""
📊 Umumiy statistika

👥 Foydalanuvchilar:
• Jami: {total_users}
• Faol: {active_users}
• Bloklangan: {blocked_users}
• 💻 IT: {it_teachers}
• 🇰🇷 Koreys tili: {korean_teachers}
• 🏢 Ofis xodimlari: {office_workers}

📋 Davomatlar:
• Jami: {total_attendances}
• Bugun: {today_attendances}
• Shu oyda: {monthly_attendances}

🏆 Eng faol filial: {top_branch[0]} ({top_branch[1]} ta)

👑 Eng faol o'qituvchi: {top_teacher_display} ({top_teacher_id[1]} ta)

📅 Oxirgi yangilanish: {now_uzb.strftime('%Y-%m-%d %H:%M')}
"""
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_stats_main"))
        
        await callback.message.edit_text(
            text,
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_stats_general error: {e}")
        await callback.message.edit_text("❌ Statistikani olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_stats_branches")
async def admin_stats_branches(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        branch_stats = defaultdict(int)
        for (uid, branch, date, time) in daily_attendance_log:
            branch_stats[branch] += 1
        
        if not branch_stats:
            await callback.message.edit_text("📭 Hali davomat ma'lumotlari yo'q.")
            await callback.answer()
            return
        
        sorted_branches = sorted(branch_stats.items(), key=lambda x: x[1], reverse=True)
        
        text = "🏆 Filiallar reytingi\n\n"
        for i, (branch, count) in enumerate(sorted_branches, 1):
            medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            text += f"{medal} {branch}: {count} ta davomat\n"
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_stats_main"))
        
        await callback.message.edit_text(
            text,
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_stats_branches error: {e}")
        await callback.message.edit_text("❌ Filiallar reytingini olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_stats_teachers")
async def admin_stats_teachers(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        teacher_stats = defaultdict(int)
        for (uid, branch, date, time) in daily_attendance_log:
            teacher_stats[uid] += 1
        
        if not teacher_stats:
            await callback.message.edit_text("📭 Hali davomat ma'lumotlari yo'q.")
            await callback.answer()
            return
        
        sorted_teachers = sorted(teacher_stats.items(), key=lambda x: x[1], reverse=True)[:20]
        
        text = "👥 Eng faol o'qituvchilar\n\n"
        for i, (uid, count) in enumerate(sorted_teachers, 1):
            name = user_names.get(uid, f"ID: {uid}")
            specialty = user_specialty.get(uid, '')
            specialty_display = f" [{specialty}]" if specialty else ""
            medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            text += f"{medal} {name}{specialty_display}: {count} ta davomat\n"
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_stats_main"))
        
        await callback.message.edit_text(
            text,
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_stats_teachers error: {e}")
        await callback.message.edit_text("❌ O'qituvchilar reytingini olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_monthly")
async def admin_monthly(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        now_uzb = datetime.now(UZB_TZ)
        current_month = now_uzb.strftime("%Y-%m")
        month_name = now_uzb.strftime("%B %Y")
        
        monthly_stats = defaultdict(lambda: defaultdict(int))
        for (uid, branch, date, time) in daily_attendance_log:
            if date.startswith(current_month):
                monthly_stats[branch][uid] += 1
        
        if not monthly_stats:
            await callback.message.edit_text("📭 Shu oy uchun davomat ma'lumotlari yo'q.")
            await callback.answer()
            return
        
        report = f"📊 {month_name} oyi uchun hisobot\n\n"
        
        for branch, users in monthly_stats.items():
            total = sum(users.values())
            unique_users = len(users)
            report += f"🏢 {branch}\n"
            report += f"   Jami: {total} ta davomat\n"
            report += f"   O'qituvchilar: {unique_users} ta\n\n"
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_stats_main"))
        
        await callback.message.edit_text(
            report,
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_monthly error: {e}")
        await callback.message.edit_text("❌ Oylik hisobotni olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_users_main")
async def admin_users_main(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="✅ Foydalanuvchilar ro'yxati", callback_data="admin_users_active")
        )
        builder.row(
            InlineKeyboardButton(text="⛔ Bloklanganlar", callback_data="admin_users_blocked")
        )
        builder.row(
            InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
        )
        
        await callback.message.edit_text(
            "👥 Foydalanuvchilarni boshqarish\n\nKerakli amalni tanlang:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_users_main error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(F.data.startswith("admin_user_info_"))
async def admin_user_info(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    logging.info(f"admin_user_info handler: {callback.data}")
    
    try:
        try:

            uid = int(callback.data.replace("admin_user_info_", ""))

        except (ValueError, KeyError):

            await callback.answer("Xatolik!", show_alert=True)

            return
        logging.info(f"admin_user_info called for uid: {uid}")
    except ValueError as e:
        logging.error(f"admin_user_info parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    
    name = user_names.get(uid, "Noma'lum")
    status = user_status.get(uid, 'active')
    lang = user_languages.get(uid, 'uz')
    specialty = user_specialty.get(uid, '')
    specialty_display = f" [{specialty}]" if specialty else ""
    
    user_attendances = len([k for k in daily_attendance_log if k[0] == uid])
    user_schedules_count = len([g for g in groups.values() if g.get('teacher_id') == uid])
    
    last_attendance = "Yo'q"
    user_logs = [k for k in daily_attendance_log if k[0] == uid]
    if user_logs:
        last = max(user_logs, key=lambda x: x[2])
        last_attendance = f"{last[2]} {last[3]} ({last[1]})"
    
    text = f"""
👤 Foydalanuvchi ma'lumoti

ID: `{uid}`
Ism: {name}{specialty_display}
Holat: {"✅ Faol" if status == 'active' else "⛔ Bloklangan"}
Til: {lang}

📊 Statistika:
• Jami davomatlar: {user_attendances}
• Guruhlari: {user_schedules_count} ta
• Oxirgi davomat: {last_attendance}
"""
    
    builder = InlineKeyboardBuilder()
    if status == 'active':
        builder.row(InlineKeyboardButton(text="⛔ Bloklash", callback_data=f"admin_user_block_{uid}"))
    else:
        builder.row(InlineKeyboardButton(text="✅ Faollashtirish", callback_data=f"admin_user_unblock_{uid}"))
    builder.row(
        InlineKeyboardButton(text="📊 Statistika", callback_data=f"admin_user_stats_{uid}"),
        InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"admin_user_delete_{uid}")
    )
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_users_main"))
    
    await callback.message.edit_text(
        text,
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("admin_user_block_"))
async def admin_user_block(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        try:

            uid = int(callback.data.replace("admin_user_block_", ""))

        except (ValueError, KeyError):

            await callback.answer("Xatolik!", show_alert=True)

            return
        logging.info(f"admin_user_block called for uid: {uid}")
    except ValueError as e:
        logging.error(f"admin_user_block parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    
    user_status[uid] = 'blocked'
    await db.update_user_status(uid, 'blocked')
    
    await callback.answer("✅ Foydalanuvchi bloklandi!")
    await admin_user_info(callback)

@dp.callback_query(F.data.startswith("admin_user_unblock_"))
async def admin_user_unblock(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        try:

            uid = int(callback.data.replace("admin_user_unblock_", ""))

        except (ValueError, KeyError):

            await callback.answer("Xatolik!", show_alert=True)

            return
        logging.info(f"admin_user_unblock called for uid: {uid}")
    except ValueError as e:
        logging.error(f"admin_user_unblock parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    
    user_status[uid] = 'active'
    await db.update_user_status(uid, 'active')
    
    await callback.answer("✅ Foydalanuvchi faollashtirildi!")
    await admin_user_info(callback)

@dp.callback_query(F.data.startswith("admin_user_delete_confirm_"))
async def admin_user_delete_confirm(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    logging.info(f"admin_user_delete_confirm handler: {callback.data}")
    
    try:
        try:

            uid = int(callback.data.replace("admin_user_delete_confirm_", ""))

        except (ValueError, KeyError):

            await callback.answer("Xatolik!", show_alert=True)

            return
        logging.info(f"admin_user_delete_confirm called for uid: {uid}")
        
        user_name = user_names.get(uid, "Noma'lum")
        user_spec = user_specialty.get(uid, "")
        spec_display = f" [{user_spec}]" if user_spec else ""
        
        await callback.message.edit_text(
            f"⏳ Foydalanuvchi o'chirilmoqda...\n\nID: `{uid}`\nIsm: {user_name}{spec_display}"
        )
        
    except ValueError as e:
        logging.error(f"admin_user_delete_confirm parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    
    try:
        async with db.pool.acquire() as conn:
            # O'chirish emas — arxivlash (statistika saqlansin)
            archived_name = user_names.get(uid, f"ID:{uid}")
            archived_spec = user_specialty.get(uid, '')
            await conn.execute(
                "UPDATE users SET status='deleted', full_name=$1 WHERE user_id=$2",
                f"[ARXIV] {archived_name}", uid
            )
            await conn.execute("DELETE FROM schedules WHERE user_id = $1", uid)
            # Guruhlarni egasiz qoldirmaslik — teacher_id ni NULL qilamiz
            teacher_group_ids = await conn.fetch("SELECT id FROM groups WHERE teacher_id = $1", uid)
            for grp_row in teacher_group_ids:
                await conn.execute("UPDATE groups SET teacher_id=NULL WHERE id=$1", grp_row['id'])
                if grp_row['id'] in groups:
                    groups[grp_row['id']]['teacher_id'] = None
        
        # RAM dan faqat faol ro'yxatdan chiqaramiz — statistika saqlanadi
        if uid in user_ids:
            user_ids.remove(uid)
        user_status[uid] = 'deleted'
        user_names[uid] = f"[ARXIV] {user_names.get(uid, f'ID:{uid}')}"

        if uid in user_schedules:
            for schedule_id in user_schedules[uid]:
                schedules.pop(schedule_id, None)
            user_schedules.pop(uid, None)

        await callback.message.edit_text(
            f"✅ Foydalanuvchi arxivlandi!\n\n"
            f"ID: `{uid}`\n"
            f"Ism: {user_name}{spec_display}\n\n"
            f"📁 Statistika va davomat tarixi saqlab qolindi."
        )
        
        await callback.answer("✅ Foydalanuvchi o'chirildi!")
        
        await asyncio.sleep(2)
        
        active_users = [u for u in user_ids if user_status.get(u) != 'blocked']
        if active_users:
            builder = InlineKeyboardBuilder()
            for u in sorted(active_users)[:15]:
                name = user_names.get(u, f"ID: {u}")
                specialty = user_specialty.get(u, '')
                spec_display = f" [{specialty}]" if specialty else ""
                
                if len(name) > 30:
                    name = name[:27] + "..."
                    
                builder.row(
                    InlineKeyboardButton(
                        text=f"👤 {name}{spec_display}", 
                        callback_data=f"admin_user_info_{u}"
                    )
                )
            
            builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_users_main"))
            
            await callback.message.answer(
                "✅ Faol foydalanuvchilar ro'yxati:",
                reply_markup=builder.as_markup()
            )
        else:
            await callback.message.answer("📭 Faol foydalanuvchilar yo'q.")
        
    except Exception as e:
        logging.error(f"admin_user_delete_confirm error: {e}")
        traceback.print_exc()
        
        await callback.message.edit_text(
            f"❌ Xatolik yuz berdi: {str(e)}\n\n"
            f"Batafsil ma'lumot uchun loglarni tekshiring."
        )
        await callback.answer("Xatolik yuz berdi!")

@dp.callback_query(F.data.regexp(r"^admin_user_delete_\d+$"))
async def admin_user_delete(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    logging.info(f"admin_user_delete handler: {callback.data}")
    
    try:
        try:

            uid = int(callback.data.replace("admin_user_delete_", ""))

        except (ValueError, KeyError):

            await callback.answer("Xatolik!", show_alert=True)

            return
        logging.info(f"admin_user_delete called for uid: {uid}")
        
        user_name = user_names.get(uid, "Noma'lum")
        user_spec = user_specialty.get(uid, "")
        spec_display = f" [{user_spec}]" if user_spec else ""
        
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="✅ Ha, o'chirish", callback_data=f"admin_user_delete_confirm_{uid}"),
            InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"admin_user_info_{uid}")
        )
        
        await callback.message.edit_text(
            f"⚠️ Foydalanuvchini o'chirish\n\n"
            f"ID: `{uid}`\n"
            f"Ism: {user_name}{spec_display}\n\n"
            f"Bu foydalanuvchini butunlay o'chirmoqchimisiz?\n"
            f"Barcha ma'lumotlari (davomatlar, dars jadvallari) ham o'chib ketadi!",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
        
    except ValueError as e:
        logging.error(f"admin_user_delete parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    except Exception as e:
        logging.error(f"admin_user_delete error: {e}")
        traceback.print_exc()
        await callback.answer("Xatolik yuz berdi!")

@dp.callback_query(F.data.startswith("admin_user_stats_"))
async def admin_user_stats(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        uid = int(callback.data.replace("admin_user_stats_", ""))
    except ValueError:
        await callback.answer("Noto'g'ri format!")
        return
    
    name = user_names.get(uid, "Noma'lum")
    
    branch_stats = defaultdict(int)
    month_stats = defaultdict(int)
    
    for (user_id, branch, date, time) in daily_attendance_log:
        if user_id == uid:
            branch_stats[branch] += 1
            month = date[:7]
            month_stats[month] += 1
    
    text = f"📊 {name} statistikasi\n\n"
    
    if branch_stats:
        text += "🏢 Filiallar bo'yicha:\n"
        for branch, count in sorted(branch_stats.items(), key=lambda x: x[1], reverse=True):
            text += f"• {branch}: {count} ta\n"
        text += "\n"
    
    if month_stats:
        text += "📅 Oylar bo'yicha:\n"
        for month, count in sorted(month_stats.items(), reverse=True):
            text += f"• {month}: {count} ta\n"
    else:
        text += "📭 Hali davomat yo'q"
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data=f"admin_user_info_{uid}"))
    
    await callback.message.edit_text(
        text,
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@dp.callback_query(F.data == "admin_users_active")
async def admin_users_active(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    await callback.answer()
    try:
        active = [uid for uid in user_ids if user_status.get(uid) != 'blocked']
        if not active:
            await callback.message.edit_text("📭 Faol foydalanuvchilar yo'q.")
            return

        spec_order = ['IT', 'Koreys tili', 'Ofis xodimi']
        spec_icons = {'IT': '💻', 'Koreys tili': '🇰🇷', 'Ofis xodimi': '🗂'}

        builder = InlineKeyboardBuilder()
        for spec in spec_order:
            spec_users = [uid for uid in active if user_specialty.get(uid) == spec]
            if not spec_users:
                continue
            icon = spec_icons.get(spec, '👤')
            # Bo'lim sarlavhasi (bosilmaydi)
            builder.row(InlineKeyboardButton(
                text=f"━━━ {icon} {spec.upper()} ({len(spec_users)} ta) ━━━",
                callback_data="ignore"
            ))
            for uid in sorted(spec_users, key=lambda x: user_names.get(x, '')):
                name = user_names.get(uid, f"ID:{uid}")
                builder.row(InlineKeyboardButton(
                    text=f"👤 {name}",
                    callback_data=f"admin_user_info_{uid}"
                ))
        # Ixtisosligi belgilanmaganlar
        no_spec = [uid for uid in active if not user_specialty.get(uid)]
        if no_spec:
            builder.row(InlineKeyboardButton(text="━━━ ❓ BELGILANMAGAN ━━━", callback_data="ignore"))
            for uid in no_spec:
                name = user_names.get(uid, f"ID:{uid}")
                builder.row(InlineKeyboardButton(text=f"👤 {name}", callback_data=f"admin_user_info_{uid}"))

        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_users_main"))
        await callback.message.edit_text(
            f"✅ Faol foydalanuvchilar ({len(active)} ta):",
            reply_markup=builder.as_markup()
        )
    except Exception as e:
        logging.error(f"admin_users_active error: {e}")
        await callback.message.edit_text("❌ Ro'yxatni olishda xatolik")

@dp.callback_query(F.data == "admin_users_blocked")
async def admin_users_blocked(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        blocked =[uid for uid in user_ids if user_status.get(uid) == 'blocked']
        
        if not blocked:
            await callback.message.edit_text("📭 Bloklangan foydalanuvchilar yo'q.")
            await callback.answer()
            return
        
        builder = InlineKeyboardBuilder()
        for uid in blocked[:20]:
            name = user_names.get(uid, f"ID: {uid}")
            specialty = user_specialty.get(uid, '')
            specialty_display = f" [{specialty}]" if specialty else ""
            builder.row(
                InlineKeyboardButton(text=f"⛔ {name}{specialty_display}", callback_data=f"admin_user_info_{uid}")
            )
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_users_main"))
        
        await callback.message.edit_text(
            "⛔ Bloklangan foydalanuvchilar:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_users_blocked error: {e}")
        await callback.message.edit_text("❌ Bloklangan foydalanuvchilar ro'yxatini olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_broadcast")
async def admin_broadcast_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    user_id = callback.from_user.id
    lang = user_languages.get(user_id, 'uz')
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text=TRANSLATIONS[lang]['specialty_it'], callback_data="broadcast_spec_IT"),
        InlineKeyboardButton(text=TRANSLATIONS[lang]['specialty_korean'], callback_data="broadcast_spec_Koreys tili")
    )
    builder.row(
        InlineKeyboardButton(text="🏢 Ofis xodimi", callback_data="broadcast_spec_Ofis xodimi"),
        InlineKeyboardButton(text=TRANSLATIONS[lang]['all_teachers'], callback_data="broadcast_spec_all")
    )
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    
    await state.set_state(Broadcast.selecting_specialty)
    await callback.message.edit_text(
        get_text(user_id, 'select_broadcast_specialty'),
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@dp.callback_query(Broadcast.selecting_specialty, F.data.startswith("broadcast_spec_"))
async def admin_broadcast_specialty(callback: types.CallbackQuery, state: FSMContext):
    try:
        specialty = callback.data.replace("broadcast_spec_", "")
        if specialty == "all":
            specialty = None
        
        await state.update_data(specialty=specialty)
        await state.set_state(Broadcast.waiting_for_message)
        
        await callback.message.edit_text(
            "📢 Xabar yuborish\n\nYubormoqchi bo'lgan xabaringizni kiriting (matn, rasm, hujjat):"
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_broadcast_specialty error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.message(Broadcast.waiting_for_message)
async def admin_broadcast_message(message: types.Message, state: FSMContext):
    if not check_admin(message.chat.id):
        await state.clear()
        return
    
    try:
        await state.update_data(
            message_text=message.text or message.caption,
            message_type=message.content_type,
            message_data=message
        )
        
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="✅ Ha", callback_data="broadcast_confirm"),
            InlineKeyboardButton(text="❌ Yo'q", callback_data="broadcast_cancel")
        )
        
        data = await state.get_data()
        specialty = data.get('specialty')
        
        if specialty:
            target_users =[uid for uid in user_ids if user_status.get(uid) != 'blocked' and user_specialty.get(uid) == specialty]
            specialty_text = f" ({specialty})"
        else:
            target_users =[uid for uid in user_ids if user_status.get(uid) != 'blocked']
            specialty_text = " (barcha)"
        
        total_users = len(target_users)
        
        await state.set_state(Broadcast.waiting_for_confirm)
        await message.answer(
            f"📢 Xabar yuborishni tasdiqlang{specialty_text}\n\n"
            f"Xabar: {message.text or 'Rasm/hujjat'}\n"
            f"Qabul qiluvchilar: {total_users} ta foydalanuvchi\n\n"
            f"Yuborishni boshlaymizmi?",
            reply_markup=builder.as_markup()
        )
    except Exception as e:
        logging.error(f"admin_broadcast_message error: {e}")
        await message.answer("❌ Xatolik yuz berdi")
        await state.clear()

@dp.callback_query(Broadcast.waiting_for_confirm, F.data == "broadcast_confirm")
async def admin_broadcast_confirm(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        data = await state.get_data()
        specialty = data.get('specialty')
        
        if specialty:
            target_users =[uid for uid in user_ids if user_status.get(uid) != 'blocked' and user_specialty.get(uid) == specialty]
        else:
            target_users =[uid for uid in user_ids if user_status.get(uid) != 'blocked']
        
        await callback.message.edit_text("⏳ Xabarlar yuborilmoqda...")
        
        sent_count = 0
        failed_count = 0
        
        for user_id in target_users:
            try:
                msg_data = data['message_data']
                if data['message_type'] == 'text':
                    await bot.send_message(user_id, msg_data.text)
                elif data['message_type'] == 'photo':
                    await bot.send_photo(user_id, msg_data.photo[-1].file_id, caption=msg_data.caption)
                elif data['message_type'] == 'document':
                    await bot.send_document(user_id, msg_data.document.file_id, caption=msg_data.caption)
                sent_count += 1
                await asyncio.sleep(0.05)
            except Exception as e:
                failed_count += 1
                err_str = str(e).lower()
                if 'blocked' in err_str or 'forbidden' in err_str or 'deactivated' in err_str:
                    user_status[user_id] = 'blocked'
                    logging.warning(f"Bot bloklangan (broadcast): {user_id}")
                else:
                    logging.error(f"Broadcast error for user {user_id}: {e}")
        
        broadcast_history.append({
            'text': data.get('message_text', ''),
            'date': datetime.now(UZB_TZ).strftime("%Y-%m-%d %H:%M:%S"),
            'sent_count': sent_count,
            'failed_count': failed_count,
            'specialty': specialty
        })
        
        await db.save_broadcast(data.get('message_text', ''), sent_count, failed_count, specialty)
        
        specialty_text = f" ({specialty})" if specialty else " (barcha)"
        
        await callback.message.edit_text(
            f"✅ Xabar yuborildi{specialty_text}!\n\n"
            f"✓ Yuborildi: {sent_count}\n"
            f"✗ Xatolik: {failed_count}"
        )
        
        await state.clear()
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_broadcast_confirm error: {e}")
        await callback.message.edit_text("❌ Xabar yuborishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(Broadcast.waiting_for_confirm, F.data == "broadcast_cancel")
async def admin_broadcast_cancel(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        await state.clear()
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Admin panel", callback_data="admin_back"))
        
        await callback.message.edit_text(
            "❌ Xabar yuborish bekor qilindi.",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_broadcast_cancel error: {e}")
        await callback.answer("Xatolik yuz berdi")

# =============================================
# ADMIN BACK + FILIALLAR HANDLERLARI
# =============================================
@dp.message(lambda m: m.text == '🖥 Admin Panel' or m.text == '/admin_panel')
async def open_admin_miniapp(message: types.Message):
    if not check_admin(message.chat.id):
        return
    webapp_url = f"{BASE_URL}/admin"
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="🖥 Admin Panelni ochish",
            url=webapp_url
        )
    ]])
    await message.answer("📊 Admin Dashboard:", reply_markup=kb)

@dp.callback_query(F.data == "admin_back")
async def admin_back(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    await callback.answer()
    await state.clear()
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="💰 Oylik hisoblash", callback_data="admin_salary_calc"))
    builder.row(InlineKeyboardButton(text="📚 Guruhlar", callback_data="admin_groups_menu"))
    builder.row(
        InlineKeyboardButton(text="👥 Foydalanuvchilar", callback_data="admin_users_main"),
        InlineKeyboardButton(text="📢 Xabar yuborish", callback_data="admin_broadcast")
    )
    builder.row(
        InlineKeyboardButton(text="🏢 Filiallar", callback_data="admin_locations_main"),
        InlineKeyboardButton(text="📅 Dars jadvallari", callback_data="admin_schedules_main")
    )
    builder.row(
        InlineKeyboardButton(text="📊 Oylik hisobot (Excel)", callback_data="admin_excel_menu"),
        InlineKeyboardButton(text="📊 Kunlik PDF", callback_data="admin_pdf_report")
    )
    builder.row(InlineKeyboardButton(text="🧑‍🎓 O'quvchilar davomati", callback_data="admin_student_att_branches"))
    try:
        await callback.message.edit_text(
            "👨‍💼 Admin Panel\n\nKerakli bo'limni tanlang:",
            reply_markup=builder.as_markup()
        )
    except:
        await bot.send_message(
            callback.message.chat.id,
            "👨‍💼 Admin Panel\n\nKerakli bo'limni tanlang:",
            reply_markup=builder.as_markup()
        )

@dp.callback_query(F.data == "admin_locations_main")
async def admin_locations_main(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    await callback.answer()
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📋 Filiallar ro'yxati", callback_data="admin_location_list"))
    builder.row(InlineKeyboardButton(text="➕ Yangi filial qo'shish", callback_data="admin_location_add"))
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    await callback.message.edit_text(
        "🏢 Filiallar boshqaruvi\n\nKerakli amalni tanlang:",
        reply_markup=builder.as_markup()
    )

@dp.callback_query(F.data == "admin_location_list")
async def admin_location_list(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    await callback.answer()
    builder = InlineKeyboardBuilder()
    for loc in LOCATIONS:
        maps_link = f"https://yandex.com/maps/?pt={loc['lon']},{loc['lat']}&z=17&l=map"
        builder.row(InlineKeyboardButton(
            text=f"📍 {loc['name']}",
            url=maps_link
        ))
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_locations_main"))
    await callback.message.edit_text(
        "🏢 Filiallar ro'yxati\n\nFilial nomiga bosing — Yandex Maps da ko'ring:",
        reply_markup=builder.as_markup()
    )

@dp.callback_query(F.data == "admin_location_add")
async def admin_location_add_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    await callback.answer()
    await state.set_state(AddLocation.waiting_for_name)
    await callback.message.edit_text(
        "➕ Yangi filial qo'shish\n\nFilial nomini kiriting:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="🔙 Bekor qilish", callback_data="admin_locations_main")
        ]])
    )

@dp.message(AddLocation.waiting_for_name)
async def admin_location_name_received(message: types.Message, state: FSMContext):
    await state.update_data(loc_name=message.text.strip())
    await state.set_state(AddLocation.waiting_for_coords)
    await message.answer(
        "📍 Lokatsiyani yuboring yoki koordinatalarni kiriting:\n"
        "Format: 41.257490, 69.220109"
    )

@dp.message(AddLocation.waiting_for_coords)
async def admin_location_coords_received(message: types.Message, state: FSMContext):
    data = await state.get_data()
    loc_name = data.get('loc_name', 'Yangi filial')
    try:
        if message.location:
            lat = message.location.latitude
            lon = message.location.longitude
        else:
            parts = message.text.replace(' ', '').split(',')
            lat, lon = float(parts[0]), float(parts[1])
        LOCATIONS.append({"name": loc_name, "lat": lat, "lon": lon})
        await state.clear()
        await message.answer(
            f"✅ Filial qo'shildi!\n\n🏢 {loc_name}\n📍 {lat:.6f}, {lon:.6f}\n"
            f"Jami filiallar: {len(LOCATIONS)} ta"
        )
    except Exception as e:
        await message.answer(f"❌ Noto'g'ri format: {e}\nQaytadan kiriting:")

# =============================================

@dp.callback_query(F.data == "admin_schedules_main")
async def admin_schedules_main(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="📋 Dars jadvali PDF", callback_data="admin_schedules_pdf")
        )
        builder.row(
            InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
        )
        await callback.message.edit_text(
            "📅 Dars jadvallari",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_schedules_main error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(F.data == "admin_schedules_pdf")
async def admin_schedules_pdf(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    await callback.answer()
    await callback.message.edit_text("⏳ Dars jadvallari Excel tayyorlanmoqda...")
    try:
        buf = await create_all_schedules_pdf()
        filename = f"dars_jadvali_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.xlsx"
        await bot.send_document(
            callback.message.chat.id,
            types.BufferedInputFile(buf.getvalue(), filename=filename),
            caption=f"📋 Barcha o'qituvchilarning dars jadvallari\n📅 {datetime.now(UZB_TZ).strftime('%d.%m.%Y %H:%M')}"
        )
    except Exception as e:
        logging.error(f"admin_schedules_pdf error: {e}")
        await callback.message.edit_text(f"❌ Excel xatolik: {str(e)}")

async def create_all_schedules_pdf() -> io.BytesIO:
    """Barcha o'qituvchilarning dars jadvali — Excel format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    red_fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
    spec_order = ['IT', 'Koreys tili', 'Ofis xodimi']
    for spec in spec_order:
        ws = wb.create_sheet(title=spec[:15])
        # Sarlavha
        ws.merge_cells('A1:E1')
        ws['A1'] = f"{spec.upper()} — DARS JADVALI ({datetime.now(UZB_TZ).strftime('%d.%m.%Y')})"
        ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill = hdr_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        # Header
        headers = ["O'qituvchi", "Guruh", "Filial", "Hafta kuni", "Dars vaqti"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        row = 4
        teachers = [uid for uid in user_ids if user_specialty.get(uid) == spec and user_status.get(uid) != 'blocked']
        for uid in sorted(teachers, key=lambda x: user_names.get(x, '')):
            name = user_names.get(uid, f"ID:{uid}")
            teacher_groups = [(gid, gd) for gid, gd in groups.items() if gd.get('teacher_id') == uid]
            if not teacher_groups:
                continue
            for gid, gdata in teacher_groups:
                branch = gdata.get('branch', '')
                group_name = gdata.get('group_name', '')
                days_list = gdata.get('days', [])
                day_times_g = gdata.get('day_times', {})
                time_text = gdata.get('time_text') or gdata.get('time', '')
                if days_list:
                    for day in days_list:
                        vaqt = day_times_g.get(day, time_text)
                        row_vals = [name, group_name, branch, day, vaqt]
                        for col, val in enumerate(row_vals, 1):
                            cell = ws.cell(row=row, column=col, value=val)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center')
                        row += 1
                else:
                    row_vals = [name, group_name, branch, '—', time_text]
                    for col, val in enumerate(row_vals, 1):
                        cell = ws.cell(row=row, column=col, value=val)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                    row += 1
        for i, w in enumerate([25, 20, 20, 15, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@dp.callback_query(F.data == "admin_active_groups")
async def admin_active_groups(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    try:
        async with db.pool.acquire() as conn:
            all_grps = await conn.fetch("SELECT * FROM groups ORDER BY created_at DESC")
        if not all_grps:
            builder = InlineKeyboardBuilder()
            builder.row(InlineKeyboardButton(text="Ortga", callback_data="admin_back"))
            await callback.message.edit_text("Hozircha guruhlar yo'q.", reply_markup=builder.as_markup())
            await callback.answer()
            return
        builder = InlineKeyboardBuilder()
        for grp in all_grps:
            tname = user_names.get(grp["teacher_id"], "Noma'lum")
            builder.row(InlineKeyboardButton(
                text=f"👥 {grp['group_name']} | {grp['branch']} | {tname}",
                callback_data=f"grp_view_{grp['id']}"
            ))
        builder.row(InlineKeyboardButton(text="Ortga", callback_data="admin_back"))
        await callback.message.edit_text(
            f"Faol guruhlar ({len(all_grps)} ta):",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_active_groups error: {e}")
        await callback.answer("Xatolik!", show_alert=True)

@dp.callback_query(F.data.startswith("grp_view_"))
async def grp_view_detail(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    try:
        raw_id = callback.data.replace("grp_view_", "")
        # Eski cache: "detail_21" -> "21"
        if raw_id.startswith("detail_"):
            raw_id = raw_id.replace("detail_", "")
        group_id = int(raw_id)
        async with db.pool.acquire() as conn:
            grp = await conn.fetchrow("SELECT * FROM groups WHERE id = $1", group_id)
            students = await conn.fetch("SELECT * FROM group_students WHERE group_id = $1 ORDER BY id", group_id)
        if not grp:
            await callback.answer("Guruh topilmadi!", show_alert=True)
            return
        tname = user_names.get(grp["teacher_id"], "Noma'lum")
        import json as _json
        raw = grp["days_data"]
        if isinstance(raw, str):
            raw = _json.loads(raw)
        if isinstance(raw, dict):
            # Yangi format: {kun: vaqt}
            days_times_str = "\n".join([f"  📅 {d}: {t}" for d, t in raw.items()])
        else:
            # Eski format: [kun]
            t = grp['time_text'] or ''
            days_times_str = "\n".join([f"  📅 {d}: {t}" for d in raw])
        text = (
            f"👥 Guruh: {grp['group_name']}\n"
            f"🏢 Filial: {grp['branch']}\n"
            f"📚 Fan: {grp['lesson_type']}\n"
            f"👤 O'qituvchi: {tname}\n\n"
            f"⏰ Dars vaqtlari:\n{days_times_str}\n\n"
            f"🧑‍🎓 O'quvchilar ({len(students)} ta):\n"
        )
        for idx, std in enumerate(students, 1):
            text += f"{idx}. {std['student_name']} - {std['student_phone']}\n"
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔄 O'quvchilar ro'yxatini yangilash", callback_data=f"grp_update_students_{group_id}"))
        builder.row(
            InlineKeyboardButton(text="📅 Vaqt/Kunlarni o'zgartirish", callback_data=f"grp_edit_schedule_{group_id}"),
            InlineKeyboardButton(text="👤 O'qituvchini almashtirish", callback_data=f"grp_edit_teacher_{group_id}")
        )
        builder.row(InlineKeyboardButton(text="🗑 Guruhni o'chirish", callback_data=f"grp_delete_{group_id}"))
        builder.row(InlineKeyboardButton(text="⬅️ Guruhlar ro'yxati", callback_data="admin_active_groups"))
        await callback.message.edit_text(text, reply_markup=builder.as_markup())
        await callback.answer()
    except Exception as e:
        logging.error(f"grp_view_detail error: {e}")
        await callback.answer("Xatolik!", show_alert=True)

@dp.callback_query(F.data.startswith("grp_edit_"), ~F.data.startswith("grp_edit_schedule_"), ~F.data.startswith("grp_edit_teacher_"))
async def grp_edit_students(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    try:
        try:

            group_id = int(callback.data.replace("grp_edit_", ""))

        except (ValueError, KeyError):

            await callback.answer("Xatolik!", show_alert=True)

            return
        async with db.pool.acquire() as conn:
            students = await conn.fetch("SELECT * FROM group_students WHERE group_id = $1 ORDER BY id", group_id)
        if not students:
            await callback.answer("Bu guruhda o'quvchilar yo'q!", show_alert=True)
            return
        builder = InlineKeyboardBuilder()
        for std in students:
            builder.row(InlineKeyboardButton(
                text=f"{std['student_name']}",
                callback_data=f"grp_std_edit_{std['id']}_{group_id}"
            ))
        builder.row(InlineKeyboardButton(text="Ortga", callback_data=f"grp_view_{group_id}"))
        await callback.message.edit_text("Tahrirlash uchun o'quvchini tanlang:", reply_markup=builder.as_markup())
        await callback.answer()
    except Exception as e:
        logging.error(f"grp_edit_students error: {e}")
        await callback.answer("Xatolik!", show_alert=True)

@dp.callback_query(F.data.startswith("grp_std_edit_"))
async def grp_std_edit_options(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    try:
        parts = callback.data.replace("grp_std_edit_", "").split("_")
        std_id, group_id = int(parts[0]), int(parts[1])
        async with db.pool.acquire() as conn:
            std = await conn.fetchrow("SELECT * FROM group_students WHERE id = $1", std_id)
        if not std:
            await callback.answer("O'quvchi topilmadi!", show_alert=True)
            return
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="Ismini o'zgartirish", callback_data=f"grp_std_rename_{std_id}_{group_id}"))
        builder.row(InlineKeyboardButton(text="Raqamini o'zgartirish", callback_data=f"grp_std_rephone_{std_id}_{group_id}"))
        builder.row(InlineKeyboardButton(text="O'chirish", callback_data=f"grp_std_del_{std_id}_{group_id}"))
        builder.row(InlineKeyboardButton(text="Ortga", callback_data=f"grp_edit_{group_id}"))
        await callback.message.edit_text(
            f"{std['student_name']}\n{std['student_phone']}\n\nNima qilmoqchisiz?",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"grp_std_edit_options error: {e}")
        await callback.answer("Xatolik!", show_alert=True)

@dp.callback_query(F.data.startswith("grp_std_rename_"))
async def grp_std_rename_start(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.replace("grp_std_rename_", "").split("_")
    std_id, group_id = int(parts[0]), int(parts[1])
    await state.update_data(edit_std_id=std_id, edit_group_id=group_id, edit_type="name")
    await state.set_state(EditGroupStudents.entering_new_name)
    await callback.message.edit_text("Yangi ismni kiriting:")
    await callback.answer()

@dp.callback_query(F.data.startswith("grp_std_rephone_"))
async def grp_std_rephone_start(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.replace("grp_std_rephone_", "").split("_")
    std_id, group_id = int(parts[0]), int(parts[1])
    await state.update_data(edit_std_id=std_id, edit_group_id=group_id, edit_type="phone")
    await state.set_state(EditGroupStudents.entering_new_phone)
    await callback.message.edit_text("Yangi telefon raqamini kiriting:")
    await callback.answer()

@dp.message(EditGroupStudents.entering_new_name)
async def grp_std_save_name(message: types.Message, state: FSMContext):
    data = await state.get_data()
    new_val = message.text.strip()
    if data.get("add_std_step") == "name":
        group_id = data["add_std_group_id"]
        await state.update_data(add_std_name=new_val, add_std_step="phone")
        await state.set_state(EditGroupStudents.entering_new_phone)
        await message.answer(f"{new_val} ning telefon raqamini kiriting:")
        return
    std_id, group_id = data["edit_std_id"], data["edit_group_id"]
    try:
        async with db.pool.acquire() as conn:
            await conn.execute("UPDATE group_students SET student_name=$1 WHERE id=$2", new_val, std_id)
        await message.answer(f"Ism yangilandi: {new_val}")
    except Exception as e:
        await message.answer(f"Xatolik: {e}")
    await state.clear()
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="Guruhga qaytish", callback_data=f"grp_view_{group_id}"))
    await message.answer("Davom etish:", reply_markup=builder.as_markup())

@dp.message(EditGroupStudents.entering_new_phone)
async def grp_std_save_phone(message: types.Message, state: FSMContext):
    data = await state.get_data()
    new_val = message.text.strip()
    if data.get("add_std_step") == "phone":
        group_id, new_name = data["add_std_group_id"], data["add_std_name"]
        try:
            async with db.pool.acquire() as conn:
                await conn.execute(
                    "INSERT INTO group_students (group_id, student_name, student_phone) VALUES ($1,$2,$3)",
                    group_id, new_name, new_val
                )
            await message.answer(f"{new_name} qo'shildi!")
        except Exception as e:
            await message.answer(f"Xatolik: {e}")
        await state.clear()
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="Guruhga qaytish", callback_data=f"grp_view_{group_id}"))
        await message.answer("Davom etish:", reply_markup=builder.as_markup())
        return
    std_id, group_id = data["edit_std_id"], data["edit_group_id"]
    try:
        async with db.pool.acquire() as conn:
            await conn.execute("UPDATE group_students SET student_phone=$1 WHERE id=$2", new_val, std_id)
        await message.answer(f"Raqam yangilandi: {new_val}")
    except Exception as e:
        await message.answer(f"Xatolik: {e}")
    await state.clear()
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="Guruhga qaytish", callback_data=f"grp_view_{group_id}"))
    await message.answer("Davom etish:", reply_markup=builder.as_markup())

@dp.callback_query(F.data.startswith("grp_std_del_"))
async def grp_std_delete(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.replace("grp_std_del_", "").split("_")
    std_id, group_id = int(parts[0]), int(parts[1])
    try:
        async with db.pool.acquire() as conn:
            std = await conn.fetchrow("SELECT * FROM group_students WHERE id=$1", std_id)
            await conn.execute("DELETE FROM group_students WHERE id=$1", std_id)
        await callback.answer(f"{std['student_name']} o'chirildi!", show_alert=True)
    except Exception as e:
        await callback.answer(f"Xatolik: {e}", show_alert=True)
        return
    async with db.pool.acquire() as conn:
        grp = await conn.fetchrow("SELECT * FROM groups WHERE id=$1", group_id)
        students = await conn.fetch("SELECT * FROM group_students WHERE group_id=$1 ORDER BY id", group_id)
    tname = user_names.get(grp["teacher_id"], "Noma'lum")
    text = f"Guruh: {grp['group_name']}\nO'qituvchi: {tname}\n\nO'quvchilar ({len(students)} ta):\n"
    for idx, std in enumerate(students, 1):
        text += f"{idx}. {std['student_name']} - {std['student_phone']}\n"
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="Tahrirlash", callback_data=f"grp_edit_{group_id}"))
    builder.row(
        InlineKeyboardButton(text="📥 Excel yuklab olish", callback_data=f"grp_excel_download_{group_id}"),
        InlineKeyboardButton(text="📤 Excel yuklash", callback_data=f"grp_excel_upload_{group_id}")
    )
    builder.row(InlineKeyboardButton(text="O'quvchi qo'shish", callback_data=f"grp_add_std_{group_id}"))
    builder.row(InlineKeyboardButton(text="Guruhni o'chirish", callback_data=f"grp_delete_{group_id}"))
    builder.row(InlineKeyboardButton(text="Guruhlar ro'yxati", callback_data="admin_active_groups"))
    await callback.message.edit_text(text, reply_markup=builder.as_markup())

@dp.callback_query(F.data.startswith("grp_add_std_"))
async def grp_add_student_start(callback: types.CallbackQuery, state: FSMContext):
    try:

        group_id = int(callback.data.replace("grp_add_std_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    await state.update_data(add_std_group_id=group_id, add_std_step="name")
    await state.set_state(EditGroupStudents.entering_new_name)
    await callback.message.edit_text("Yangi o'quvchining ismini kiriting:")
    await callback.answer()

@dp.callback_query(F.data.startswith("grp_delete_confirm_"))
async def grp_delete_execute(callback: types.CallbackQuery, state: FSMContext):
    try:

        group_id = int(callback.data.replace("grp_delete_confirm_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    try:
        async with db.pool.acquire() as conn:
            grp_name = await conn.fetchval("SELECT group_name FROM groups WHERE id=$1", group_id)
            await conn.execute("DELETE FROM groups WHERE id=$1", group_id)
        groups.pop(group_id, None)
        group_students.pop(group_id, None)
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="Guruhlar ro'yxati", callback_data="admin_active_groups"))
        await callback.message.edit_text(f"{grp_name} guruhi o'chirildi!", reply_markup=builder.as_markup())
    except Exception as e:
        await callback.answer(f"Xatolik: {e}", show_alert=True)
    await callback.answer()

@dp.callback_query(F.data.startswith("grp_delete_"))
async def grp_delete_confirm(callback: types.CallbackQuery, state: FSMContext):
    try:

        group_id = int(callback.data.replace("grp_delete_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    async with db.pool.acquire() as conn:
        grp = await conn.fetchrow("SELECT group_name FROM groups WHERE id=$1", group_id)
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="Ha, o'chirish", callback_data=f"grp_delete_confirm_{group_id}"),
        InlineKeyboardButton(text="Yo'q", callback_data=f"grp_view_{group_id}")
    )
    await callback.message.edit_text(
        f"{grp['group_name']} guruhini o'chirishni tasdiqlaysizmi? Barcha o'quvchilar ham o'chib ketadi!",
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("grp_excel_download_"))
async def grp_excel_download(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    try:
        try:

            group_id = int(callback.data.replace("grp_excel_download_", ""))

        except (ValueError, KeyError):

            await callback.answer("Xatolik!", show_alert=True)

            return
        async with db.pool.acquire() as conn:
            grp = await conn.fetchrow("SELECT * FROM groups WHERE id=$1", group_id)
            students = await conn.fetch("SELECT * FROM group_students WHERE group_id=$1 ORDER BY id", group_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "O'quvchilar"
        ws["A1"] = f"Guruh: {grp['group_name']}"
        ws["A1"].font = Font(bold=True, size=13)
        ws.merge_cells("A1:C1")
        ws["A2"] = f"Filial: {grp['branch']}  |  Fan: {grp['lesson_type']}  |  Vaqt: {grp['time_text']}"
        ws.merge_cells("A2:C2")
        ws["A3"] = "group_id"
        ws["B3"] = group_id
        ws.row_dimensions[3].hidden = True
        for col, h in enumerate(["No", "Ism Familiya", "Telefon raqami"], 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for idx, std in enumerate(students, 1):
            ws.cell(row=4+idx, column=1, value=idx)
            ws.cell(row=4+idx, column=2, value=std["student_name"])
            ws.cell(row=4+idx, column=3, value=std["student_phone"])
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        note_row = 4 + len(students) + 2
        ws.cell(row=note_row, column=1,
            value="ESLATMA: Yangi o'quvchi - yangi qator qo'shing. O'chirish - qatorni turing. Tahrirlang. Keyin faylni qaytarib yuboring.")
        ws.merge_cells(f"A{note_row}:C{note_row}")
        ws.cell(row=note_row, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=note_row, column=1).font = Font(italic=True, size=9)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"guruh_{grp['group_name'].replace(' ', '_')}.xlsx"
        await callback.message.answer_document(
            types.BufferedInputFile(buf.read(), filename=filename),
            caption=(
                f"Guruh: {grp['group_name']} - o'quvchilar ro'yxati\n\n"
                f"Faylni tahrirlang va 'Excel yuklash' tugmasini bosib qaytarib yuboring."
            )
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"grp_excel_download error: {e}")
        await callback.answer(f"Xatolik: {e}", show_alert=True)

# --- GURUH DARS VAQTI/KUNLARI O'ZGARTIRISH ---
@dp.callback_query(F.data.startswith("grp_edit_schedule_"))
async def grp_edit_schedule_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    try:

        group_id = int(callback.data.replace("grp_edit_schedule_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    grp = groups.get(group_id)
    if not grp:
        await callback.answer("Guruh topilmadi!", show_alert=True)
        return
    current_days = grp.get('days', [])
    await state.update_data(edit_grp_id=group_id, edit_days=list(current_days))
    await state.set_state(EditGroupSchedule.selecting_days)

    builder = InlineKeyboardBuilder()
    for day in WEEKDAYS_UZ:
        check = "✅" if day in current_days else "⬜"
        builder.row(InlineKeyboardButton(text=f"{check} {day}", callback_data=f"egrp_day_{day}"))
    builder.row(InlineKeyboardButton(text="✔️ Kunlarni tasdiqlash", callback_data="egrp_days_done"))
    builder.row(InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"grp_view_{group_id}"))
    await callback.message.edit_text(
        f"📅 *{grp['group_name']}* guruhining dars kunlarini tanlang:\n\n"
        f"Joriy kunlar: {', '.join(current_days)}",
        reply_markup=builder.as_markup(),
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(EditGroupSchedule.selecting_days, F.data.startswith("egrp_day_"))
async def egrp_day_toggle(callback: types.CallbackQuery, state: FSMContext):
    day = callback.data.replace("egrp_day_", "")
    data = await state.get_data()
    days = data.get('edit_days', [])
    if day in days:
        days.remove(day)
    else:
        days.append(day)
    await state.update_data(edit_days=days)
    group_id = data['edit_grp_id']
    grp = groups.get(group_id, {})

    builder = InlineKeyboardBuilder()
    for d in WEEKDAYS_UZ:
        check = "✅" if d in days else "⬜"
        builder.row(InlineKeyboardButton(text=f"{check} {d}", callback_data=f"egrp_day_{d}"))
    builder.row(InlineKeyboardButton(text="✔️ Kunlarni tasdiqlash", callback_data="egrp_days_done"))
    builder.row(InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"grp_view_{group_id}"))
    await callback.message.edit_reply_markup(reply_markup=builder.as_markup())
    await callback.answer()

@dp.callback_query(EditGroupSchedule.selecting_days, F.data == "egrp_days_done")
async def egrp_days_done(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    days = data.get('edit_days', [])
    if not days:
        await callback.answer("Kamida 1 kun tanlang!", show_alert=True)
        return
    # Har kun uchun alohida vaqt so'rash
    await state.update_data(edit_day_times={}, edit_days_remaining=list(days))
    await state.set_state(EditGroupSchedule.entering_day_times)
    group_id = data['edit_grp_id']
    grp = groups.get(group_id, {})
    first_day = days[0]
    current = grp.get('day_times', {}).get(first_day, grp.get('time', '—'))
    await callback.message.edit_text(
        f"⏰ *{first_day}* uchun dars vaqtini kiriting (HH:MM)\n\n"
        f"Joriy vaqt: {current}\n"
        f"({len(days)} kundan 1-chi)",
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.message(EditGroupSchedule.entering_day_times)
async def egrp_day_time_entered(message: types.Message, state: FSMContext):
    if not check_admin(message.chat.id):
        return
    import re
    time_text = message.text.strip()
    if not re.match(r'^\d{2}:\d{2}$', time_text):
        await message.answer("❌ Noto'g'ri format! HH:MM shaklida kiriting (masalan: 14:30)")
        return
    data = await state.get_data()
    group_id = data['edit_grp_id']
    days = data['edit_days']
    day_times = data.get('edit_day_times', {})
    remaining = data.get('edit_days_remaining', [])

    # Joriy kunni saqlash
    current_day = remaining[0]
    day_times[current_day] = time_text
    remaining = remaining[1:]
    await state.update_data(edit_day_times=day_times, edit_days_remaining=remaining)

    if remaining:
        # Keyingi kun uchun so'rash
        next_day = remaining[0]
        grp = groups.get(group_id, {})
        current = grp.get('day_times', {}).get(next_day, grp.get('time', '—'))
        done = len(days) - len(remaining)
        await message.answer(
            f"⏰ *{next_day}* uchun dars vaqtini kiriting (HH:MM)\n\n"
            f"Joriy vaqt: {current}\n"
            f"({len(days)} kundan {done+1}-chi)",
            parse_mode="Markdown"
        )
    else:
        # Barchasi kiritildi — saqlash
        first_time = list(day_times.values())[0]
        time_display = ", ".join([f"{d}: {t}" for d, t in day_times.items()])
        try:
            async with db.pool.acquire() as conn:
                await conn.execute(
                    "UPDATE groups SET days_data=$1, time_text=$2 WHERE id=$3",
                    json.dumps(day_times), first_time, group_id
                )
            if group_id in groups:
                groups[group_id]['days'] = days
                groups[group_id]['day_times'] = day_times
                groups[group_id]['time'] = first_time
                groups[group_id]['time_text'] = first_time
            grp = groups.get(group_id, {})
            teacher_id = grp.get('teacher_id')
            if teacher_id:
                try:
                    await bot.send_message(
                        teacher_id,
                        f"📅 *{grp['group_name']}* guruhingizning dars jadvali yangilandi!\n\n"
                        f"📆 Yangi kunlar va vaqtlar:\n{time_display}",
                        parse_mode="Markdown"
                    )
                except:
                    pass
            await message.answer(
                f"✅ Dars jadvali yangilandi!\n\n"
                f"📆 Kunlar va vaqtlar:\n{time_display}"
            )
        except Exception as e:
            logging.error(f"egrp_time_entered error: {e}")
            await message.answer(f"❌ Xatolik: {e}")
        await state.clear()

@dp.message(EditGroupSchedule.entering_time)
async def egrp_time_entered(message: types.Message, state: FSMContext):
    """Eski compat — to'g'ridan foydalanilmaydi"""
    await state.clear()

# --- GURUH O'QITUVCHISINI ALMASHTIRISH ---
@dp.callback_query(F.data.startswith("grp_edit_teacher_"))
async def grp_edit_teacher_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    try:

        group_id = int(callback.data.replace("grp_edit_teacher_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    grp = groups.get(group_id)
    if not grp:
        await callback.answer("Guruh topilmadi!", show_alert=True)
        return

    branch = grp.get('branch', '')
    lesson_type = grp.get('lesson_type', '')

    # Filialdagi o'qituvchilarni topamiz
    teachers = [
        (uid, user_names.get(uid, f"ID:{uid}"))
        for uid, spec in user_specialty.items()
        if spec and (
            (lesson_type == 'IT' and 'IT' in spec) or
            (lesson_type == 'Koreys tili' and 'Koreys' in spec) or
            True
        )
    ]

    if not teachers:
        await callback.answer("O'qituvchilar topilmadi!", show_alert=True)
        return

    await state.update_data(edit_grp_id=group_id)
    await state.set_state(EditGroupTeacher.selecting_teacher)

    builder = InlineKeyboardBuilder()
    for uid, name in teachers[:20]:
        builder.row(InlineKeyboardButton(text=f"👤 {name}", callback_data=f"egrp_teacher_{uid}"))
    builder.row(InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"grp_view_{group_id}"))
    current_teacher = user_names.get(grp.get('teacher_id'), "—")
    await callback.message.edit_text(
        f"👤 *{grp['group_name']}* guruhining o'qituvchisini tanlang:\n\n"
        f"Joriy o'qituvchi: {current_teacher}",
        reply_markup=builder.as_markup(),
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(EditGroupTeacher.selecting_teacher, F.data.startswith("egrp_teacher_"))
async def egrp_teacher_selected(callback: types.CallbackQuery, state: FSMContext):
    try:

        new_teacher_id = int(callback.data.replace("egrp_teacher_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    data = await state.get_data()
    group_id = data['edit_grp_id']
    grp = groups.get(group_id, {})
    old_teacher_id = grp.get('teacher_id')
    new_teacher_name = user_names.get(new_teacher_id, f"ID:{new_teacher_id}")
    try:
        async with db.pool.acquire() as conn:
            await conn.execute("UPDATE groups SET teacher_id=$1 WHERE id=$2", new_teacher_id, group_id)
        if group_id in groups:
            groups[group_id]['teacher_id'] = new_teacher_id
        # Eski o'qituvchiga xabar
        if old_teacher_id and old_teacher_id != new_teacher_id:
            try:
                await bot.send_message(
                    old_teacher_id,
                    f"ℹ️ *{grp['group_name']}* guruhi sizdan boshqa o'qituvchiga o'tkazildi."
                    , parse_mode="Markdown"
                )
            except:
                pass
        # Yangi o'qituvchiga xabar
        try:
            await bot.send_message(
                new_teacher_id,
                f"🎉 Siz *{grp['group_name']}* guruhining yangi o'qituvchisi sifatida tayinlandingiz!\n\n"
                f"🏢 Filial: {grp.get('branch', '—')}\n"
                f"📆 Kunlar: {', '.join(grp.get('days', []))}\n"
                f"⏰ Vaqt: {grp.get('time_text', '—')}",
                parse_mode="Markdown"
            )
        except:
            pass
        await callback.message.edit_text(
            f"✅ O'qituvchi muvaffaqiyatli almashtirildi!\n\n"
            f"Guruh: {grp['group_name']}\n"
            f"Yangi o'qituvchi: {new_teacher_name}"
        )
    except Exception as e:
        logging.error(f"egrp_teacher_selected error: {e}")
        await callback.answer(f"❌ Xatolik: {e}", show_alert=True)
    await state.clear()

@dp.callback_query(F.data.startswith("grp_update_students_"))
async def grp_update_students_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    try:

        group_id = int(callback.data.replace("grp_update_students_", ""))

    except (ValueError, KeyError):

        await callback.answer("Xatolik!", show_alert=True)

        return
    # Mavjud o'quvchilar bilan shablon Excel yuboramiz
    try:
        async with db.pool.acquire() as conn:
            grp = await conn.fetchrow("SELECT * FROM groups WHERE id=$1", group_id)
            students = await conn.fetch("SELECT * FROM group_students WHERE group_id=$1 ORDER BY id", group_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "O'quvchilar"
        ws["A1"] = f"Guruh: {grp['group_name']} | {grp['branch']}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="4472C4")
        ws.merge_cells("A1:B1")
        for col, h in enumerate(["Ism Familiya", "Telefon raqami"], 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E86AB")
            cell.alignment = Alignment(horizontal="center")
        for idx, std in enumerate(students, 1):
            ws.cell(row=2+idx, column=1, value=std["student_name"])
            ws.cell(row=2+idx, column=2, value=std["student_phone"])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        await state.update_data(excel_upload_group_id=group_id)
        await state.set_state(ExcelUploadGroup.waiting_file)
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="Bekor qilish", callback_data=f"grp_view_{group_id}"))
        await callback.message.answer_document(
            types.BufferedInputFile(buf.read(), filename=f"oquvchilar_{grp['group_name']}.xlsx"),
            caption=(
                f"📋 {grp['group_name']} guruhi o'quvchilari ro'yxati\n\n"
                "Faylni tahrirlang:\n"
                "• Yangi o'quvchi — yangi qator qo'shing\n"
                "• O'chirish — qatorni o'chiring\n"
                "• Tahrirlash — ismi/raqamni o'zgartiring\n\n"
                "Tayyor bo'lgach, faylni yuboring — ro'yxat yangilanadi va o'qituvchiga xabar boriladi."
            ),
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"grp_update_students_start error: {e}")
        await callback.answer(f"Xatolik: {e}", show_alert=True)

@dp.message(ExcelUploadGroup.waiting_file, F.document)
async def grp_excel_process(message: types.Message, state: FSMContext):
    if not check_admin(message.chat.id):
        return
    data = await state.get_data()
    group_id = data.get("excel_upload_group_id")
    if not group_id:
        await message.answer("Guruh ID topilmadi. Qaytadan boshlang.")
        await state.clear()
        return
    doc = message.document
    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await message.answer("Faqat .xlsx yoki .xls fayl yuboring!")
        return
    try:
        file = await message.bot.get_file(doc.file_id)
        file_buf = io.BytesIO()
        await message.bot.download_file(file.file_path, file_buf)
        file_buf.seek(0)
        wb = openpyxl.load_workbook(file_buf)
        ws = wb.active
        new_students = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            vals = (list(row) + [None, None])[:2]
            name, phone = vals
            if name and str(name).strip():
                new_students.append({
                    "name": str(name).strip(),
                    "phone": str(phone).strip() if phone else "-"
                })
        if not new_students:
            await message.answer("Faylda o'quvchilar topilmadi! 3-qatordan boshlab to'ldiring.")
            return
        async with db.pool.acquire() as conn:
            grp = await conn.fetchrow("SELECT * FROM groups WHERE id=$1", group_id)
            await conn.execute("DELETE FROM group_students WHERE group_id=$1", group_id)
            for std in new_students:
                await conn.execute(
                    "INSERT INTO group_students (group_id, student_name, student_phone) VALUES ($1,$2,$3)",
                    group_id, std["name"], std["phone"]
                )
        group_students[group_id] = new_students
        # O'qituvchiga xabar
        teacher_id = grp["teacher_id"]
        tname = user_names.get(teacher_id, "O'qituvchi")
        try:
            await bot.send_message(
                teacher_id,
                f"📢 *{grp['branch']} filialidagi {grp['group_name']} guruhining o'quvchilar ro'yxati yangilandi!*\n\n"
                f"Jami: {len(new_students)} ta o'quvchi",
                parse_mode="Markdown"
            )
        except Exception as e:
            logging.error(f"Teacher notify error: {e}")
        await state.clear()
        result = f"✅ {grp['group_name']} guruhi yangilandi!\n\nJami {len(new_students)} ta o'quvchi:\n"
        for idx, std in enumerate(new_students, 1):
            result += f"{idx}. {std['name']} - {std['phone']}\n"
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="Guruhga qaytish", callback_data=f"grp_view_{group_id}"))
        await message.answer(result, reply_markup=builder.as_markup())
    except Exception as e:
        logging.error(f"grp_excel_process error: {e}")
        await message.answer(f"Faylni o'qishda xatolik: {e}")
        await state.clear()

@dp.message(ExcelUploadGroup.waiting_file)
async def grp_excel_wrong_file(message: types.Message, state: FSMContext):
    await message.answer("Iltimos, .xlsx formatidagi Excel fayl yuboring!")


# ============================================================
# EXCEL ORQALI YANGI GURUH YARATISH
# ============================================================

@dp.callback_query(F.data == "admin_excel_create_group")
async def excel_create_group_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guruh Ma'lumotlari"
    ws["A1"] = "GURUH MA'LUMOTLARI - SHABLON"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2E86AB")
    ws.merge_cells("A1:D1")
    fields = [
        ("A3", "Guruh nomi:", "B3", "Masalan: IT-1 yoki Koreys-A"),
        ("A4", "Filial:", "B4", "Masalan: 78-Maktab"),
        ("A5", "Fan:", "B5", "IT yoki Koreys tili"),
        ("A6", "O'qituvchi Telegram ID:", "B6", "Masalan: 123456789"),
        ("A7", "Dars vaqti:", "B7", "Masalan: 14:00"),
        ("A8", "Dars kunlari:", "B8", "Masalan: Dushanba,Chorshanba,Juma"),
    ]
    for label_cell, label_val, val_cell, hint in fields:
        ws[label_cell] = label_val
        ws[label_cell].font = Font(bold=True)
        ws[val_cell] = hint
        ws[val_cell].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["A10"] = "O'QUVCHILAR RO'YXATI"
    ws["A10"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A10"].fill = PatternFill("solid", fgColor="4472C4")
    ws.merge_cells("A10:C10")
    for col, h in enumerate(["No", "Ism Familiya", "Telefon raqami"], 1):
        cell = ws.cell(row=11, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
    for i, (name, phone) in enumerate([("Ali Karimov", "+998901234567"), ("Barno Qosimova", "+998911234567")], 1):
        ws.cell(row=11+i, column=1, value=i)
        ws.cell(row=11+i, column=2, value=name)
        ws.cell(row=11+i, column=3, value=phone)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await callback.message.answer_document(
        types.BufferedInputFile(buf.read(), filename="yangi_guruh_shablon.xlsx"),
        caption=(
            "📋 Yangi guruh yaratish uchun shablon\n\n"
            "1. Faylni yuklab oling\n"
            "2. Ma'lumotlarni to'ldiring:\n"
            "   - Guruh nomi, filial, fan, vaqt, kunlar\n"
            "   - O'qituvchi Telegram ID\n"
            "   - 12-qatordan o'quvchilarni kiriting\n"
            "3. To'ldirilgan faylni yuboring"
        )
    )
    await state.set_state(ExcelCreateGroup.waiting_file)
    await callback.answer()

@dp.message(ExcelCreateGroup.waiting_file, F.document)
async def excel_create_group_process(message: types.Message, state: FSMContext):
    if not check_admin(message.chat.id):
        return
    doc = message.document
    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await message.answer("Faqat .xlsx yoki .xls fayl yuboring!")
        return
    try:
        await message.answer("Fayl o'qilmoqda...")
        file = await message.bot.get_file(doc.file_id)
        file_buf = io.BytesIO()
        await message.bot.download_file(file.file_path, file_buf)
        file_buf.seek(0)
        wb = openpyxl.load_workbook(file_buf)
        ws = wb.active
        group_name = str(ws["B3"].value or "").strip()
        branch = str(ws["B4"].value or "").strip()
        lesson_type = str(ws["B5"].value or "").strip()
        teacher_id_raw = ws["B6"].value
        time_text = str(ws["B7"].value or "").strip()
        days_raw = str(ws["B8"].value or "").strip()
        errors = []
        if not group_name:
            errors.append("Guruh nomi bo'sh!")
        if not branch:
            errors.append("Filial bo'sh!")
        if lesson_type not in ("IT", "Koreys tili"):
            errors.append("Fan: faqat 'IT' yoki 'Koreys tili' bo'lishi kerak!")
        try:
            teacher_id = int(teacher_id_raw)
            if teacher_id not in user_names:
                errors.append(f"O'qituvchi ID {teacher_id} topilmadi!")
        except (TypeError, ValueError):
            errors.append("O'qituvchi Telegram ID noto'g'ri!")
            teacher_id = None
        if not time_text:
            errors.append("Dars vaqti bo'sh!")
        days_list = [d.strip() for d in days_raw.split(",") if d.strip()]
        invalid_days = [d for d in days_list if d not in set(WEEKDAYS_UZ)]
        if not days_list:
            errors.append("Dars kunlari bo'sh!")
        if invalid_days:
            errors.append(f"Noto'g'ri kun nomlari: {', '.join(invalid_days)}")
        if errors:
            await message.answer("Xatoliklar:\n\n" + "\n".join(f"  {e}" for e in errors))
            return
        students = []
        for row in ws.iter_rows(min_row=12, values_only=True):
            vals = (list(row) + [None, None, None])[:3]
            num, name, phone = vals
            if name and str(name).strip():
                students.append({"name": str(name).strip(), "phone": str(phone).strip() if phone else "-"})
        if not students:
            await message.answer("O'quvchilar topilmadi! 12-qatordan boshlab kiriting.")
            return
        async with db.pool.acquire() as conn:
            group_id = await conn.fetchval("""
                INSERT INTO groups (group_name, branch, lesson_type, teacher_id, days_data, time_text)
                VALUES ($1, $2, $3, $4, $5::jsonb, $6) RETURNING id
            """, group_name, branch, lesson_type, teacher_id, json.dumps(days_list), time_text)
            for std in students:
                await conn.execute(
                    "INSERT INTO group_students (group_id, student_name, student_phone) VALUES ($1,$2,$3)",
                    group_id, std["name"], std["phone"]
                )
        groups[group_id] = {
            "group_name": group_name, "branch": branch, "lesson_type": lesson_type,
            "teacher_id": teacher_id, "days": days_list, "time": time_text, "time_text": time_text,
            "created_at": datetime.now(UZB_TZ)
        }
        group_students[group_id] = students
        days_str = ", ".join(days_list)
        try:
            await bot.send_message(
                teacher_id,
                f"🆕 *Yangi guruh biriktirildi!*\n\n"
                f"👥 Guruh: {group_name}\n"
                f"🏢 Filial: {branch}\n"
                f"📚 Fan: {lesson_type}\n"
                f"⏰ Vaqt: {time_text}\n"
                f"📅 Kunlar: {days_str}\n"
                f"🧑‍🎓 O'quvchilar soni: {len(students)} ta\n\n"
                f"📍 Botda davomat qilganingizda ushbu o'quvchilar ro'yxati chiqadi.",
                parse_mode="Markdown"
            )
        except Exception as e:
            logging.error(f"Teacher notify error: {e}")
        await state.clear()
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="Guruhni ko'rish", callback_data=f"grp_view_{group_id}"))
        builder.row(InlineKeyboardButton(text="Admin Panel", callback_data="admin_back"))
        await message.answer(
            f"Guruh yaratildi!\n\n{group_name}\n{branch} | {lesson_type}\n{time_text} | {days_str}\nOquvchilar: {len(students)} ta",
            reply_markup=builder.as_markup()
        )
    except Exception as e:
        logging.error(f"excel_create_group_process error: {e}")
        await message.answer(f"Faylni oqishda xatolik: {e}")
        await state.clear()

@dp.message(ExcelCreateGroup.waiting_file)
async def excel_create_group_wrong_file(message: types.Message, state: FSMContext):
    await message.answer("Iltimos, .xlsx formatidagi Excel fayl yuboring!")


async def auto_daily_report_task():
    while True:
        now = datetime.now(UZB_TZ)
        if now.hour == 10 and now.minute == 10:
            yesterday = now.date() - timedelta(days=1)
            logging.info(f"Avtomatik hisobot yuborilmoqda: {yesterday}")
            
            try:
                pdf_buf = await get_combined_report_pdf(yesterday)
                await bot.send_document(
                    chat_id=ADMIN_GROUP_ID,
                    document=types.BufferedInputFile(pdf_buf.read(), filename=f"hisobot_{yesterday}.xlsx"),
                    caption=f"📅 Kechagi kun ({yesterday}) uchun avtomatik davomat hisoboti."
                )
            except Exception as e:
                logging.error(f"Auto report xatosi: {e}")
            
            await asyncio.sleep(61)
        
        await asyncio.sleep(30)

async def check_schedule_reminders():
    """Guruhlar asosida dars eslatmalari va notification yuborish"""
    sent_reminders = set()  # (group_id, date, type) - takrorlanishni oldini olish

    while True:
        try:
            now_uzb = datetime.now(UZB_TZ)
            current_time = now_uzb.strftime("%H:%M")
            today_date = now_uzb.strftime("%Y-%m-%d")
            current_day_name = WEEKDAYS_UZ[now_uzb.weekday()]

            for group_id, gdata in list(groups.items()):
                teacher_id = gdata.get('teacher_id')
                if not teacher_id:
                    continue
                if user_status.get(teacher_id) == 'blocked':
                    continue

                branch = gdata.get('branch', '')
                days_list = gdata.get('days', [])
                group_name = gdata.get('group_name', '')
                day_times = gdata.get('day_times', {})

                if current_day_name not in days_list:
                    continue

                # Bugungi kun uchun vaqtni olamiz
                time_text = day_times.get(current_day_name) or gdata.get('time', '')
                if not time_text:
                    continue

                try:
                    lesson_dt = datetime.strptime(time_text, "%H:%M")
                except ValueError:
                    continue

                # Dars boshlanishi vaqti
                lesson_time_str = lesson_dt.strftime("%H:%M")
                # 5 daqiqa o'tganda tekshiruv
                check_time = (lesson_dt + timedelta(minutes=5)).strftime("%H:%M")

                # Dars boshlanishida notification
                start_key = (group_id, today_date, "start")
                if current_time == lesson_time_str and start_key not in sent_reminders:
                    attended = any(
                        k[0] == teacher_id and k[1] == branch and k[2] == today_date
                        for k in daily_attendance_log
                    )
                    sent_reminders.add(start_key)
                    try:
                        if attended:
                            # Avval kelganini tasdiqlagan — o'quvchilar davomati (miniapp + bot)
                            students_list = group_students.get(group_id, [])
                            if students_list:
                                # Bugun davomat qilinganmi?
                                already_std_done = False
                                try:
                                    async with db.pool.acquire() as conn:
                                        cnt = await conn.fetchval(
                                            "SELECT COUNT(*) FROM student_attendance WHERE group_id=$1 AND lesson_date=$2",
                                            group_id, now_uzb.date()
                                        )
                                        already_std_done = (cnt or 0) > 0
                                except: pass
                                
                                if not already_std_done:
                                    webapp_url = f"{BASE_URL}/miniapp?user_id={teacher_id}"
                                    mini_kb = InlineKeyboardMarkup(inline_keyboard=[[
                                        InlineKeyboardButton(
                                            text="📱 O'quvchilar davomatini belgilash",
                                            web_app=types.WebAppInfo(url=webapp_url)
                                        )
                                    ]])
                                    await bot.send_message(
                                        teacher_id,
                                        f"🔔 *Darsingiz boshlandi!*\n\n"
                                        f"👥 {group_name} | 🏢 {branch}\n"
                                        f"⏰ Vaqt: {time_text}\n\n"
                                        f"🧑‍🎓 O'quvchilar davomatini belgilang:",
                                        reply_markup=mini_kb,
                                        parse_mode="Markdown"
                                    )
                                    logging.info(f"Lesson start (attended, miniapp sent): teacher={teacher_id}, group={group_name}")
                        else:
                            # Kelmagan — kelganini tasdiqlashga undash
                            await bot.send_message(
                                teacher_id,
                                f"⚠️ *Darsingiz boshlandi!*\n\n"
                                f"👥 {group_name} | 🏢 {branch}\n"
                                f"⏰ Vaqt: {time_text}\n\n"
                                f"📍 Kelganingizni tasdiqlamadingiz! Iltimos, darhol *Kelganimni tasdiqlash* tugmasini bosib, o'quvchilar davomatini qiling!",
                                parse_mode="Markdown"
                            )
                            logging.info(f"Lesson start (not attended): teacher={teacher_id}, group={group_name}")
                    except Exception as e:
                        logging.error(f"Lesson start notify error: {e}")

                # 5 daqiqa o'tdi, hali kelmagan bo'lsa qayta eslatma
                check_key = (group_id, today_date, "check")
                if current_time == check_time and check_key not in sent_reminders:
                    attended = any(
                        k[0] == teacher_id and k[1] == branch and k[2] == today_date
                        for k in daily_attendance_log
                    )
                    if not attended:
                        try:
                            await bot.send_message(
                                teacher_id,
                                f"🚨 *DIQQAT: Hali davomat qilinmadi!*\n\n"
                                f"👥 {group_name} | 🏢 {branch}\n"
                                f"⏰ Dars {time_text} da boshlangan.\n\n"
                                f"Darhol davomatni tasdiqlang yoki sababini adminga bildiring!",
                                parse_mode="Markdown"
                            )
                            sent_reminders.add(check_key)
                        except Exception as e:
                            logging.error(f"Check reminder error: {e}")

            # Har kuni yarim kechada sent_reminders tozalanadi
            if now_uzb.hour == 0 and now_uzb.minute == 0:
                old_keys = {k for k in sent_reminders if k[1] != today_date}
                for k in old_keys:
                    sent_reminders.discard(k)

        except Exception as e:
            logging.error(f"check_schedule_reminders loop error: {e}")

        await asyncio.sleep(30)



async def update_all_keyboards():
    """Keyboard faqat logga yoziladi, xabar yuborilmaydi"""
    await asyncio.sleep(3)
    logging.info(f"✅ Bot ishga tushdi: {len(user_ids)} foydalanuvchi")

async def handle_static(request):
    """Static fayllarni serve qilish"""
    import os
    filename = request.match_info['filename']
    allowed_ext = {'.svg', '.png', '.jpg', '.jpeg', '.ico', '.css', '.js', '.woff', '.woff2', '.ttf'}
    ext = os.path.splitext(filename)[1].lower()
    if ext not in allowed_ext:
        return web.Response(status=403, text='Forbidden')
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', filename)
    if not os.path.exists(filepath):
        return web.Response(status=404, text='Not found')
    content_types = {
        '.svg': 'image/svg+xml', '.png': 'image/png',
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.ico': 'image/x-icon', '.css': 'text/css',
        '.js': 'application/javascript', '.ttf': 'font/ttf',
        '.woff': 'font/woff', '.woff2': 'font/woff2',
    }
    ct = content_types.get(ext, 'application/octet-stream')
    with open(filepath, 'rb') as f:
        data = f.read()
    return web.Response(body=data, content_type=ct)


async def main():
    await db.create_pool()
    await db.init_tables()
    await db.load_branches()
    await db.load_configurations()
    await db.load_to_ram()

    # Barcha foydalanuvchilar tilini 'uz' ga o'zgartirish (bir martalik migration)
    try:
        async with db.pool.acquire() as conn:
            updated = await conn.execute(
                "UPDATE users SET language = 'uz' WHERE language != 'uz'"
            )
            logging.info(f"✅ Til migration: {updated}")
        # RAM dagi user_languages ham yangilash
        for uid in list(user_languages.keys()):
            user_languages[uid] = 'uz'
    except Exception as e:
        logging.error(f"Til migration xato: {e}")

    # Background tasklar
    asyncio.create_task(check_schedule_reminders())
    asyncio.create_task(_cache_all_photos())
    asyncio.create_task(update_all_keyboards())

    # Webhook URL
    WEBHOOK_HOST = BASE_URL
    WEBHOOK_PATH = f"/webhook/{TOKEN}"
    WEBHOOK_URL  = f"{WEBHOOK_HOST}{WEBHOOK_PATH}"

    # Webhookni o'rnatamiz
    await bot.set_webhook(
        url=WEBHOOK_URL,
        drop_pending_updates=True,
        allowed_updates=["message", "callback_query", "inline_query"]
    )
    logging.info(f"✅ Webhook o'rnatildi: {WEBHOOK_URL}")

    # aiohttp server
    from aiogram.webhook.aiohttp_server import SimpleRequestHandler, setup_application

    app = web.Application()

    # Telegram update handler
    SimpleRequestHandler(dispatcher=dp, bot=bot).register(app, path=WEBHOOK_PATH)
    setup_application(app, dp, bot=bot)

    # Qo'shimcha routelar (health check, status)
    app.router.add_get('/static/{filename}', handle_static)
    app.router.add_get('/', handle)
    
    async def handle_register(request):
        import os
        html_path = os.path.join(os.path.dirname(__file__), 'royxatdan_otish.html')
        try:
            with open(html_path, 'r', encoding='utf-8') as f:
                html = f.read()
            return web.Response(text=html, content_type='text/html', charset='utf-8')
        except FileNotFoundError:
            return web.Response(text='File not found', status=404)
    
    app.router.add_get('/register', handle_register)
    import os as _os
    if _os.path.isdir('static'):
        app.router.add_static('/static', path='static', name='static')
    if _os.path.isdir('docs'):
        app.router.add_static('/docs', path='docs', name='docs')
    app.router.add_get('/health', health_check)
    app.router.add_get('/admin', admin_panel_page)
    app.router.add_post('/admin/login', admin_login)
    app.router.add_post('/admin/logout', admin_logout)
    app.router.add_get('/admin/api/data', admin_api_data)
    app.router.add_get('/admin/api/attendance', admin_api_attendance)
    app.router.add_post('/admin/api/user/status', admin_api_user_status)
    app.router.add_post('/admin/api/broadcast', admin_api_broadcast)
    app.router.add_post('/admin/api/branch/add', admin_api_branch_add)
    app.router.add_get('/admin/api/report/{type}', admin_api_report)
    app.router.add_get('/admin/api/stats', admin_api_stats)
    app.router.add_get('/admin/api/teachers', admin_api_teachers_list)
    app.router.add_get('/admin/api/student_att', admin_api_student_att)
    app.router.add_get('/admin/api/payments/summary', admin_api_payments_summary)
    app.router.add_get('/admin/api/student_payments', admin_api_student_payments)
    app.router.add_post('/admin/api/payment/save', miniapp_save_payment)
    app.router.add_post('/admin/api/salary', admin_api_salary_calc)
    app.router.add_post('/admin/api/office/salary', admin_api_office_salary_calc)
    app.router.add_post('/admin/api/office/salary/excel', admin_api_office_salary_excel)
    app.router.add_get('/admin/api/office/employees', admin_api_office_employees_list)
    app.router.add_get('/admin/api/salary/structure', admin_api_salary_structure)
    app.router.add_post('/admin/api/group/delete', admin_api_group_delete)
    app.router.add_post('/admin/api/user/delete', admin_api_user_delete)
    app.router.add_post('/admin/api/user/restore', admin_api_user_restore)
    app.router.add_post('/admin/api/user/permanent-delete', admin_api_user_permanent_delete)
    app.router.add_get('/admin/api/user/stats', admin_api_user_stats)
    app.router.add_get('/admin/api/group/detail', admin_api_group_detail)
    app.router.add_post('/admin/api/group/edit/schedule', admin_api_group_edit_schedule)
    app.router.add_post('/admin/api/group/edit/teacher', admin_api_group_edit_teacher)
    app.router.add_post('/admin/api/group/edit/branch', admin_api_group_edit_branch)
    app.router.add_post('/admin/api/group/edit/name', admin_api_group_edit_name)
    app.router.add_get('/api/site/config', admin_api_site_config_get)
    app.router.add_post('/api/apply', api_submit_application)
    app.router.add_post('/api/bootcamp/apply', api_bootcamp_apply)
    app.router.add_get('/admin/api/applications', admin_api_applications_get)
    app.router.add_post('/admin/api/application/status', admin_api_application_status)
    app.router.add_post('/admin/api/application/delete', admin_api_application_delete)
    app.router.add_get('/admin/api/bootcamp/applications', admin_api_bootcamp_applications_get)
    app.router.add_post('/admin/api/bootcamp/application/status', admin_api_bootcamp_application_status)
    app.router.add_post('/admin/api/bootcamp/application/delete', admin_api_bootcamp_application_delete)
    app.router.add_get('/admin/api/news', admin_api_news_get)
    app.router.add_post('/admin/api/news/save', admin_api_news_save)
    app.router.add_post('/admin/api/news/delete', admin_api_news_delete)  # Public - sayt uchun
    app.router.add_get('/api/partners', api_get_partners)
    app.router.add_get('/admin/api/partners', admin_api_partners_get)
    app.router.add_post('/admin/api/partners/save', admin_api_partners_save)
    app.router.add_post('/admin/api/partners/delete', admin_api_partners_delete)
    app.router.add_get('/admin/api/site/config', admin_api_site_config_get)
    app.router.add_post('/admin/api/site/config', admin_api_site_config_save)
    app.router.add_post('/admin/api/upload/image', admin_api_upload_image)
    app.router.add_post('/api/upload/resume', api_upload_resume)
    if _os.path.isdir('resumes'):
        app.router.add_static('/resumes', path='resumes', name='resumes')
    app.router.add_post('/admin/api/group/create', admin_api_group_create)
    app.router.add_get('/admin/api/group/excel', admin_api_group_excel)
    app.router.add_post('/admin/api/student/add', admin_api_student_add)
    app.router.add_post('/admin/api/student/edit', admin_api_student_edit)
    app.router.add_post('/admin/api/student/delete', admin_api_student_delete)
    app.router.add_post('/admin/api/branch/delete', admin_api_branch_delete)
    app.router.add_post('/admin/api/branch/update', admin_api_branch_update)
    app.router.add_get('/teacher', miniapp_teacher_page)
    app.router.add_get('/teacher/api/profile/photo', miniapp_get_profile_photo)
    app.router.add_post('/teacher/api/profile/update', miniapp_update_profile)
    app.router.add_post('/teacher/api/lang/change', miniapp_change_lang)
    app.router.add_get('/teacher/api/data', miniapp_teacher_data)
    app.router.add_get('/teacher/api/group/students', miniapp_group_students)
    app.router.add_post('/teacher/api/payment/save', miniapp_save_payment)
    app.router.add_post('/teacher/api/student/add', miniapp_add_student)
    app.router.add_post('/teacher/api/student/delete', miniapp_del_student)
    app.router.add_get('/teacher/api/group/attendance', miniapp_group_att_history)
    app.router.add_get('/miniapp', miniapp_page)
    app.router.add_get('/miniapp/api/init', miniapp_api_init)
    app.router.add_get('/miniapp/api/students', miniapp_api_students)
    app.router.add_post('/miniapp/api/submit', miniapp_api_submit)
    app.router.add_post('/admin/api/salary/excel', admin_api_salary_excel)
    app.router.add_get('/admin/api/report/monthly', admin_api_monthly_excel)
    app.router.add_get('/admin/api/report/daily_pdf', admin_api_daily_pdf)
    app.router.add_get('/admin/api/branch/groups', admin_api_branch_groups)
    app.router.add_get('/admin/api/schedule/view', admin_api_schedule_view)
    
    # Hisobot API endpoint'lari
    app.router.add_get('/admin/api/reports/attendance', admin_api_reports_attendance)
    app.router.add_get('/admin/api/reports/groups', admin_api_reports_groups)
    app.router.add_get('/admin/api/reports/students', admin_api_reports_students)
    app.router.add_get('/admin/api/reports/payments', admin_api_reports_payments)
    app.router.add_get('/admin/api/reports/branches', admin_api_reports_branches)
    app.router.add_get('/admin/api/reports/general', admin_api_reports_general)
    
    # Business Report API (Korean style)
    app.router.add_get('/admin/api/business/report', admin_api_business_report)
    app.router.add_post('/admin/api/business/expenses', admin_api_business_expenses_save)
    
    # Salary Configs API (Korean style)
    app.router.add_get('/admin/api/salary/configs', admin_api_salary_configs_get)
    app.router.add_post('/admin/api/salary/configs', admin_api_salary_configs_save)

    port = int(os.environ.get("PORT", 10000))
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, '0.0.0.0', port)
    await site.start()
    logging.info(f"✅ Web server {port}-portda ishga tushdi")

    # Server ishlayversin
    try:
        await asyncio.Event().wait()
    finally:
        await bot.delete_webhook()
        await runner.cleanup()

if __name__ == "__main__":
    asyncio.run(main())
